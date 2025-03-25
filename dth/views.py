# dth/views.py
from django.contrib.auth.decorators import permission_required, login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.urls import reverse_lazy, reverse
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.generic import CreateView, TemplateView
from django.views.decorators.http import require_POST
from django.shortcuts import render, get_object_or_404, redirect

import json
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import timedelta, datetime, time, date
from dateutil.relativedelta import relativedelta

from got.models import Asset
from dth.forms import UserProfileForm
from .models import *
from .forms import *
from .utils import *


@login_required
def profile_update(request):
    user = request.user
    # Intenta obtener el perfil; si no existe, créalo
    profile, created = UserProfile.objects.get_or_create(user=user)
    
    if request.method == 'POST':
        form = UserProfileForm(request.POST, request.FILES, instance=profile, user=user)
        if form.is_valid():
            # Actualizar datos del usuario
            user.first_name = form.cleaned_data.get('first_name')
            user.last_name = form.cleaned_data.get('last_name')
            user.email = form.cleaned_data.get('email')
            user.save()
            # Guardar perfil
            form.save()
            messages.success(request, 'Tu perfil ha sido actualizado exitosamente.')
            return redirect('got:asset-list')
        else:
            messages.error(request, 'Por favor corrige los errores indicados.')
    else:
        form = UserProfileForm(instance=profile, user=user, initial={
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email,
        })
    return render(request, 'dth/profile_update.html', {'form': form})


class OvertimeListView(LoginRequiredMixin, TemplateView):
    template_name = 'dth/overtime_list.html'

    def get_default_date_range(self):
        today = date.today()
        if today.day < 21:
            start_date = (today - relativedelta(months=1)).replace(day=21)
            end_date = today.replace(day=21)
        else:
            start_date = today.replace(day=21)
            end_date = (today + relativedelta(months=1)).replace(day=21)
        return start_date, end_date
    
    def parse_date_range(self):
        date_range_str = self.request.GET.get('date_range', '')
        if date_range_str:
            # Detecta separadores posibles
            if " to " in date_range_str:
                dates = date_range_str.split(" to ")
            elif " a " in date_range_str:
                dates = date_range_str.split(" a ")
            elif "," in date_range_str:
                dates = date_range_str.split(",")
            else:
                dates = []
            if len(dates) == 2:
                try:
                    start_date = datetime.strptime(dates[0].strip(), '%Y-%m-%d').date()
                    end_date = datetime.strptime(dates[1].strip(), '%Y-%m-%d').date()
                except ValueError:
                    start_date, end_date = self.get_default_date_range()
            else:
                try:
                    start_date = datetime.strptime(date_range_str.strip(), '%Y-%m-%d').date()
                    end_date = start_date + timedelta(days=1)
                except ValueError:
                    start_date, end_date = self.get_default_date_range()
        else:
            start_date, end_date = self.get_default_date_range()
        return start_date, end_date
    
    def get_queryset(self):
        start_date, end_date = self.parse_date_range()

        queryset = OvertimeProject.objects.filter(report_date__gte=start_date, report_date__lt=end_date)

        # Filtros por grupo de usuario
        current_user = self.request.user
        if current_user.groups.filter(name='mto_members').exists():
            pass  # Sin filtro adicional
        elif current_user.groups.filter(name='maq_members').exists():
            asset = Asset.objects.filter(models.Q(supervisor=current_user) | models.Q(capitan=current_user)).first()
            queryset = queryset.filter(asset=asset)
        elif current_user.groups.filter(name__in=['buzos_members', 'serport_members']).exists():
            queryset = queryset.none()

        return queryset.order_by('-report_date', 'asset', 'ovetime_set_start')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['projects'] = self.get_queryset()
        context['start_date'], context['end_date'] = self.parse_date_range()
        return context

    # def get_queryset(self):
    #     start_date, end_date = self.parse_date_range()

        # Filtros específicos
        # person_name = self.request.GET.get('name', '').strip()
        # cedula = self.request.GET.get('cedula', '').strip()
        # asset_name = self.request.GET.get('asset', '').strip()
        # aprobado_filter = self.request.GET.get('estado', 'all')

        # if person_name:
        #     queryset = queryset.filter(nombre_completo__icontains=person_name)
        # if cedula:
        #     queryset = queryset.filter(cedula__icontains=cedula)
        # if asset_name:
        #     queryset = queryset.filter(asset__name__icontains=asset_name)
        # if aprobado_filter == 'aprobado':
        #     queryset = queryset.filter(approved=True)
        # elif aprobado_filter == 'no_aprobado':
        #     queryset = queryset.filter(approved=False)

    def get_total_hours(self, queryset):
        total_seconds = 0
        for entry in queryset:
            dt_start = datetime.combine(entry.fecha, entry.hora_inicio)
            dt_end = datetime.combine(entry.fecha, entry.hora_fin)
            diff = dt_end - dt_start
            total_seconds += diff.total_seconds()
        return round(total_seconds / 3600, 2)  # Total en horas

    def get_total_sunday_holiday_hours(self, queryset):
        total_seconds = 0
        for entry in queryset:
            dt_start = datetime.combine(entry.fecha, entry.hora_inicio)
            dt_end = datetime.combine(entry.fecha, entry.hora_fin)
            diff = dt_end - dt_start
            # Si la fecha es domingo o festiva, sumar la totalidad de la duración
            if entry.fecha.weekday() == 6 or entry.fecha in COLOMBIA_HOLIDAYS:
                total_seconds += diff.total_seconds()
        return round(total_seconds / 3600, 2)

    # def get_total_nocturnal_hours(self, queryset):
    #     total_seconds = 0
    #     for entry in queryset:
    #         dt_start = datetime.combine(entry.fecha, entry.hora_inicio)
    #         dt_end = datetime.combine(entry.fecha, entry.hora_fin)
    #         if dt_end <= dt_start:
    #             dt_end += timedelta(days=1)
    #         overlap = get_nocturnal_overlap(dt_start, dt_end)
    #         total_seconds += overlap.total_seconds()
    #     return round(total_seconds / 3600, 2)


    # def get_context_data(self, **kwargs):
    #     context = super().get_context_data(**kwargs)
    #     start_date, end_date = self.parse_date_range()
    #     context['start_date'] = start_date.strftime('%Y-%m-%d')
    #     context['end_date'] = end_date.strftime('%Y-%m-%d')
    #     context['date_range'] = f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    #     context['name'] = self.request.GET.get('name', '')
    #     context['cedula'] = self.request.GET.get('cedula', '')
    #     context['asset_filter'] = self.request.GET.get('asset', '')
    #     context['estado'] = self.request.GET.get('estado', 'all')
        
    #     queryset = OvertimeProject.objects.filter(report_date__gte=start_date, report_date__lt=end_date)
    #     total_hours = self.get_total_hours(queryset)
    #     context['total_hours'] = total_hours  # Valor numérico, por ejemplo 12.5
    #     context['total_hours_hhmm'] = hours_to_hhmm(total_hours)  # Formato HH:MM, por ejemplo "12:30"
        
    #     context['total_sunday_holiday_hours'] = hours_to_hhmm(self.get_total_sunday_holiday_hours(queryset))
    #     # context['total_nocturnal_hours'] = self.get_total_nocturnal_hours(queryset)
    
    #     page_obj = context['page_obj']
    #     date_asset_justification_groups = []
    #     for fecha, fecha_entries in groupby(page_obj.object_list, key=attrgetter('fecha')):
    #         asset_groups = []
    #         fecha_entries = list(fecha_entries)
    #         for asset, asset_entries in groupby(fecha_entries, key=attrgetter('asset')):
    #             asset_entries = list(asset_entries)
    #             for justificacion, justificacion_entries in groupby(asset_entries, key=attrgetter('justificacion')):
    #                 justificacion_entries = list(justificacion_entries)
    #                 asset_groups.append({
    #                     'asset': asset,
    #                     'justificacion': justificacion,
    #                     'entries': justificacion_entries
    #                 })
    #         date_asset_justification_groups.append({'fecha': fecha, 'assets': asset_groups})
    #     context['date_asset_groups'] = date_asset_justification_groups
    #     # También enviamos la lista de activos para el select (assets de área 'a')
    #     context['assets'] = Asset.objects.filter(area='a', show=True).order_by('name')
    #     context['holiday_dates'] = json.dumps([d.strftime('%Y-%m-%d') for d in colombia_holidays.keys()])
    #     return context

@login_required
@permission_required('got.can_approve_overtime', raise_exception=True)
@require_POST
def approve_overtime(request):
    action = request.POST.get('action')
    remarks = request.POST.get('remarks', '')
    selected_overtime_ids = request.POST.getlist('selected_overtime')

    overtimes = Overtime.objects.filter(id__in=selected_overtime_ids)

    if action == 'approve':
        overtimes_updated_count = overtimes.update(state='a')
        messages.success(request, f"Se han aprobado {overtimes_updated_count} registros de horas extras correctamente.")
        
    elif action == 'reject':
        overtimes_updated_count = overtimes.update(state='b', remarks=remarks)
        messages.success(request, f"Se han rechazado {overtimes_updated_count} registros de horas extras correctamente.")

    next_url = request.POST.get('next', reverse('dth:overtime_list'))
    return redirect(next_url)


@login_required
@require_POST
def edit_overtime(request):
    overtime = get_object_or_404(Overtime, id=request.POST.get('overtime_id'))
    overtime.start = request.POST.get('start')
    overtime.end = request.POST.get('end')
    overtime.state = request.POST.get('state')
    overtime.remarks = request.POST.get('remarks', '')
    if overtime.start >= overtime.end:
        messages.error(request, "La hora de inicio no puede ser posterior a la hora final.")
    else:
        overtime.save()
        messages.success(request, "Horas extras actualizadas.")
    return redirect(request.POST.get('next', '/'))

@login_required
@require_POST
def delete_overtime(request):
    overtime = get_object_or_404(Overtime, id=request.POST.get('overtime_id'))
    overtime.delete()
    messages.success(request, "Registro eliminado correctamente.")
    return redirect(request.POST.get('next', '/'))


def get_default_date_range():
    today = date.today()
    if today.day < 21:
        start_date = (today - relativedelta(months=1)).replace(day=21)
        end_date = today.replace(day=21)
    else:
        start_date = today.replace(day=21)
        end_date = (today + relativedelta(months=1)).replace(day=21)
    return start_date, end_date

def parse_date_range(request):
    date_range_str = request.GET.get('date_range', '')
    if date_range_str:
        # Detecta separadores posibles
        if " to " in date_range_str:
            dates = date_range_str.split(" to ")
        elif " a " in date_range_str:
            dates = date_range_str.split(" a ")
        elif "," in date_range_str:
            dates = date_range_str.split(",")
        else:
            dates = []
        if len(dates) == 2:
            try:
                start_date = datetime.strptime(dates[0].strip(), '%Y-%m-%d').date()
                end_date = datetime.strptime(dates[1].strip(), '%Y-%m-%d').date()
            except ValueError:
                start_date, end_date = get_default_date_range()
        else:
            try:
                start_date = datetime.strptime(date_range_str.strip(), '%Y-%m-%d').date()
                end_date = start_date + timedelta(days=1)
            except ValueError:
                start_date, end_date = get_default_date_range()
    else:
        start_date, end_date = get_default_date_range()
    return start_date, end_date

def filter_overtime_queryset(request):
    # Rango de fechas
    start_date, end_date = parse_date_range(request)
    qs = Overtime.objects.filter(fecha__gte=start_date, fecha__lt=end_date)

    # Filtros por nombre, cédula y asset
    person_name = request.GET.get('name', '').strip()
    cedula = request.GET.get('cedula', '').strip()
    asset_name = request.GET.get('asset', '').strip()
    estado = request.GET.get('estado', 'all')

    if person_name:
        qs = qs.filter(nombre_completo__icontains=person_name)
    if cedula:
        qs = qs.filter(cedula__icontains=cedula)
    if asset_name:
        qs = qs.filter(asset__name__icontains=asset_name)
    if estado == 'approved':
        qs = qs.filter(approved=True)
    elif estado == 'no_aprobado':
        qs = qs.filter(approved=False)

    # Filtros según grupo de usuario
    current_user = request.user
    if current_user.groups.filter(name='mto_members').exists():
        pass  # Sin filtro adicional
    elif current_user.groups.filter(name='maq_members').exists():
        asset = Asset.objects.filter(models.Q(supervisor=current_user) | models.Q(capitan=current_user)).first()
        qs = qs.filter(asset=asset)
    elif current_user.groups.filter(name__in=['buzos_members', 'serport_members']).exists():
        qs = qs.none()

    return qs.order_by('-fecha', 'asset', 'hora_inicio', 'justificacion')

@login_required
def export_overtime_excel(request):
    qs = filter_overtime_queryset(request)

    wb = openpyxl.Workbook()
    ws = wb.active

    today = date.today()
    last_month = today - relativedelta(months=1)
    report_title = f"REPORTE 24 {last_month.strftime('%B')} hasta 24 {today.strftime('%B')} {today.year}"
    ws.append([report_title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(row=1, column=1).font = openpyxl.styles.Font(size=14, bold=True)

    headers = ['Nombre completo', 'Cédula', 'Fecha', 'Justificación', 'Hora inicio', 'Hora fin', 'Aprueba', 'Transporte']
    ws.append(headers)

    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    # Añadir los datos
    for entry in qs:
        # Calcular transporte (puedes ajustar esta lógica según lo necesites)
        transporte = 'Sí' if entry.hora_fin >= time(19, 0) else 'No'

        row = [
            entry.nombre_completo,
            entry.cedula,
            entry.fecha.strftime('%d/%m/%Y') if entry.fecha else '',
            entry.justificacion,
            entry.hora_inicio.strftime('%I:%M %p') if entry.hora_inicio else '',
            entry.hora_fin.strftime('%I:%M %p') if entry.hora_fin else '',
            'Sí' if entry.approved else 'No',
            transporte,
        ]
        ws.append(row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Reporte_Horas_Extras_{today.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response


@login_required
def buscar_nomina(request):
    cedula = request.GET.get('cedula')
    try:
        persona = Nomina.objects.get(id_number=cedula)
        return JsonResponse({'success': True, 'name': f"{persona.name} {persona.surname}", 'position': persona.position})
    except Nomina.DoesNotExist:
        return JsonResponse({'success': False})


class OvertimeProjectCreateView(LoginRequiredMixin, CreateView):
    model = OvertimeProject
    form_class = OvertimeProjectForm
    template_name = "dth/create_overtime_report.html"
    success_url = reverse_lazy("dth:overtime_list")  # Ajusta con tu URL de éxito

    def get_form_kwargs(self):
        """
        Inyectar el 'user' al formulario para que 
        OvertimeProjectForm pueda ocultar/mostrar el asset si es maq_members.
        """
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        """
        Añadimos:
          - La lista de festivos en 'holiday_dates'.
          - La fecha inicial que usará el calendario en 'initial_date'.
        """
        context = super().get_context_data(**kwargs)

        # 1) Festivos
        current_year = date.today().year
        next_year = current_year + 1
        holiday_list = []
        for hol_date in COLOMBIA_HOLIDAYS:
            if hol_date.year in [current_year, next_year]:
                holiday_list.append(hol_date.isoformat())
        context["holiday_dates"] = json.dumps(holiday_list)

        # 2) Determinar la fecha inicial (si ya llenaron el form o no).
        #    Si el form tiene un .instance con .report_date, la usamos.
        #    Si no, usamos la fecha de hoy.
        chosen_date = None
        form = context.get("form")
        if form and form.instance and form.instance.report_date:
            chosen_date = form.instance.report_date
        if not chosen_date:
            chosen_date = date.today()  # por defecto, hoy

        context["initial_date"] = chosen_date

        return context

    def form_valid(self, form):
        """
        Lógica principal para:
          - Guardar el proyecto (commit=False primero)
          - Validar horas extras, cédulas, etc.
          - Crear los sub-intervalos (Overtime)
        """
        project = form.save(commit=False)
        project.reported_by = self.request.user

        # Asignar asset si user es maq_members
        if self.request.user.groups.filter(name='maq_members').exists():
            asset = Asset.objects.filter(
                Q(supervisor=self.request.user) | Q(capitan=self.request.user)
            ).first()
            project.asset = asset

        # Datos del formulario
        report_date = form.cleaned_data['report_date']
        start_time  = form.cleaned_data['start']
        end_time    = form.cleaned_data['end']

        # Calcular intervalos válidos
        overtime_periods = calcular_horas_extras(report_date, start_time, end_time)
        if not overtime_periods:
            messages.error(self.request, "Las horas ingresadas no califican como horas extras.")
            return self.form_invalid(form)  # Re-renderiza con errores

        # Leer cédulas / externos
        cedulas_json  = form.cleaned_data.get('cedulas', '[]')
        externos_json = form.cleaned_data.get('personas_externas', '[]')

        try:
            cedulas_list = json.loads(cedulas_json) if cedulas_json.strip() else []
        except:
            cedulas_list = []
        try:
            externos_list = json.loads(externos_json) if externos_json.strip() else []
        except:
            externos_list = []

        # Validar: al menos 1 persona
        if not cedulas_list and not externos_list:
            messages.error(self.request, "Debes ingresar al menos una persona (cédula o registro manual).")
            return self.form_invalid(form)

        # Guardar el proyecto
        project.save()

        # Función local para crear los sub-intervalos no solapados
        def create_intervals_for_overtime(worker_obj, ext_nombre, ext_cedula, ext_cargo):
            created_count = 0
            if worker_obj:
                existing = Overtime.objects.filter(
                    worker=worker_obj,
                    project__report_date=report_date
                ).order_by('start')
            else:
                existing = Overtime.objects.filter(
                    cedula=ext_cedula,
                    project__report_date=report_date
                ).order_by('start')

            existing_ranges = [(ov.start, ov.end) for ov in existing]

            for (ot_start, ot_end) in overtime_periods:
                final_subintervals = subtract_time_ranges(ot_start, ot_end, existing_ranges)
                for (sub_start, sub_end) in final_subintervals:
                    Overtime.objects.create(
                        start=sub_start,
                        end=sub_end,
                        worker=worker_obj,
                        project=project,
                        state='c',  # 'Pendiente'
                        nombre_completo=ext_nombre or '',
                        cedula=ext_cedula or '',
                        cargo=ext_cargo or ''
                    )
                    created_count += 1
                    existing_ranges.append((sub_start, sub_end))
                    existing_ranges.sort(key=lambda x: x[0])
            return created_count

        total_created = 0
        total_omitted = 0

        # Crear Overtime para Nomina
        for cedula in cedulas_list:
            try:
                persona = Nomina.objects.get(id_number=cedula)
            except Nomina.DoesNotExist:
                total_omitted += 1
                continue
            created = create_intervals_for_overtime(
                worker_obj=persona,
                ext_nombre=None,
                ext_cedula=None,
                ext_cargo=None
            )
            if created == 0:
                total_omitted += 1
            else:
                total_created += created

        # Crear Overtime para externos
        for externo_data in externos_list:
            ext_ced   = externo_data.get('cedula', '')
            ext_name  = externo_data.get('nombre_completo', '')
            ext_cargo = externo_data.get('cargo', '')
            created = create_intervals_for_overtime(
                worker_obj=None,
                ext_nombre=ext_name,
                ext_cedula=ext_ced,
                ext_cargo=ext_cargo
            )
            if created == 0:
                total_omitted += 1
            else:
                total_created += created

        # Mensajes finales
        if total_created > 0:
            messages.success(
                self.request,
                f"Reporte creado. Se generaron {total_created} registros de horas extras. Omitidos: {total_omitted}."
            )
            return super().form_valid(form)
        else:
            project.delete()
            messages.error(
                self.request,
                "No se pudo crear ninguna hora extra (todos los intervalos estaban solapados)."
            )
            return self.form_invalid(form)

    def form_invalid(self, form):
        """
        Si el form no es válido (o detectamos un error manual en form_valid),
        se re-renderiza la plantilla con los mismos datos y los mensajes de error.
        """
        return super().form_invalid(form)


def gerencia_nomina_view(request):
    """
    Vista exclusiva para la gerencia (Jennifer Padilla).
    Permite subir un Excel y procesar la creación/actualización de NominaReport.
    """
    template_name = "dth/gerencia_nomina.html"

    if request.method == "POST":
        form = UploadNominaReportForm(request.POST, request.FILES)
        if form.is_valid():
            # Se extrae el archivo excel
            excel_file = request.FILES['excel_file']

            # Deshabilitar el botón y mostrar mensaje de "procesando..."
            # -> Esto se maneja desde el template con JavaScript, ver ejemplo abajo.

            # Lógica para procesar el archivo
            try:
                # Abrimos el workbook con openpyxl
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                # Asumimos que solo hay UNA hoja
                sheet = wb.active

                # Mapeo de codigos -> nombre de campo en NominaReport
                # Clave = el prefijo con el que inicia la celda de "Concepto"
                CODE_FIELD_MAP = {
                    'DV01': 'dv01',
                    'DV23': 'dv25',
                    'DV03': 'dv03',
                    'DV103': 'dv103',
                    'DV27': 'dv27',
                    'DV30': 'dv30',
                    'DX03': 'dx03',
                    'DX05': 'dx05',
                    'DX01': 'dx01',
                    'DX07': 'dx07',
                    'DX12': 'dx12',
                    'DX63': 'dx63',
                    'DX64': 'dx64',
                    'DX66': 'dx66',
                    # Ejemplo: si tienes más códigos, agrégalos aquí.
                    # 'DV48': 'dv48', etc.
                }

                rows_created = 0
                rows_updated = 0

                # Buscar el índice de columna (asumiendo encabezados en la primera fila)
                # Para evitar problemas con la posición exacta, lo hacemos flexible:
                headers = {cell.value: idx for idx, cell in enumerate(sheet[1], start=1)}
                print("DEBUG - Headers leídos:", headers)

                # Revisar que existan las columnas mínimas
                required_cols = [
                    "Empleado", "FechaMovimiento", "Concepto",
                    "ValorDevengo", "ValorDeduccion"
                ]
                clean_headers = {str(k).strip(): v for k, v in headers.items() if k is not None}

                for col in required_cols:
                    if col not in clean_headers:
                        messages.error(request, f"Falta la columna requerida '{col}' en el Excel.")
                        return render(request, template_name, {"form": form})

                # Iterar sobre las filas (omitir la primera, que es encabezado)
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # Tomar valores de la fila según las columnas
                    empleado = row[headers["Empleado"] - 1]
                    fecha_mov = row[headers["FechaMovimiento"] - 1]  # string con fecha

                    fecha_celda = row[headers["FechaMovimiento"] - 1]
                    if isinstance(fecha_celda, datetime):
                        # Es un objeto datetime => tomar su fecha
                        parsed_date = fecha_celda.date()
                        mes = parsed_date.month
                        anio = parsed_date.year
                    elif isinstance(fecha_celda, date):
                        # A veces openpyxl puede devolver date directamente (sin tiempo)
                        mes = fecha_celda.month
                        anio = fecha_celda.year
                    elif isinstance(fecha_celda, str):
                        # Intentar parsear como "d/m/yyyy"
                        # Ojo con más validaciones según tu archivo
                        try:
                            parsed_date = datetime.strptime(fecha_celda.strip(), "%d/%m/%Y")
                            mes = parsed_date.month
                            anio = parsed_date.year
                        except ValueError:
                            # Manejar formato inválido
                            continue
                    else:
                        # No es ninguno de los tipos esperados => se ignora
                        continue

                    concepto_original = row[clean_headers["Concepto"] - 1] or ""
                    concepto_upper = str(concepto_original).strip().upper()

                    print("DEBUG - Concepto leido:", concepto_original) 

                    # Settear found_field en None
                    found_field = None
                    for code_prefix, field_name in CODE_FIELD_MAP.items():
                        # Buscamos si concepto_upper arranca con code_prefix
                        if concepto_upper.startswith(code_prefix):
                            found_field = field_name
                            break

                    if found_field:
                        print("DEBUG - matched", found_field, "para concepto", concepto_upper)
                    else:
                        print("DEBUG - sin match para", concepto_upper)

                    valor_devengo = row[headers["ValorDevengo"] - 1] or 0
                    valor_deduccion = row[headers["ValorDeduccion"] - 1] or 0

                    # Verificamos si hay en Nomina
                    if empleado is None:
                        continue  # Ignorar filas vacías
                    try:
                        nomina_obj = Nomina.objects.get(id_number=str(empleado))
                    except Nomina.DoesNotExist:
                        # Si no existe en Nomina, puedes decidir crearlo o ignorarlo
                        # Aquí optamos por ignorar y continuar
                        continue

                    # Extraer mes y año de 'FechaMovimiento' (ej: "1/02/2025")
                    # Asumiendo formato: d/m/yyyy
                    # try:
                    #     parsed_date = datetime.strptime(str(fecha_mov), "%d/%m/%Y")
                    #     mes = parsed_date.month
                    #     anio = parsed_date.year
                    # except ValueError:
                    #     # Si no logra parsear, omitir
                    #     continue

                    # Verificar si ya existe un NominaReport con esa Nomina y mes/año
                    with transaction.atomic():
                        nomina_report, created = NominaReport.objects.get_or_create(
                            nomina=nomina_obj,
                            mes=mes,
                            anio=anio
                        )
                        if created:
                            rows_created += 1
                        else:
                            rows_updated += 1

                        # Detectar si el Concepto inicia con algún código
                        # p.ej. "DV01 SUELDO BASICO" -> arranca con "DV01"
                        found_field = None
                        concepto_upper = concepto_upper.strip().upper()
                        for code_prefix, field_name in CODE_FIELD_MAP.items():
                            if concepto_upper.startswith(code_prefix):
                                found_field = field_name
                                break

                        # Tomar el valor "no cero" entre ValorDevengo y ValorDeduccion
                        # Si ambos son 0, no hacemos nada
                        valor = Decimal('0')
                        if valor_devengo not in [None, 0]:
                            valor = Decimal(valor_devengo)
                        elif valor_deduccion not in [None, 0]:
                            valor = Decimal(valor_deduccion)

                        if found_field and valor != 0:
                            # Asignar dinámicamente al campo correspondiente
                            setattr(nomina_report, found_field, valor)
                            nomina_report.save()

                # Al terminar
                messages.success(
                    request,
                    (f"Proceso finalizado. Se crearon {rows_created} registros nuevos "
                     f"y se actualizaron {rows_updated} registros existentes.")
                )
                return redirect('dth:gerencia_nomina')  # o simplemente recargar
            except Exception as e:
                messages.error(request, f"Ocurrió un error procesando el archivo: {str(e)}")
    else:
        form = UploadNominaReportForm()

    # ==== CONSULTAMOS LOS REGISTROS PARA LA TABLA ====
    # Recuperamos todos los NominaReport, con su Nomina relacionada
    # Orden: primero por el nombre del empleado, luego mes y año.
    # Nota: Para “nombre completo”, un truco es concatenar en un campo simulado
    #       o podemos sólo ordenar por “nomina__name” y “nomina__surname”.
    reports = (NominaReport.objects
               .select_related('nomina')
               .order_by('nomina__name', 'nomina__surname', 'mes', 'anio'))

    # Aquí hacemos los cálculos necesarios claramente:
    processed_reports = []
    for nr in reports:
        salario = nr.nomina.salary or 0
        dv01 = nr.dv01 or 0
        dv25 = nr.dv25 or 0
        dias_sueldo_basico = round((dv01 / (salario / 30)), 2) if salario else 0
        dias_vacaciones = round((dv25 / (salario / 30)), 2) if salario else 0

        processed_reports.append({
            'report': nr,
            'dias_sueldo_basico': dias_sueldo_basico,
            'dias_vacaciones': dias_vacaciones,
            # Aquí puedes añadir otros cálculos que necesites
        })

    nomina_qs = Nomina.objects.all().order_by('name', 'surname')

    context = {
        "form": form,
        "processed_reports": processed_reports,
        'nomina_list': nomina_qs
    }

    return render(request, template_name, context)


# dth/views.py (continuación)

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, NamedStyle
from decimal import Decimal
from django.http import HttpResponse

@login_required
def export_gerencia_nomina_excel(request):
    """
    Exporta la tabla de NominaReport a un archivo Excel con estilos y formatos 
    similares a la tabla de la vista de Nómina Gerencia.
    """
    # 1) Obtención de registros
    reports = (NominaReport.objects
               .select_related('nomina')
               .order_by('nomina__name', 'nomina__surname', 'mes', 'anio'))
    
    # 2) Crear workbook y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nómina Gerencia"
    
    # 3) Definir encabezados y configurar estilos
    headers = [
        "Cédula", "Nombre", "Cargo", "MES", "AÑO", "Salario",
        "# días sueldo básico", "Sueldo Básico (dv01)", "# días vacaciones", "Pago Vacaciones (dv25)",
        "Subsidio transporte (dv03)", "Licencia familia (dv103)", "Provisión vacaciones (4.17%)",
        "DV27 Intereses Ces. Ant", "Prima de Servicio (8.33%)", "Cesantías (dv30)",
        "Intereses de Cesantías (1%)", "Salud colab. (4%)", "Pensión colab (dx03)",
        "Fdo. solidar. (dx05)", "Pensión empleador (12%)", "ARL empleador (6.96%)",
        "Caja comp. (4%)", "Ret. Fuente", "Exequias Lordoy (dx07)",
        "Desc. pensión voluntaria (dx12)", "Banco Occidente (dx63)",
        "Confenalco (dx64)", "Préstamo (dx66)", "Neto a pagar"
    ]
    
    # Estilo para encabezados
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(border_style="thin", color="000000")
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # Escribir encabezados y ajustar anchos
    for col_num, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

    # Número de columna de cada campo de moneda (según índice en la lista de encabezados)
    currency_columns = [6, 8, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20,
                        21, 22, 23, 24, 25, 26, 27, 28, 29, 30]
    currency_format = '"COP " #,##0.00'

    # 4) Agregar filas de datos
    row_index = 2
    for nr in reports:
        emp = nr.nomina
        # Calcular # días sueldo básico y # días vacaciones
        dias_sb = Decimal('0.00')
        dias_vac = Decimal('0.00')
        if emp.salary and nr.dv01:
            dias_sb = nr.dv01 / (Decimal(emp.salary) / Decimal('30'))
        if emp.salary and nr.dv25:
            dias_vac = nr.dv25 / (Decimal(emp.salary) / Decimal('30'))
        
        data_row = [
            emp.id_number,
            f"{emp.name} {emp.surname}",
            emp.position,
            nr.mes,
            nr.anio,
            float(emp.salary or 0),
            float(dias_sb),
            float(nr.dv01 or 0),
            float(dias_vac),
            float(nr.dv25 or 0),
            float(nr.dv03 or 0),
            float(nr.dv103 or 0),
            float(nr.provision_vacaciones or 0),  # dv25 * 4.17%
            float(nr.dv27 or 0),
            float(nr.prima_servicio or 0),       # (dv01 + dv25 + dv03)*0.0833
            float(nr.dv30 or 0),
            float(nr.intereses_cesantias or 0),  # dv30 * 0.01
            float(nr.salud_aporte or 0),         # (dv01 + dv25) * 0.04
            float(nr.dx03 or 0),
            float(nr.dx05 or 0),
            float(nr.pension_aporte_empleador or 0),
            float(nr.arl_aporte or 0),
            float(nr.caja_compensacion_aporte or 0),
            float(nr.dx01 or 0),
            float(nr.dx07 or 0),
            float(nr.dx12 or 0),
            float(nr.dx63 or 0),
            float(nr.dx64 or 0),
            float(nr.dx66 or 0),
            float(nr.neto_a_pagar or 0),
        ]
        # Escribir la fila en la hoja
        for col_num, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row_index, column=col_num, value=value)
            cell.border = header_border
            # Alinear a la derecha para números (excepto cédula y texto)
            if col_num in currency_columns or col_num in [7, 9]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            # Si es columna de moneda, aplica el formato
            if col_num in currency_columns:
                cell.number_format = currency_format
        row_index += 1

    # 5) Retornar el archivo Excel como respuesta
    filename = "NominaGerencia.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
def nomina_edit(request, pk):
    """
    Vista que edita o elimina un registro de Nomina.

    - Si POST incluye 'action=delete', se elimina el registro.
      -> Si es AJAX, se responde con JSON.
      -> Si no es AJAX, se redirige con un mensaje.
    - Si no incluye 'action=delete', se asume edición de campos.
    """
    nomina_obj = get_object_or_404(Nomina, pk=pk)

    if request.method == 'POST':
        # Verificamos si es petición de borrado (action=delete)
        action = request.POST.get('action')
        if action == 'delete':
            nomina_obj.delete()

            # Si es AJAX, devolvemos JSON para que JS manipule la tabla sin recargar
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'deleted': True})

            # No es AJAX => redirigimos con mensaje
            messages.success(request, "Empleado eliminado con éxito.")
            return redirect('dth:gerencia_nomina')

        # De lo contrario, es edición
        form = NominaForm(request.POST, instance=nomina_obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Registro de nómina actualizado.")
            return redirect('dth:gerencia_nomina')
        else:
            messages.error(request, "Por favor revisa los campos del formulario.")
    else:
        # GET => mostrar el formulario (opcional si deseas cargar un template)
        form = NominaForm(instance=nomina_obj)

    # Podrías renderizar un template si deseas un formulario de edición fuera del modal
    return render(request, 'dth/nomina_form.html', {'form': form, 'object': nomina_obj})


# dth/views.py (complemento)
@login_required
def nomina_create(request):
    """
    Crea un nuevo registro de Nomina.
    """
    if request.method == 'POST':
        form = NominaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Empleado creado con éxito.')
            return redirect('dth:gerencia_nomina') # O la misma página modal
        else:
            messages.error(request, 'Hay errores en el formulario.')
    else:
        form = NominaForm()

    return render(request, 'dth/nomina_form.html', {'form': form})
