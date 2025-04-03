# dth/management/commands/import_positions.py
import openpyxl
import difflib
from django.core.management.base import BaseCommand
from dth.models import Position, Document, PositionDocument


class Command(BaseCommand):
    help = "Importa cargos y documentos desde un archivo Excel"

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, help='Ruta del archivo Excel')
    
    def handle(self, *args, **options):
        excel_path = options['file']
        if not excel_path:
            self.stdout.write(self.style.ERROR("Debes especificar --file=/ruta/del/archivo.xlsx"))
            return
        
        try:
            wb = openpyxl.load_workbook(excel_path)
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error al abrir {excel_path}: {e}"))
            return

        sheet = wb.active
        # Lee la primera fila (encabezado) como lista de valores
        try:
            header_row = next(sheet.iter_rows(values_only=True))
        except StopIteration:
            self.stdout.write(self.style.ERROR("La hoja no contiene filas."))
            return
        
        header_row = list(header_row) if header_row else []

        # Diccionario para no preguntar Document repetidamente
        # doc_header_cache: dict[str, Document]
        # Ej: {"OMI 1.13 Primeros Auxilios": doc_obj}
        self.doc_header_cache = {}

        # Ubicamos columnas (ajusta a tus nombres reales)
        col_cargo = None
        col_adtop = None
        col_titulonaveg = None
        col_regla = None
        col_capacidad = None

        for idx, col_name in enumerate(header_row):
            if not col_name:
                continue
            lower_name = str(col_name).strip().lower()
            if lower_name.startswith("cargo"):
                col_cargo = idx
            elif "operativo" in lower_name or "adttvo" in lower_name:
                col_adtop = idx
            elif "título de naveg" in lower_name or "titulo de naveg" in lower_name:
                col_titulonaveg = idx
            elif "regla" in lower_name:
                col_regla = idx
            elif "capacidad" in lower_name:
                col_capacidad = idx
        
        # Las columnas de documentos (lo que no es cargo/adtop/regla/capacidad)
        doc_columns = []
        for idx, col_name in enumerate(header_row):
            if idx not in [col_cargo, col_adtop, col_titulonaveg, col_regla, col_capacidad] and col_name:
                doc_columns.append((idx, str(col_name).strip()))
        
        # Recorremos filas restantes
        all_rows = list(sheet.iter_rows(min_row=2, values_only=True))
        for row_index, row_data in enumerate(all_rows, start=2):
            if not row_data:
                continue
            
            cargo_name = row_data[col_cargo] if col_cargo is not None else None
            if not cargo_name:
                continue
            
            cargo_name = str(cargo_name).strip()
            if not cargo_name or cargo_name.upper() == "N/A":
                # Si es N/A, ignorar
                continue

            # -- Buscar coincidencias en DB
            matches = Position.objects.filter(name__iexact=cargo_name)

            if matches.count() > 1:
                self.stdout.write(self.style.WARNING(
                    f"Fila {row_index}: Cargo '{cargo_name}' aparece varias veces en la DB."
                ))
                # Preguntamos hasta que escoja opción válida
                while True:
                    action = input("¿[i]gnorar, [n]uevo cargo o [m]anual ID? (i/n/m): ").strip().lower()
                    if action in ['i','n','m']:
                        break
                    print("Opción inválida. Ingresa i, n o m.")
                
                if action == 'i':
                    self.stdout.write("Ignorando fila...")
                    continue
                elif action == 'n':
                    cargo_obj = self.create_position_interactive(
                        cargo_name, row_data, col_adtop, col_titulonaveg, col_regla
                    )
                else:  # action == 'm'
                    cargo_obj = self.find_position_by_id_interactive()

            elif matches.count() == 1:
                # Ya existe uno
                cargo_obj = matches.first()
                cargo_obj = self.update_position_interactive(
                    cargo_obj, row_data, col_adtop, col_titulonaveg, col_regla
                )
            else:
                # No se encontró
                self.stdout.write(
                    self.style.WARNING(f"Fila {row_index}: Cargo '{cargo_name}' no existe en DB.")
                )
                while True:
                    action = input("¿[c]rear nuevo, [i]gnorar fila, [m]anual ID? (c/i/m): ").strip().lower()
                    if action in ['c','i','m']:
                        break
                    print("Opción inválida. Ingresa c, i o m.")
                
                if action == 'c':
                    cargo_obj = self.create_position_interactive(
                        cargo_name, row_data, col_adtop, col_titulonaveg, col_regla
                    )
                elif action == 'm':
                    cargo_obj = self.find_position_by_id_interactive()
                    if not cargo_obj:
                        self.stdout.write("No se asignó cargo, ignorando fila.")
                        continue
                else:
                    self.stdout.write("Ignorando fila...")
                    continue
            
            if not cargo_obj:
                continue  # No se pudo obtener cargo

            # 3) Procesar columns => Documentos
            for (col_idx, doc_header) in doc_columns:
                cell_value = row_data[col_idx]
                if cell_value is None:
                    # interpretamos como "sí aplica"
                    self.link_document_to_position(doc_header, cargo_obj)
                else:
                    cell_str = str(cell_value).strip()
                    if cell_str.upper() == "N/A":
                        # no aplica => ignoramos
                        pass
                    else:
                        # interpretamos como aplica => link doc
                        self.link_document_to_position(doc_header, cargo_obj)

    # ------------------------------------------------------------------
    # Funciones auxiliares
    # ------------------------------------------------------------------

    def create_position_interactive(self, cargo_name, row_data, col_adtop, col_titulonaveg, col_regla):
        """
        Crea un nuevo Position.
        - No se puede crear con name vacío: si el usuario ingresa '' => se usa cargo_name.
        - "N/A" se trata como vacío en description/rule.
        """
        self.stdout.write(self.style.WARNING(f"Creando cargo '{cargo_name}'..."))

        # Preguntar una descripción
        desc_input = input("Descripción del cargo (ENTER para usar 'TITULO DE NAVEG'): ").strip()
        if desc_input.upper() == 'N/A':
            desc_input = ''

        category_input = input("Categoría [o=Operativo, a=Administrativo, m=Mixto] (ENTER=deduce de Excel 'Adtvo/Operativo'): ").strip().lower()
        
        # deducir adtop
        adtop_value = row_data[col_adtop] if col_adtop is not None else None
        adtop_value = str(adtop_value).strip() if adtop_value else ""
        if adtop_value.upper() == 'N/A':
            adtop_value = ''

        if not category_input:
            if adtop_value.lower().startswith("oper"):
                category = 'o'
            elif adtop_value.lower().startswith("ad"):
                category = 'a'
            else:
                category = 'm'
        else:
            category = category_input

        # deducir description
        if not desc_input and col_titulonaveg is not None:
            val_titulonaveg = row_data[col_titulonaveg]
            if val_titulonaveg and str(val_titulonaveg).strip().upper() != 'N/A':
                desc_input = str(val_titulonaveg).strip()

        # deducir regla
        rule = None
        if col_regla is not None:
            rule_val = row_data[col_regla]
            if rule_val:
                rule_val = str(rule_val).strip()
                if rule_val.upper() != 'N/A':
                    rule = rule_val

        # Nombre final
        final_name = cargo_name
        if not final_name:
            final_name = "CargoSinNombre"  # fallback mínimo

        self.stdout.write(f"Creando Position con name='{final_name}'...")

        cargo_obj = Position.objects.create(
            name=final_name,
            description=desc_input,
            category=category
        )
        if hasattr(cargo_obj, 'rule'):
            cargo_obj.rule = rule or ''
        if cargo_obj.description and cargo_obj.description.upper() == 'N/A':
            cargo_obj.description = ''
        if cargo_obj.rule and cargo_obj.rule.upper() == 'N/A':
            cargo_obj.rule = ''

        cargo_obj.save()
        
        self.stdout.write(self.style.SUCCESS(f"Cargo '{cargo_obj.name}' creado con ID={cargo_obj.id}"))
        return cargo_obj

    def update_position_interactive(self, cargo_obj, row_data, col_adtop, col_titulonaveg, col_regla):
        # Primero, si cargo_obj.description o cargo_obj.rule es "N/A", limpiarlo
        if cargo_obj.description and cargo_obj.description.upper() == "N/A":
            cargo_obj.description = ''
        if hasattr(cargo_obj, 'rule') and cargo_obj.rule and cargo_obj.rule.upper() == "N/A":
            cargo_obj.rule = ''

        adtop_value = row_data[col_adtop] if col_adtop is not None else None
        adtop_value = str(adtop_value).strip() if adtop_value else ""
        if adtop_value.upper() == 'N/A':
            adtop_value = ''

        if adtop_value.lower().startswith("oper"):
            new_cat = 'o'
        elif adtop_value.lower().startswith("ad"):
            new_cat = 'a'
        else:
            new_cat = 'm'
        
        if cargo_obj.category != new_cat:
            self.stdout.write(f"Cargo '{cargo_obj.name}' => category actual='{cargo_obj.category}', Excel='{new_cat}'")
            while True:
                action = input("¿Actualizar? [s/n]: ").strip().lower()
                if action in ['s','n']:
                    break
                print("Ingresa s o n.")
            if action == 's':
                cargo_obj.category = new_cat

        # titulo_naveg => possible desc
        if col_titulonaveg is not None:
            val_tn = row_data[col_titulonaveg]
            if val_tn:
                val_tn = str(val_tn).strip()
                if val_tn.upper() == "N/A":
                    val_tn = ''
                if val_tn:
                    if cargo_obj.description:
                        self.stdout.write(f"Cargo '{cargo_obj.name}' ya tiene desc='{cargo_obj.description}'. Excel='{val_tn}'")
                        while True:
                            action = input("[O]verwrite, [C]oncat, [I]gnore? ").strip().lower()
                            if action in ['o','c','i']:
                                break
                            print("Opción inválida. Ingresa o, c o i.")
                        if action == 'o':
                            cargo_obj.description = val_tn
                        elif action == 'c':
                            cargo_obj.description += " " + val_tn
                    else:
                        cargo_obj.description = val_tn
        
        # Regla
        if col_regla is not None and hasattr(cargo_obj, 'rule'):
            rule_val = row_data[col_regla]
            if rule_val:
                rule_val = str(rule_val).strip()
                if rule_val.upper() == "N/A":
                    rule_val = ''
                if rule_val:
                    if cargo_obj.rule:
                        self.stdout.write(f"Cargo '{cargo_obj.name}' => rule actual='{cargo_obj.rule}', Excel='{rule_val}'")
                        while True:
                            act = input("[O]verwrite, [I]gnore? ").strip().lower()
                            if act in ['o','i']:
                                break
                            print("Opción inválida. Ingresa o o i.")
                        if act == 'o':
                            cargo_obj.rule = rule_val
                    else:
                        cargo_obj.rule = rule_val

        if cargo_obj.description and cargo_obj.description.upper() == 'N/A':
            cargo_obj.description = ''
        if cargo_obj.rule and cargo_obj.rule.upper() == 'N/A':
            cargo_obj.rule = ''

        cargo_obj.save()
        return cargo_obj

    def find_position_by_id_interactive(self):
        """Permite ingresar manualmente un ID de Position ya existente."""
        while True:
            pos_id = input("Ingresa ID del cargo (o 'x' para cancelar): ").strip().lower()
            if pos_id == 'x':
                return None
            if pos_id.isdigit():
                pos_obj = Position.objects.filter(id=int(pos_id)).first()
                if pos_obj:
                    self.stdout.write(self.style.SUCCESS(f"Asig. cargo ID={pos_id}, name='{pos_obj.name}'"))
                    return pos_obj
                else:
                    self.stdout.write(self.style.ERROR(f"No existe un cargo con ID={pos_id}"))
            else:
                self.stdout.write("Debes ingresar un número entero o 'x' para cancelar.")

    def link_document_to_position(self, doc_header, cargo_obj):
        """
        Asocia doc_header a cargo_obj si no existe PositionDocument.
        Si doc_header se ha resuelto previamente, se reusa la doc en self.doc_header_cache.
        """
        # Primero, si doc_header ya está en cache => no preguntamos nada:
        if doc_header in self.doc_header_cache:
            doc_obj = self.doc_header_cache[doc_header]
            self._create_position_document_if_missing(cargo_obj, doc_obj)
            return

        # Buscamos doc por nombre parecido
        all_docs = list(Document.objects.all())
        names = [d.name for d in all_docs]
        close = difflib.get_close_matches(doc_header, names, n=1, cutoff=0.6)
        if close:
            matched_name = close[0]
            doc_obj = next(d for d in all_docs if d.name == matched_name)
            # Guardamos en cache
            self.doc_header_cache[doc_header] = doc_obj
        else:
            self.stdout.write(self.style.WARNING(f"No se encontró Document similar a '{doc_header}'"))
            while True:
                action = input("¿Deseas [c]rear doc, [m]anual ID, [i]gnorar? (c/m/i): ").strip().lower()
                if action in ['c','m','i']:
                    break
                print("Opción inválida (c/m/i).")
            
            doc_obj = None
            if action == 'c':
                doc_name = input(f"Ingresa name para Document (ENTER => usar '{doc_header}'): ").strip()
                if not doc_name:
                    doc_name = doc_header  # no permitir vacio
                desc = input("Descripción (opcional): ").strip()
                doc_obj = Document.objects.create(name=doc_name, description=desc)
                # guardamos en cache
                self.doc_header_cache[doc_header] = doc_obj
            elif action == 'm':
                while True:
                    did = input("Ingresa ID del Document o 'x' para cancelar: ").strip().lower()
                    if did == 'x':
                        break
                    if did.isdigit():
                        dtemp = Document.objects.filter(id=int(did)).first()
                        if dtemp:
                            doc_obj = dtemp
                            self.stdout.write(self.style.SUCCESS(f"Asig. doc ID={did} => '{doc_obj.name}'"))
                            self.doc_header_cache[doc_header] = doc_obj
                            break
                        else:
                            self.stdout.write(self.style.ERROR(f"No hay Document con ID={did}"))
                    else:
                        self.stdout.write("Ingresa un número o 'x'.")
            
            if action == 'i' or (action == 'm' and not doc_obj):
                return  # ignoramos

        if not doc_obj:
            return
        
        self._create_position_document_if_missing(cargo_obj, doc_obj)

    def _create_position_document_if_missing(self, cargo_obj, doc_obj):
        """Crea PositionDocument si no existe (para link_document_to_position)."""
        pd_exist = PositionDocument.objects.filter(position=cargo_obj, document=doc_obj).first()
        if pd_exist:
            self.stdout.write(f"El cargo '{cargo_obj.name}' ya tiene doc '{doc_obj.name}'.")
        else:
            PositionDocument.objects.create(
                position=cargo_obj,
                document=doc_obj,
                mandatory=True
            )
            self.stdout.write(self.style.SUCCESS(
                f"Asociado doc '{doc_obj.name}' al cargo '{cargo_obj.name}'."
            ))
