# dth/management/commands/fill_payroll_details.py

from django.core.management.base import BaseCommand
from django.db import transaction, DataError
from openpyxl import load_workbook

from dth.models import Nomina, PayrollDetails, EPS

import re
import datetime

class Command(BaseCommand):
    help = (
        "Lee un Excel (hoja 'MARZO') y actualiza campos en Nomina y PayrollDetails.\n"
        "Muestra debug info de cualquier valor que exceda max_length."
    )

    def add_arguments(self, parser):
        parser.add_argument('excel_path', type=str, help="Ruta local del archivo Excel a procesar")

    @transaction.atomic
    def handle(self, *args, **options):
        excel_path = options['excel_path']

        try:
            wb = load_workbook(filename=excel_path, data_only=True)
        except FileNotFoundError:
            self.stdout.write(self.style.ERROR(f"El archivo '{excel_path}' no fue encontrado."))
            return
        
        sheet_name = 'MARZO'
        if sheet_name not in wb.sheetnames:
            self.stdout.write(self.style.ERROR(f"No se encontró la hoja '{sheet_name}' en el Excel."))
            return
        
        ws = wb[sheet_name]
        self.stdout.write(self.style.NOTICE(f"Abriendo hoja '{sheet_name}'..."))

        # Diccionarios de mapeo...
        gender_map = {
            'M': 'h',  # Masculino => 'h'
            'F': 'm',  # Femenino => 'm'
        }
        education_map = {
            'Ninguno': 'none',
            'Primaria Incompleta': 'sec_incompleta',
            'Primaria Completa': 'sec_incompleta',
            'Secundaria Incompleta': 'sec_incompleta',
            'Secundaria Completa': 'sec_completa',
            'Técnico': 'tecnico',
            'Tecnólogo': 'tecnologo',
            'Pregrado': 'pregrado',
            'Postgrado': 'postgrado',
        }
        marital_map = {
            'casado': ['Casado/da', 'casado/da', 'casado/a', 'casado(da)'],
            'unionlibre': ['Union Libre', 'Unión Libre'],
            'soltero': ['Soltero/ra', 'soltero/ra', 'soltero(a)'],
            'viudo': ['Viudo(a)', 'viudo(a)'],
            'divorciado': ['Divorciado(a)', 'divorciado(a)'],
        }
        shift_map = {
            '5 x 2 (Lunes a Viernes)': '5x2',
            '6 x 1': '6x1',
            '6 x 1 (Navegando 2 x 1)': '6x1_nav',
            '14 x 7': '14x7',
            'Jornada Flexible': 'flexible',
            '14 x 14 (1 x 1)': '14x14',
        }
        criticity_map = {
            'BAJO': 'bajo',
            'MEDIO': 'medio',
            'ALTO': 'alto',
        }
        salary_type_map = {
            'Ordinario': 'ordinario',
            'Variable': 'variable',
            'Integral': 'integral',
            'Especie': 'especie',
        }

        eps_cache = {}
        
        total_rows_processed = 0
        updated_nomina_count = 0
        updated_details_count = 0
        no_nomina_count = 0

        row_index = 4
        while True:
            # Identificacion => columna C
            id_number_cell = f"C{row_index}"
            id_number_val = ws[id_number_cell].value
            if not id_number_val:
                self.stdout.write(self.style.WARNING(
                    f"Fin de datos al llegar a fila {row_index}. No se encontró 'id_number'."
                ))
                break

            id_number_str = str(id_number_val).strip()
            if id_number_str.lower() == "identificacion":
                self.stdout.write(self.style.WARNING(
                    f"No se encontró registro en Nomina con id_number={id_number_str} (fila {row_index})."
                ))
                user_decision = input("¿Deseas continuar con la siguiente fila? (s/n): ").lower()
                if user_decision.startswith('n'):
                    break
                row_index += 1
                continue

            # Buscamos Nomina
            try:
                nomina_obj = Nomina.objects.get(id_number=id_number_str)
            except Nomina.DoesNotExist:
                self.stdout.write(self.style.WARNING(
                    f"No se encontró registro en Nomina con id_number={id_number_str} (fila {row_index})."
                ))
                no_nomina_count += 1
                user_decision = input("¿Deseas continuar con la siguiente fila? (s/n): ").lower()
                if user_decision.startswith('n'):
                    break
                row_index += 1
                continue
            except Nomina.MultipleObjectsReturned:
                self.stdout.write(self.style.ERROR(
                    f"Existen múltiples registros en Nomina con id_number={id_number_str} (fila {row_index})."
                ))
                user_decision = input("¿Deseas continuar con la siguiente fila? (s/n): ").lower()
                if user_decision.startswith('n'):
                    break
                row_index += 1
                continue

            details_obj, _created = PayrollDetails.objects.get_or_create(nomina=nomina_obj)

            row_updates_nomina = False
            row_updates_details = False

            ### PEQUEÑA FUNCIÓN DE DEBUG
            def debug_update(model_name, field_name, old_val, new_val):
                """Imprime info de la actualización."""
                old_str = str(old_val) if old_val is not None else ""
                new_str = str(new_val) if new_val is not None else ""
                self.stdout.write(
                    f"  Updating {model_name}.{field_name} from '{old_str[:40]}' to '{new_str[:40]}' "
                    f"(length {len(new_str)})"
                )

            # Col G => genero => M / F
            g_cell = f"G{row_index}"
            excel_gender_str = safe_str_cell(ws[g_cell].value).upper()
            if excel_gender_str in gender_map:
                desired_gender = gender_map[excel_gender_str]
                if nomina_obj.gender != desired_gender:
                    debug_update("Nomina", "gender", nomina_obj.gender, desired_gender)  ### DEBUG
                    nomina_obj.gender = desired_gender
                    row_updates_nomina = True

            # Col K => birth_date => parse date
            k_cell = f"K{row_index}"
            birth_val = parse_excel_date(ws[k_cell].value)
            if birth_val is not None and details_obj.birth_date != birth_val:
                debug_update("PayrollDetails", "birth_date", details_obj.birth_date, birth_val)  ### DEBUG
                details_obj.birth_date = birth_val
                row_updates_details = True

            # Col M => place_of_birth
            place_of_birth_val = safe_str_cell(ws[f"M{row_index}"].value)
            if place_of_birth_val and details_obj.place_of_birth != place_of_birth_val:
                debug_update("PayrollDetails", "place_of_birth", details_obj.place_of_birth, place_of_birth_val)
                details_obj.place_of_birth = place_of_birth_val
                row_updates_details = True

            # Col N => doc_expedition_date
            doc_exp_val = parse_excel_date(ws[f"N{row_index}"].value)
            if doc_exp_val and details_obj.doc_expedition_date != doc_exp_val:
                debug_update("PayrollDetails", "doc_expedition_date", details_obj.doc_expedition_date, doc_exp_val)
                details_obj.doc_expedition_date = doc_exp_val
                row_updates_details = True

            # Col O => doc_expedition_department
            depto_val = safe_str_cell(ws[f"O{row_index}"].value)
            if depto_val and details_obj.doc_expedition_department != depto_val:
                debug_update("PayrollDetails", "doc_expedition_department", details_obj.doc_expedition_department, depto_val)
                details_obj.doc_expedition_department = depto_val
                row_updates_details = True

            # Col P => doc_expedition_municipality
            muni_val = safe_str_cell(ws[f"P{row_index}"].value)
            if muni_val and details_obj.doc_expedition_municipality != muni_val:
                debug_update("PayrollDetails", "doc_expedition_municipality", details_obj.doc_expedition_municipality, muni_val)
                details_obj.doc_expedition_municipality = muni_val
                row_updates_details = True

            # Col Q => education_level
            edu_raw = safe_str_cell(ws[f"Q{row_index}"].value)
            if edu_raw in education_map:
                final_edu = education_map[edu_raw]
                if details_obj.education_level != final_edu:
                    debug_update("PayrollDetails", "education_level", details_obj.education_level, final_edu)
                    details_obj.education_level = final_edu
                    row_updates_details = True

            # Col R => profession
            prof_val = safe_str_cell(ws[f"R{row_index}"].value)
            if prof_val and details_obj.profession != prof_val:
                debug_update("PayrollDetails", "profession", details_obj.profession, prof_val)
                details_obj.profession = prof_val
                row_updates_details = True

            # Col V => last_academic_institution
            last_inst_val = safe_str_cell(ws[f"V{row_index}"].value)
            if last_inst_val and details_obj.last_academic_institution != last_inst_val:
                debug_update("PayrollDetails", "last_academic_institution", details_obj.last_academic_institution, last_inst_val)
                details_obj.last_academic_institution = last_inst_val
                row_updates_details = True

            # Col W => phone => Nomina
            phone_val = safe_str_cell(ws[f"W{row_index}"].value)
            if phone_val and nomina_obj.phone != phone_val:
                debug_update("Nomina", "phone", nomina_obj.phone, phone_val)
                nomina_obj.phone = phone_val
                row_updates_nomina = True

            # Col X => municipality_of_residence => details
            mun_res_val = safe_str_cell(ws[f"X{row_index}"].value)
            if mun_res_val and details_obj.municipality_of_residence != mun_res_val:
                debug_update("PayrollDetails", "municipality_of_residence", details_obj.municipality_of_residence, mun_res_val)
                details_obj.municipality_of_residence = mun_res_val
                row_updates_details = True

            # Col Y => address => details
            addr_val = safe_str_cell(ws[f"Y{row_index}"].value)
            if addr_val and details_obj.address != addr_val:
                debug_update("PayrollDetails", "address", details_obj.address, addr_val)
                details_obj.address = addr_val
                row_updates_details = True

            # Col Z => email => Nomina
            email_val = safe_str_cell(ws[f"Z{row_index}"].value)
            if email_val and nomina_obj.email != email_val:
                debug_update("Nomina", "email", nomina_obj.email, email_val)
                nomina_obj.email = email_val
                row_updates_nomina = True

            # Col AA => rh => details
            rh_val = safe_str_cell(ws[f"AA{row_index}"].value)
            if rh_val and details_obj.rh != rh_val:
                debug_update("PayrollDetails", "rh", details_obj.rh, rh_val)
                details_obj.rh = rh_val
                row_updates_details = True

            # Col AB => marital_status => details
            ms_val = safe_str_cell(ws[f"AB{row_index}"].value)
            final_ms = map_any(ms_val, marital_map)
            if final_ms and details_obj.marital_status != final_ms:
                debug_update("PayrollDetails", "marital_status", details_obj.marital_status, final_ms)
                details_obj.marital_status = final_ms
                row_updates_details = True

            # Col AD => EPS => details.eps
            eps_name_val = safe_str_cell(ws[f"AD{row_index}"].value)
            if eps_name_val:
                if eps_name_val in eps_cache:
                    eps_obj = eps_cache[eps_name_val]
                else:
                    try:
                        eps_obj = EPS.objects.get(name__iexact=eps_name_val)
                        eps_cache[eps_name_val] = eps_obj
                    except EPS.DoesNotExist:
                        self.stdout.write(self.style.WARNING(
                            f"No se encontró EPS '{eps_name_val}' en la BD."
                        ))
                        user_inp = input("Ingresa manualmente el ID de EPS o deja vacío para omitir: ").strip()
                        if user_inp:
                            try:
                                eps_obj = EPS.objects.get(pk=user_inp)
                                eps_cache[eps_name_val] = eps_obj
                            except EPS.DoesNotExist:
                                self.stdout.write(self.style.ERROR("ID de EPS inválido. Omitiendo..."))
                                eps_obj = None
                        else:
                            eps_obj = None
                if eps_obj and details_obj.eps != eps_obj:
                    debug_update("PayrollDetails", "eps", details_obj.eps, eps_obj)
                    details_obj.eps = eps_obj
                    row_updates_details = True

            # Col AE => afp => details.afp
            afp_val = safe_str_cell(ws[f"AE{row_index}"].value).lower()
            if afp_val:
                if afp_val.startswith('prot'):
                    afp_val = 'proteccion'
                elif afp_val.startswith('porv'):
                    afp_val = 'porvenir'
                elif afp_val.startswith('colp'):
                    afp_val = 'colpensiones'
                if details_obj.afp != afp_val:
                    debug_update("PayrollDetails", "afp", details_obj.afp, afp_val)
                    details_obj.afp = afp_val
                    row_updates_details = True

            # Col AG => caja_compensacion
            ccomp_val = safe_str_cell(ws[f"AG{row_index}"].value)
            if ccomp_val and details_obj.caja_compensacion != ccomp_val:
                debug_update("PayrollDetails", "caja_compensacion", details_obj.caja_compensacion, ccomp_val)
                details_obj.caja_compensacion = ccomp_val
                row_updates_details = True

            # Col AM => center_of_work
            cow_val = safe_str_cell(ws[f"AM{row_index}"].value).lower()
            if 'cartagena' in cow_val:
                cow_val = 'cartagena'
            elif 'guyana' in cow_val:
                cow_val = 'guyana'
            if cow_val and details_obj.center_of_work != cow_val:
                debug_update("PayrollDetails", "center_of_work", details_obj.center_of_work, cow_val)
                details_obj.center_of_work = cow_val
                row_updates_details = True

            # Col AO => contract_type
            ctype_val = safe_str_cell(ws[f"AO{row_index}"].value).lower()
            if ctype_val.startswith('inde'):
                ctype_val = 'indefinido'
            elif ctype_val.startswith('defin'):
                ctype_val = 'definido'
            elif ctype_val.startswith('aprend'):
                ctype_val = 'aprendizaje'
            elif ctype_val.startswith('obra'):
                ctype_val = 'obra'
            if ctype_val and details_obj.contract_type != ctype_val:
                debug_update("PayrollDetails", "contract_type", details_obj.contract_type, ctype_val)
                details_obj.contract_type = ctype_val
                row_updates_details = True

            # Col AR => months_term
            mt_val = ws[f"AR{row_index}"].value
            if isinstance(mt_val, int) and mt_val > 0:
                if details_obj.months_term != mt_val:
                    debug_update("PayrollDetails", "months_term", details_obj.months_term, mt_val)
                    details_obj.months_term = mt_val
                    row_updates_details = True

            # Col AZ => shift
            shift_raw_val = safe_str_cell(ws[f"AZ{row_index}"].value)
            final_shift = shift_map.get(shift_raw_val, None)
            if final_shift and details_obj.shift != final_shift:
                debug_update("PayrollDetails", "shift", details_obj.shift, final_shift)
                details_obj.shift = final_shift
                row_updates_details = True

            # Col BB => criticity_level => BAJO/MEDIO/ALTO
            crit_raw_val = safe_str_cell(ws[f"BB{row_index}"].value).upper()
            final_crit = criticity_map.get(crit_raw_val, None)
            if final_crit and details_obj.criticity_level != final_crit:
                debug_update("PayrollDetails", "criticity_level", details_obj.criticity_level, final_crit)
                details_obj.criticity_level = final_crit
                row_updates_details = True

            # Col BC => salary_type
            stype_raw_val = safe_str_cell(ws[f"BC{row_index}"].value)
            final_stype = salary_type_map.get(stype_raw_val, None)
            if final_stype and details_obj.salary_type != final_stype:
                debug_update("PayrollDetails", "salary_type", details_obj.salary_type, final_stype)
                details_obj.salary_type = final_stype
                row_updates_details = True

            # Col BE => bank_account
            bacc_val = safe_str_cell(ws[f"BE{row_index}"].value)
            if bacc_val and details_obj.bank_account != bacc_val:
                debug_update("PayrollDetails", "bank_account", details_obj.bank_account, bacc_val)
                details_obj.bank_account = bacc_val
                row_updates_details = True

            # Col BF => bank
            bank_val = safe_str_cell(ws[f"BF{row_index}"].value)
            if bank_val and details_obj.bank != bank_val:
                debug_update("PayrollDetails", "bank", details_obj.bank, bank_val)
                details_obj.bank = bank_val
                row_updates_details = True

            # Guardamos
            if row_updates_nomina:
                ### Debug for Nomina
                try:
                    nomina_obj.save()
                except DataError as e:
                    self.stdout.write(self.style.ERROR(
                        f"Ocurrió DataError al guardar Nomina (fila {row_index}, id_number={id_number_str}). Revisa logs arriba."
                    ))
                    raise e

                updated_nomina_count += 1

            if row_updates_details:
                ### Debug for PayrollDetails
                try:
                    details_obj.save()
                except DataError as e:
                    self.stdout.write(self.style.ERROR(
                        f"Ocurrió DataError al guardar PayrollDetails (fila {row_index}, id_number={id_number_str}). Revisa logs arriba."
                    ))
                    raise e

                updated_details_count += 1

            total_rows_processed += 1
            self.stdout.write(f"Fila {row_index} => id_number={id_number_str} procesada.")
            row_index += 1

        self.stdout.write(self.style.SUCCESS("=== RESUMEN DE PROCESAMIENTO ==="))
        self.stdout.write(f"Filas procesadas: {total_rows_processed}")
        self.stdout.write(f"Nominas sin coincidencia: {no_nomina_count}")
        self.stdout.write(f"Nominas actualizadas: {updated_nomina_count}")
        self.stdout.write(f"Details actualizados: {updated_details_count}")
        self.stdout.write(self.style.SUCCESS("Proceso finalizado."))


def safe_str_cell(value):
    """Convierte cualquier valor a str y aplica strip(); si None => ''."""
    if value is None:
        return ''
    return str(value).strip()


def parse_excel_date(value):
    """Intenta parsear un valor en date. Devuelve None si no se puede."""
    if not value:
        return None
    if isinstance(value, datetime.date):
        return value

    text = str(value).strip()
    m1 = re.match(r'^(\d{1,2})-(\d{1,2})-(\d{4})$', text)
    if m1:
        d, mn, y = m1.groups()
        try:
            return datetime.date(int(y), int(mn), int(d))
        except ValueError:
            return None

    m2 = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})$', text)
    if m2:
        y, mn, d = m2.groups()
        try:
            return datetime.date(int(y), int(mn), int(d))
        except ValueError:
            return None

    return None


def map_any(value_str, mapping_dict):
    """
    Toma un string y lo busca en el mapping_dict.
    Ej:
      {
        'casado': ['Casado/da','Casado/a','casado(da)'],
        'unionlibre': ['Union Libre','Unión Libre']
      }
    Retorna la key si encuentra coincidencia; None si no la hay.
    """
    val_lower = value_str.lower()
    for final_key, synonyms in mapping_dict.items():
        for syn in synonyms:
            if syn.lower() == val_lower:
                return final_key
    return None
