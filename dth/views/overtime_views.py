from django.contrib.auth.decorators import permission_required, login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.db.models import Q, Prefetch
from django.http import HttpResponse, JsonResponse
from django.urls import reverse_lazy, reverse
from django.views.decorators.http import require_POST
from django.views.generic import CreateView, TemplateView
from django.shortcuts import get_object_or_404, redirect

import json
import openpyxl
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import time, date
from dateutil.relativedelta import relativedelta

from got.models import Asset
from dth.models.overtime import OvertimeProject, Overtime
from dth.models.payroll import Nomina
from dth.forms import OvertimeProjectForm
from dth.utils import parse_date_range, calculate_overtime, subtract_time_ranges, COLOMBIA_HOLIDAYS
from ntf.models import Notification


def filter_overtime_queryset(request):
    start_date, end_date = parse_date_range(request)
    asset_id = request.GET.get('asset', '')
    state_filter = request.GET.get('state', 'all')
    name_filter = request.GET.get('name', '').strip()
    cedula_filter = request.GET.get('cedula', '').strip()
    qs = OvertimeProject.objects.filter(report_date__gte=start_date, report_date__lt=end_date)

    # Filtrar por nombre
    if name_filter:
        qs = qs.filter(
            Q(overtime__worker__name__icontains=name_filter) |
            Q(overtime__worker__surname__icontains=name_filter) |
            Q(overtime__nombre_completo__icontains=name_filter)
        ).distinct()
    # Filtrar por cédula
    if cedula_filter:
        qs = qs.filter(
            Q(overtime__worker__id_number__icontains=cedula_filter) |
            Q(overtime__cedula__icontains=cedula_filter)
        ).distinct()
    # Filtrar por asset (abbreviation)
    if asset_id:
        qs = qs.filter(asset__abbreviation=asset_id)
    # Filtrar por estado
    if state_filter in ['a','b','c']:
        qs = qs.filter(overtime__state=state_filter).distinct()
        filtered_overtime = Overtime.objects.filter(state=state_filter)
        prefetch = Prefetch('overtime_set', queryset=filtered_overtime, to_attr='filtered_overtimes')
        qs = qs.prefetch_related(prefetch)
    else:
        # state = 'all'
        qs = qs.filter(overtime__isnull=False).distinct()
        filtered_overtime = Overtime.objects.all()
        prefetch = Prefetch('overtime_set', queryset=filtered_overtime, to_attr='filtered_overtimes')
        qs = qs.prefetch_related(prefetch)

    # Filtros por grupo de usuario:
    current_user = request.user
    if current_user.groups.filter(name='mto_members').exists():
        pass
    elif current_user.groups.filter(name='maq_members').exists():
        asset = Asset.objects.filter(Q(supervisor=current_user) | Q(capitan=current_user)).first()
        qs = qs.filter(asset=asset)
    elif current_user.groups.filter(name__in=['buzos_members','serport_members']).exists():
        qs = qs.none()
    return qs.order_by('-report_date', 'asset')


class OvertimeListView(LoginRequiredMixin, TemplateView):
    template_name = 'dth/overtime_list.html'
    
    def get_queryset(self):
        return filter_overtime_queryset(self.request)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['projects'] = self.get_queryset()
        start_date, end_date = parse_date_range(self.request)
        context['start_date'] = start_date.strftime('%Y-%m-%d')
        context['end_date'] = end_date.strftime('%Y-%m-%d')
        context['state'] = self.request.GET.get('state', 'all')
        context['asset_id'] = self.request.GET.get('asset', '')
        context['assets'] = Asset.objects.filter(show=True).order_by('name')
        context['name'] = self.request.GET.get('name', '')
        context['cedula'] = self.request.GET.get('cedula', '')
        context['holiday_dates'] = json.dumps([d.strftime('%Y-%m-%d') for d in COLOMBIA_HOLIDAYS.keys()])
        return context


@login_required
@permission_required('got.can_approve_overtime', raise_exception=True)
@require_POST
def approve_overtime(request):
    action = request.POST.get('action')
    remarks = request.POST.get('remarks', '')
    selected_overtime_ids = request.POST.getlist('selected_overtime')
    overtimes = Overtime.objects.filter(id__in=selected_overtime_ids)
    overtimes_list = list(overtimes)

    if action == 'approve':
        overtimes_updated_count = overtimes.update(state='a')
        messages.success(request, f"Se han aprobado {overtimes_updated_count} registros de horas extras correctamente.")
        
    elif action == 'reject':
        overtimes_updated_count = overtimes.update(state='b', remarks=remarks)
        messages.success(request, f"Se han rechazado {overtimes_updated_count} registros de horas extras correctamente.")

    user_overtimes_map = defaultdict(list)
    for ov in overtimes_list:
        if ov.project and ov.project.reported_by:
            user_overtimes_map[ov.project.reported_by].append(ov)

    for user_to_notify, ov_list in user_overtimes_map.items():
        if not ov_list:
            continue

        earliest_start = min([ov.start for ov in ov_list])
        latest_end = max([ov.end for ov in ov_list])

        first_project = min(ov_list, key=lambda o: (o.project.id if o.project else 9999999)).project
        if not first_project:
            continue

        earliest_str = earliest_start.strftime('%H:%M') if earliest_start else ''
        latest_str = latest_end.strftime('%H:%M') if latest_end else ''

        if action == 'approve':
            body_action = "aprobado"
            title = "Aprobación de Horas Extras"
        else:
            body_action = "rechazado"
            title = "Rechazo de Horas Extras"

        message_lines = []
        message_lines.append(f"Se han {body_action} las horas extras de {earliest_str} a {latest_str}.")
        if action == 'reject' and remarks:
            message_lines.append(f"Observación: {remarks}")
        message_final = "\n".join(message_lines)
        redirect_url = reverse('dth:overtime_list') + f"#project_{first_project.id}"

        Notification.objects.create(user=user_to_notify, title=title, message=message_final, redirect_url=redirect_url)

    next_url = request.POST.get('next', reverse('dth:overtime_list'))
    return redirect(next_url)


@login_required
@require_POST
def edit_overtime(request):
    overtime = get_object_or_404(Overtime, id=request.POST.get('overtime_id'))
    overtime.start = request.POST.get('start')
    overtime.end = request.POST.get('end')
    overtime.state = request.POST.get('state')
    overtime.remarks = request.POST.get('remarks', '')
    if overtime.start >= overtime.end:
        messages.error(request, "La hora de inicio no puede ser posterior a la hora final.")
    else:
        overtime.save()
        messages.success(request, "Horas extras actualizadas.")
    return redirect(request.POST.get('next', '/'))


@login_required
@require_POST
def delete_overtime(request):
    overtime = get_object_or_404(Overtime, id=request.POST.get('overtime_id'))
    overtime.delete()
    messages.success(request, "Registro eliminado correctamente.")
    return redirect(request.POST.get('next', '/'))


@login_required
def export_overtime_excel(request):
    qs = filter_overtime_queryset(request)

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    today = date.today()
    last_month = today - relativedelta(months=1)
    report_title = f"REPORTE 24 {last_month.strftime('%B')} hasta 24 {today.strftime('%B')} {today.year}"
    ws.append([report_title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.cell(row=1, column=1).font = openpyxl.styles.Font(size=14, bold=True)

    headers = [
        'Nombre completo', 'Fecha', 'Justificación',
        'Hora inicio', 'Hora fin', 'Estado', 'Transporte'
    ]
    ws.append(headers)
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    row_idx = 3
    for project in qs:
        for entry in project.filtered_overtimes:
            transporte = 'Sí' if entry.end >= time(19,0) else 'No'

            if entry.worker:
                cell_val = f"{entry.worker.name} {entry.worker.surname} - {entry.worker.id_number}"
            else:
                cell_val = f"{entry.nombre_completo} - {entry.cedula}"
            ws.cell(row=row_idx, column=1, value=cell_val)
            # ws.cell(row=row_idx, column=2, value=(entry.worker.id_number if entry.worker else entry.cedula))
            ws.cell(row=row_idx, column=2, value=project.report_date.strftime('%d/%m/%Y') if project.report_date else '')
            ws.cell(row=row_idx, column=3, value=project.description or '')
            ws.cell(row=row_idx, column=4, value=entry.start.strftime('%I:%M %p') if entry.start else '')
            ws.cell(row=row_idx, column=5, value=entry.end.strftime('%I:%M %p') if entry.end else '')

            # Estado
            state_text = ''
            if entry.state == 'a':
                state_text = "Aprobado"
            elif entry.state == 'b':
                state_text = "No aprobado"
            else:
                state_text = "Pendiente"
            ws.cell(row=row_idx, column=6, value=state_text)

            ws.cell(row=row_idx, column=7, value=transporte)
            row_idx += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Reporte_Horas_Extras_{today.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def buscar_nomina(request):
    cedula = request.GET.get('cedula')
    try:
        persona = Nomina.objects.get(id_number=cedula)
        return JsonResponse({'success': True, 'name': f"{persona.name} {persona.surname}", 'position': persona.position_id.name})
    except Nomina.DoesNotExist:
        return JsonResponse({'success': False})


class OvertimeProjectCreateView(LoginRequiredMixin, CreateView):
    model = OvertimeProject
    form_class = OvertimeProjectForm
    template_name = "dth/create_overtime_report.html"
    success_url = reverse_lazy("dth:overtime_list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        holiday_list = []
        for hol_date in COLOMBIA_HOLIDAYS:
            holiday_list.append(hol_date.isoformat())
        context["holiday_dates"] = json.dumps(holiday_list)

        chosen_date = None
        form = context.get("form")
        if form and form.instance and form.instance.report_date:
            chosen_date = form.instance.report_date
        if not chosen_date:
            chosen_date = date.today()  # por defecto, hoy
        context["initial_date"] = chosen_date
        return context

    def form_valid(self, form):
        project = form.save(commit=False)
        project.reported_by = self.request.user

        if self.request.user.groups.filter(name='maq_members').exists(): # Asignar asset si user es maq_members
            asset = Asset.objects.filter(Q(supervisor=self.request.user) | Q(capitan=self.request.user)).first()
            project.asset = asset

        # Datos del formulario
        report_date = form.cleaned_data['report_date']
        start_time  = form.cleaned_data['start']
        end_time    = form.cleaned_data['end']

        # Calcular intervalos válidos
        overtime_periods = calculate_overtime(report_date, start_time, end_time)
        if not overtime_periods:
            messages.error(self.request, "Las horas ingresadas no califican como horas extras.")
            return self.form_invalid(form)  # Re-renderiza con errores

        # Leer cédulas / externos
        cedulas_json  = form.cleaned_data.get('cedulas', '[]')
        externos_json = form.cleaned_data.get('personas_externas', '[]')

        try:
            cedulas_list = json.loads(cedulas_json) if cedulas_json.strip() else []
        except:
            cedulas_list = []
        try:
            externos_list = json.loads(externos_json) if externos_json.strip() else []
        except:
            externos_list = []

        # Validar: al menos 1 persona
        if not cedulas_list and not externos_list:
            messages.error(self.request, "Debes ingresar al menos una persona (cédula o registro manual).")
            return self.form_invalid(form)
        project.save()

        # Función local para crear los sub-intervalos no solapados
        def create_intervals_for_overtime(worker_obj, ext_nombre, ext_cedula, ext_cargo):
            created_count = 0
            if worker_obj:
                existing = Overtime.objects.filter(worker=worker_obj, project__report_date=report_date)
            else:
                existing = Overtime.objects.filter(cedula=ext_cedula, project__report_date=report_date)
            existing_ranges = [(ov.start, ov.end) for ov in existing]

            for (ot_start, ot_end) in overtime_periods:
                final_subintervals = subtract_time_ranges(ot_start, ot_end, existing_ranges)
                for (sub_start, sub_end) in final_subintervals:
                    Overtime.objects.create(
                        start=sub_start,
                        end=sub_end,
                        worker=worker_obj,
                        project=project,
                        state='c',  # 'Pendiente'
                        nombre_completo=ext_nombre or '',
                        cedula=ext_cedula or '',
                        cargo=ext_cargo or ''
                    )
                    created_count += 1
                    existing_ranges.append((sub_start, sub_end))
                    existing_ranges.sort(key=lambda x: x[0])
            return created_count

        total_created = 0
        total_omitted = 0

        # Crear Overtime para Nomina
        for cedula in cedulas_list:
            try:
                persona = Nomina.objects.get(id_number=cedula)
            except Nomina.DoesNotExist:
                total_omitted += 1
                continue
            created = create_intervals_for_overtime(
                worker_obj=persona,
                ext_nombre=None,
                ext_cedula=None,
                ext_cargo=None
            )
            if created == 0:
                total_omitted += 1
            else:
                total_created += created

        # Crear Overtime para externos
        for externo_data in externos_list:
            ext_ced   = externo_data.get('cedula', '')
            ext_name  = externo_data.get('nombre_completo', '')
            ext_cargo = externo_data.get('cargo', '')
            created = create_intervals_for_overtime(
                worker_obj=None,
                ext_nombre=ext_name,
                ext_cedula=ext_ced,
                ext_cargo=ext_cargo
            )
            if created == 0:
                total_omitted += 1
            else:
                total_created += created

        # Mensajes finales
        if total_created > 0:
            messages.success(self.request, f"Reporte creado. Se generaron {total_created} registros de horas extras. Omitidos: {total_omitted}.")
            return super().form_valid(form)
        else:
            project.delete()
            messages.error(self.request, "E1: No se pudo crear ninguna hora extra.")
            return self.form_invalid(form)

    def form_invalid(self, form):
        return super().form_invalid(form)