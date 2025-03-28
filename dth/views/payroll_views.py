# dth/views/payroll_views.py
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse
from django.views.generic import ListView, UpdateView
from django.views.decorators.http import require_GET
from django.urls import reverse_lazy
from django.template.loader import render_to_string

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, NamedStyle
from datetime import datetime, date
from decimal import Decimal

from dth.models.payroll import Nomina, NominaReport, UserProfile
from dth.forms import UserProfileForm, UploadNominaReportForm, NominaForm

@login_required
def profile_update(request):
    user = request.user
    # Intenta obtener el perfil; si no existe, créalo
    profile, created = UserProfile.objects.get_or_create(user=user)
    
    if request.method == 'POST':
        form = UserProfileForm(request.POST, request.FILES, instance=profile, user=user)
        if form.is_valid():
            # Actualizar datos del usuario
            user.first_name = form.cleaned_data.get('first_name')
            user.last_name = form.cleaned_data.get('last_name')
            user.email = form.cleaned_data.get('email')
            user.save()
            # Guardar perfil
            form.save()
            messages.success(request, 'Tu perfil ha sido actualizado exitosamente.')
            return redirect('got:asset-list')
        else:
            messages.error(request, 'Por favor corrige los errores indicados.')
    else:
        form = UserProfileForm(instance=profile, user=user, initial={'first_name': user.first_name, 'last_name': user.last_name, 'email': user.email,})
    return render(request, 'dth/profile_update.html', {'form': form})


class NominaListView(ListView):
    model = Nomina
    template_name = 'dth/payroll_list.html'
    context_object_name = 'nominas'

    # Si deseas filtrar o personalizar la consulta, sobreescribe get_queryset
    # def get_queryset(self):
    #     return Nomina.objects.filter(...)

class NominaUpdateView(UpdateView):
    model = Nomina
    form_class = NominaForm
    template_name = 'dth/payroll_update.html'
    success_url = reverse_lazy('nomina_list')  # Redirige a la lista luego de actualizar

    # Si deseas manejar esto en modal con AJAX, 
    # no uses success_url sino un JSON Response, etc.


@require_GET
def nomina_detail_partial(request, pk):
    nomina = get_object_or_404(Nomina, pk=pk)
    html = render_to_string('dth/payroll_detail.html', {'nomina': nomina}, request=request)
    return HttpResponse(html, content_type='text/html')


def gerencia_nomina_view(request):
    template_name = "dth/gerencia_nomina.html"

    if request.method == "POST":
        form = UploadNominaReportForm(request.POST, request.FILES)
        if form.is_valid():
            # Se extrae el archivo excel
            excel_file = request.FILES['excel_file']
            try:
                # Abrimos el workbook con openpyxl
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                # Asumimos que solo hay UNA hoja
                sheet = wb.active

                # Mapeo de codigos -> nombre de campo en NominaReport
                # Clave = el prefijo con el que inicia la celda de "Concepto"
                CODE_FIELD_MAP = {
                    'DV01': 'dv01',
                    'DV23': 'dv25',
                    'DV03': 'dv03',
                    'DV103': 'dv103',
                    'DV27': 'dv27',
                    'DV30': 'dv30',
                    'DX03': 'dx03',
                    'DX05': 'dx05',
                    'DX01': 'dx01',
                    'DX07': 'dx07',
                    'DX12': 'dx12',
                    'DX63': 'dx63',
                    'DX64': 'dx64',
                    'DX66': 'dx66',
                    # Ejemplo: si tienes más códigos, agrégalos aquí.
                    # 'DV48': 'dv48', etc.
                }

                rows_created = 0
                rows_updated = 0

                # Buscar el índice de columna (asumiendo encabezados en la primera fila)
                # Para evitar problemas con la posición exacta, lo hacemos flexible:
                headers = {cell.value: idx for idx, cell in enumerate(sheet[1], start=1)}
                print("DEBUG - Headers leídos:", headers)

                # Revisar que existan las columnas mínimas
                required_cols = [
                    "Empleado", "FechaMovimiento", "Concepto",
                    "ValorDevengo", "ValorDeduccion"
                ]
                clean_headers = {str(k).strip(): v for k, v in headers.items() if k is not None}

                for col in required_cols:
                    if col not in clean_headers:
                        messages.error(request, f"Falta la columna requerida '{col}' en el Excel.")
                        return render(request, template_name, {"form": form})

                # Iterar sobre las filas (omitir la primera, que es encabezado)
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # Tomar valores de la fila según las columnas
                    empleado = row[headers["Empleado"] - 1]
                    fecha_mov = row[headers["FechaMovimiento"] - 1]  # string con fecha

                    fecha_celda = row[headers["FechaMovimiento"] - 1]
                    if isinstance(fecha_celda, datetime):
                        # Es un objeto datetime => tomar su fecha
                        parsed_date = fecha_celda.date()
                        mes = parsed_date.month
                        anio = parsed_date.year
                    elif isinstance(fecha_celda, date):
                        # A veces openpyxl puede devolver date directamente (sin tiempo)
                        mes = fecha_celda.month
                        anio = fecha_celda.year
                    elif isinstance(fecha_celda, str):
                        # Intentar parsear como "d/m/yyyy"
                        # Ojo con más validaciones según tu archivo
                        try:
                            parsed_date = datetime.strptime(fecha_celda.strip(), "%d/%m/%Y")
                            mes = parsed_date.month
                            anio = parsed_date.year
                        except ValueError:
                            # Manejar formato inválido
                            continue
                    else:
                        # No es ninguno de los tipos esperados => se ignora
                        continue

                    concepto_original = row[clean_headers["Concepto"] - 1] or ""
                    concepto_upper = str(concepto_original).strip().upper()

                    print("DEBUG - Concepto leido:", concepto_original) 

                    # Settear found_field en None
                    found_field = None
                    for code_prefix, field_name in CODE_FIELD_MAP.items():
                        # Buscamos si concepto_upper arranca con code_prefix
                        if concepto_upper.startswith(code_prefix):
                            found_field = field_name
                            break

                    if found_field:
                        print("DEBUG - matched", found_field, "para concepto", concepto_upper)
                    else:
                        print("DEBUG - sin match para", concepto_upper)

                    valor_devengo = row[headers["ValorDevengo"] - 1] or 0
                    valor_deduccion = row[headers["ValorDeduccion"] - 1] or 0

                    # Verificamos si hay en Nomina
                    if empleado is None:
                        continue  # Ignorar filas vacías
                    try:
                        nomina_obj = Nomina.objects.get(id_number=str(empleado))
                    except Nomina.DoesNotExist:
                        # Si no existe en Nomina, puedes decidir crearlo o ignorarlo
                        # Aquí optamos por ignorar y continuar
                        continue

                    # Extraer mes y año de 'FechaMovimiento' (ej: "1/02/2025")
                    # Asumiendo formato: d/m/yyyy
                    # try:
                    #     parsed_date = datetime.strptime(str(fecha_mov), "%d/%m/%Y")
                    #     mes = parsed_date.month
                    #     anio = parsed_date.year
                    # except ValueError:
                    #     # Si no logra parsear, omitir
                    #     continue

                    # Verificar si ya existe un NominaReport con esa Nomina y mes/año
                    with transaction.atomic():
                        nomina_report, created = NominaReport.objects.get_or_create(
                            nomina=nomina_obj,
                            mes=mes,
                            anio=anio
                        )
                        if created:
                            rows_created += 1
                        else:
                            rows_updated += 1

                        # Detectar si el Concepto inicia con algún código
                        # p.ej. "DV01 SUELDO BASICO" -> arranca con "DV01"
                        found_field = None
                        concepto_upper = concepto_upper.strip().upper()
                        for code_prefix, field_name in CODE_FIELD_MAP.items():
                            if concepto_upper.startswith(code_prefix):
                                found_field = field_name
                                break

                        # Tomar el valor "no cero" entre ValorDevengo y ValorDeduccion
                        # Si ambos son 0, no hacemos nada
                        valor = Decimal('0')
                        if valor_devengo not in [None, 0]:
                            valor = Decimal(valor_devengo)
                        elif valor_deduccion not in [None, 0]:
                            valor = Decimal(valor_deduccion)

                        if found_field and valor != 0:
                            # Asignar dinámicamente al campo correspondiente
                            setattr(nomina_report, found_field, valor)
                            nomina_report.save()

                # Al terminar
                messages.success(
                    request,
                    (f"Proceso finalizado. Se crearon {rows_created} registros nuevos "
                     f"y se actualizaron {rows_updated} registros existentes.")
                )
                return redirect('dth:gerencia_nomina')  # o simplemente recargar
            except Exception as e:
                messages.error(request, f"Ocurrió un error procesando el archivo: {str(e)}")
    else:
        form = UploadNominaReportForm()

    # ==== CONSULTAMOS LOS REGISTROS PARA LA TABLA ====
    # Recuperamos todos los NominaReport, con su Nomina relacionada
    # Orden: primero por el nombre del empleado, luego mes y año.
    # Nota: Para “nombre completo”, un truco es concatenar en un campo simulado
    #       o podemos sólo ordenar por “nomina__name” y “nomina__surname”.
    reports = (NominaReport.objects
               .select_related('nomina')
               .order_by('nomina__name', 'nomina__surname', 'mes', 'anio'))

    # Aquí hacemos los cálculos necesarios claramente:
    processed_reports = []
    for nr in reports:
        salario = nr.nomina.salary or 0
        dv01 = nr.dv01 or 0
        dv25 = nr.dv25 or 0
        dias_sueldo_basico = round((dv01 / (salario / 30)), 2) if salario else 0
        dias_vacaciones = round((dv25 / (salario / 30)), 2) if salario else 0

        processed_reports.append({
            'report': nr,
            'dias_sueldo_basico': dias_sueldo_basico,
            'dias_vacaciones': dias_vacaciones,
            # Aquí puedes añadir otros cálculos que necesites
        })

    nomina_qs = Nomina.objects.all().order_by('name', 'surname')

    context = {
        "form": form,
        "processed_reports": processed_reports,
        'nomina_list': nomina_qs
    }

    return render(request, template_name, context)


@login_required
def export_gerencia_nomina_excel(request):
    """
    Exporta la tabla de NominaReport a un archivo Excel con estilos y formatos 
    similares a la tabla de la vista de Nómina Gerencia.
    """
    # 1) Obtención de registros
    reports = (NominaReport.objects
               .select_related('nomina')
               .order_by('nomina__name', 'nomina__surname', 'mes', 'anio'))
    
    # 2) Crear workbook y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nómina Gerencia"
    
    # 3) Definir encabezados y configurar estilos
    headers = [
        "Cédula", "Nombre", "Cargo", "MES", "AÑO", "Salario",
        "# días sueldo básico", "Sueldo Básico (dv01)", "# días vacaciones", "Pago Vacaciones (dv25)",
        "Subsidio transporte (dv03)", "Licencia familia (dv103)", "Provisión vacaciones (4.17%)",
        "DV27 Intereses Ces. Ant", "Prima de Servicio (8.33%)", "Cesantías (dv30)",
        "Intereses de Cesantías (1%)", "Salud colab. (4%)", "Pensión colab (dx03)",
        "Fdo. solidar. (dx05)", "Pensión empleador (12%)", "ARL empleador (6.96%)",
        "Caja comp. (4%)", "Ret. Fuente", "Exequias Lordoy (dx07)",
        "Desc. pensión voluntaria (dx12)", "Banco Occidente (dx63)",
        "Confenalco (dx64)", "Préstamo (dx66)", "Neto a pagar"
    ]
    
    # Estilo para encabezados
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(border_style="thin", color="000000")
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # Escribir encabezados y ajustar anchos
    for col_num, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

    # Número de columna de cada campo de moneda (según índice en la lista de encabezados)
    currency_columns = [6, 8, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20,
                        21, 22, 23, 24, 25, 26, 27, 28, 29, 30]
    currency_format = '"COP " #,##0.00'

    # 4) Agregar filas de datos
    row_index = 2
    for nr in reports:
        emp = nr.nomina
        # Calcular # días sueldo básico y # días vacaciones
        dias_sb = Decimal('0.00')
        dias_vac = Decimal('0.00')
        if emp.salary and nr.dv01:
            dias_sb = nr.dv01 / (Decimal(emp.salary) / Decimal('30'))
        if emp.salary and nr.dv25:
            dias_vac = nr.dv25 / (Decimal(emp.salary) / Decimal('30'))
        
        data_row = [
            emp.id_number,
            f"{emp.name} {emp.surname}",
            emp.position,
            nr.mes,
            nr.anio,
            float(emp.salary or 0),
            float(dias_sb),
            float(nr.dv01 or 0),
            float(dias_vac),
            float(nr.dv25 or 0),
            float(nr.dv03 or 0),
            float(nr.dv103 or 0),
            float(nr.provision_vacaciones or 0),  # dv25 * 4.17%
            float(nr.dv27 or 0),
            float(nr.prima_servicio or 0),       # (dv01 + dv25 + dv03)*0.0833
            float(nr.dv30 or 0),
            float(nr.intereses_cesantias or 0),  # dv30 * 0.01
            float(nr.salud_aporte or 0),         # (dv01 + dv25) * 0.04
            float(nr.dx03 or 0),
            float(nr.dx05 or 0),
            float(nr.pension_aporte_empleador or 0),
            float(nr.arl_aporte or 0),
            float(nr.caja_compensacion_aporte or 0),
            float(nr.dx01 or 0),
            float(nr.dx07 or 0),
            float(nr.dx12 or 0),
            float(nr.dx63 or 0),
            float(nr.dx64 or 0),
            float(nr.dx66 or 0),
            float(nr.neto_a_pagar or 0),
        ]
        # Escribir la fila en la hoja
        for col_num, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row_index, column=col_num, value=value)
            cell.border = header_border
            # Alinear a la derecha para números (excepto cédula y texto)
            if col_num in currency_columns or col_num in [7, 9]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            # Si es columna de moneda, aplica el formato
            if col_num in currency_columns:
                cell.number_format = currency_format
        row_index += 1

    # 5) Retornar el archivo Excel como respuesta
    filename = "NominaGerencia.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def nomina_edit(request, pk):
    """
    Vista que edita o elimina un registro de Nomina.

    - Si POST incluye 'action=delete', se elimina el registro.
      -> Si es AJAX, se responde con JSON.
      -> Si no es AJAX, se redirige con un mensaje.
    - Si no incluye 'action=delete', se asume edición de campos.
    """
    nomina_obj = get_object_or_404(Nomina, pk=pk)

    if request.method == 'POST':
        # Verificamos si es petición de borrado (action=delete)
        action = request.POST.get('action')
        if action == 'delete':
            nomina_obj.delete()

            # Si es AJAX, devolvemos JSON para que JS manipule la tabla sin recargar
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'deleted': True})

            # No es AJAX => redirigimos con mensaje
            messages.success(request, "Empleado eliminado con éxito.")
            return redirect('dth:gerencia_nomina')

        # De lo contrario, es edición
        form = NominaForm(request.POST, instance=nomina_obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Registro de nómina actualizado.")
            return redirect('dth:gerencia_nomina')
        else:
            messages.error(request, "Por favor revisa los campos del formulario.")
    else:
        # GET => mostrar el formulario (opcional si deseas cargar un template)
        form = NominaForm(instance=nomina_obj)
    return render(request, 'dth/nomina_form.html', {'form': form, 'object': nomina_obj})


@login_required
def nomina_create(request):
    """
    Crea un nuevo registro de Nomina.
    """
    if request.method == 'POST':
        form = NominaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Empleado creado con éxito.')
            return redirect('dth:gerencia_nomina') # O la misma página modal
        else:
            messages.error(request, 'Hay errores en el formulario.')
    else:
        form = NominaForm()

    return render(request, 'dth/nomina_form.html', {'form': form})