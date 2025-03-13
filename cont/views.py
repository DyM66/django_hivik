# con/views.py
from decimal import Decimal
from cont.models import *
from openpyxl import load_workbook
from .forms import *
from ope.models import Operation
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404, redirect, render
from cont.forms import AssetCostUpdateForm
from django.urls import reverse_lazy
from django.views.generic.edit import CreateView
from django.views.generic import TemplateView, DetailView
from cont.mixins import FinanceMembersRequiredMixin
from collections import defaultdict


AREAS = {'a': 'Barcos', 'c': 'Barcazas',}
TODAY = timezone.now().date()

class AssetCostListView(FinanceMembersRequiredMixin, TemplateView):
    template_name = "cont/asset_list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        assets = Asset.objects.filter(show=True, area__in=AREAS.keys())
        operations = Operation.objects.filter(start__lte=TODAY, end__gte=TODAY, asset__in=assets).select_related('asset')
        projects = {op.asset.abbreviation: op.proyecto for op in operations}
        context['assets_by_area'] = {area_name: [asset for asset in assets if asset.area == area_code] for area_code, area_name in AREAS.items()}
        context['projects'] = projects
        context['area_icons'] = {'a': 'fa-ship', 'c': 'fa-solid fa-ferry'}
        return context

class FinanciacionCreateView(CreateView):
    model = Financiacion
    form_class = FinanciacionForm
    template_name = "cont/financiacion_form.html"
    
    def get_initial(self):
        initial = super().get_initial()
        asset_cost_pk = self.kwargs.get('asset_cost_pk')
        initial['asset_cost'] = asset_cost_pk
        return initial
    
    def form_valid(self, form):
        form.instance.asset_cost_id = self.kwargs.get('asset_cost_pk')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        asset_cost_pk = self.kwargs.get('asset_cost_pk')
        context['asset_cost_obj'] = AssetCost.objects.get(pk=asset_cost_pk)
        return context
    
    def get_success_url(self):
        return reverse_lazy('con:asset-detail', kwargs={'pk': self.kwargs.get('asset_cost_pk')})


class AssetCostDetailView(FinanceMembersRequiredMixin, DetailView):
    model = AssetCost
    template_name = "cont/asset_detail.html"
    context_object_name = "asset_cost"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ac = self.object
        total_cost = (ac.initial_cost or Decimal('0')) + (ac.costo_adicional or Decimal('0'))
        depreciation = total_cost / Decimal('3600') if total_cost > 0 else Decimal('0')
        context.update({
            'total_cost': total_cost,
            'depreciation': depreciation,
            'financiaciones': ac.financiaciones.all(),
            'total_interest': sum((fin.monto * fin.tasa_interes / Decimal('12') * fin.plazo) for fin in ac.financiaciones.all()),
            'total_interest_monthly': sum((fin.monto * fin.tasa_interes / Decimal('12')) for fin in ac.financiaciones.all()),
            'fp': ac.fp,
        })

        # Codigos contables agrupados por categoría
        categorias = {'ga': GastosAdministrativos, 'cd': CostoDirecto}
        codigos_por_categoria = {}

        for cat, modelo in categorias.items():
            codigos = CodigoContable.objects.filter(categoria=cat)
            registros = modelo.objects.all() if cat == 'ga' else modelo.objects.filter(assetcost=ac)

            meses_unicos = sorted(set(registros.values_list('mes', flat=True)))
            codigos_data = []
            for codigo in codigos:
                meses_data = []
                total_acumulado = Decimal('0.00')

                for mes in meses_unicos:
                    total_mes = registros.filter(codigo__startswith=codigo.codigo, mes=mes).aggregate(total_mes=models.Sum('total'))['total_mes'] or Decimal('0.00')
                    meses_data.append(total_mes)
                    total_acumulado += total_mes


                # Filtramos todos los registros que coincidan con el prefijo del código contable
                qs = registros.filter(codigo__startswith=codigo.codigo)
                detalles_list = []
                # Obtenemos los códigos únicos de GastosAdministrativos que empiecen con el código contable
                unique_codes = sorted(set(qs.values_list('codigo', flat=True)))
                for unique_code in unique_codes:
                    qs_unique = qs.filter(codigo=unique_code)
                    # Suponemos que la descripción es la misma para todos los registros con ese código
                    description = qs_unique.first().descripcion if qs_unique.exists() else ''
                    detalle_meses_data = []
                    detalle_total = Decimal('0.00')
                    for mes in meses_unicos:
                        total_mes_detalle = qs_unique.filter(mes=mes).aggregate(total_mes=models.Sum('total'))['total_mes'] or Decimal('0.00')
                        detalle_meses_data.append(total_mes_detalle)
                        detalle_total += total_mes_detalle
                    detalles_list.append({
                        'codigo': unique_code,
                        'descripcion': description,
                        'meses_data': detalle_meses_data,
                        'total_acumulado': detalle_total,
                    })

                codigos_data.append({
                    'codigo': codigo.codigo,
                    'descripcion': codigo.nombre,
                    'meses_data': meses_data,
                    'total_acumulado': total_acumulado,
                    'detalles_list': detalles_list,
                })

            codigos_por_categoria[cat] = {
                'meses_unicos': meses_unicos,
                'codigos_data': codigos_data,
            }

        context['codigos_por_categoria'] = codigos_por_categoria


        # Calcular totales globales para la categoría "ga"
        ga_data = codigos_por_categoria.get('ga', {})
        meses_unicos = ga_data.get('meses_unicos', [])
        totales_por_mes = [Decimal('0.00')] * len(meses_unicos)
        total_general = Decimal('0.00')
        for codigo in ga_data.get('codigos_data', []):
            for i, total_mes in enumerate(codigo['meses_data']):
                totales_por_mes[i] += total_mes
            total_general += codigo['total_acumulado']
        context['totales_gastos'] = {
            'totales_por_mes': [total * ac.fp for total in totales_por_mes],
            'total_general': total_general * ac.fp,
        }

        
        # Meses para mostrar nombre
        context['mes_nombres'] = {
            1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
            5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
            9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
        }
        return context


def update_assetcost(request, pk):
    assetcost = get_object_or_404(AssetCost, pk=pk)
    if request.method == "POST":
        form = AssetCostUpdateForm(request.POST, instance=assetcost)
        if form.is_valid():
            form.save()
            return redirect('con:asset-detail', pk=assetcost.pk)
    else:
        form = AssetCostUpdateForm(instance=assetcost)
    return render(request, 'cont/assetcost_form.html', {'form': form})
    
@require_POST
def gastos_upload_ajax(request):
    form = GastosUploadForm(request.POST, request.FILES)
    overwrite = request.POST.get("overwrite", "false").lower() == "true"
    if form.is_valid():
        excel_file = form.cleaned_data['excel_file']
        try:
            wb = load_workbook(filename=excel_file, data_only=True)
        except Exception as e:
            return JsonResponse({'success': False, 'error': 'Error al abrir el archivo.'})
        if 'query' not in wb.sheetnames:
            return JsonResponse({'success': False, 'error': 'La hoja "query" no existe en el archivo.'})
        ws = wb['query']
        
        # Se asume que la primera fila contiene los encabezados
        headers = [cell.value for cell in ws[1]]
        header_map = {header: idx for idx, header in enumerate(headers)}
        required_columns = ['cuenta', 'nombrecuenta', 'fecha', 'Total', 'centrocosto']
        for col in required_columns:
            if col not in header_map:
                return JsonResponse({'success': False, 'error': f'La columna requerida "{col}" no se encontró.'})
        
        # Obtener códigos válidos desde CodigoContable segregados por categoría
        codigos = CodigoContable.objects.all()
        valid_codes_ga = {c.codigo: c.nombre for c in codigos if c.categoria == 'ga'}
        valid_codes_cd = {c.codigo: c.nombre for c in codigos if c.categoria == 'cd'}
        
        # Diccionarios para agrupar filas según funcionalidad
        # Claves para GA: (cuenta, anio, mes, nombre_cuenta)
        # Claves para CD: (cuenta, anio, mes, nombre_cuenta, centrocostos)
        processed_groups_ga = {}
        detailed_records_ga = {}
        processed_groups_cd = {}
        detailed_records_cd = {}
        
        total_rows = 0
        matched_rows_ga = 0
        matched_rows_cd = 0
        
        from datetime import datetime
        import logging
        logger = logging.getLogger(__name__)
        for row in ws.iter_rows(min_row=2, values_only=True):
            total_rows += 1
            try:
                # Extraer y validar "cuenta"
                cuenta_val = row[header_map['cuenta']]
                if not cuenta_val:
                    continue
                cuenta_str = str(cuenta_val).strip()
                
                # Determinar si la fila corresponde a GA o CD
                is_ga = any(cuenta_str.startswith(code) for code in valid_codes_ga.keys())
                is_cd = any(cuenta_str.startswith(code) for code in valid_codes_cd.keys())
                if not is_ga and not is_cd:
                    continue
                
                # Campos comunes
                nombre_cuenta = row[header_map['nombrecuenta']]
                fecha_val = row[header_map['fecha']]
                if isinstance(fecha_val, str):
                    try:
                        # Se asume formato "d/m/Y" o "dd/mm/YYYY"
                        fecha_val = datetime.strptime(fecha_val.strip(), "%d/%m/%Y").date()
                    except Exception as e:
                        logger.warning(f"Error parseando fecha en fila {total_rows}: {fecha_val}. Error: {e}")
                        continue
                anio = fecha_val.year if fecha_val else None
                mes = fecha_val.month if fecha_val else None  # Se calcula el mes a partir de la fecha
                total_val = row[header_map['Total']]
                try:
                    # Se remueven comas y espacios
                    total_decimal = Decimal(str(total_val).replace(",", "").strip())
                except Exception as e:
                    logger.warning(f"Error convirtiendo Total en fila {total_rows}: {total_val}. Error: {e}")
                    total_decimal = Decimal('0.00')
                
                # Procesar Gastos Administrativos (GA)
                if is_ga:
                    key_ga = (cuenta_str, anio, mes, nombre_cuenta)
                    processed_groups_ga.setdefault(key_ga, Decimal('0.00'))
                    processed_groups_ga[key_ga] += total_decimal
                    detailed_records_ga.setdefault(key_ga, []).append({
                        'cuenta': cuenta_str,
                        'nombre_cuenta': nombre_cuenta,
                        'fecha': fecha_val,
                        'total': total_decimal,
                        'mes': mes
                    })
                    matched_rows_ga += 1
                
                # Procesar Costos Directos (CD)
                if is_cd:
                    centrocostos_val = row[header_map['centrocosto']]
                    if not centrocostos_val:
                        continue  # Omitir si está vacío
                    centrocostos_str = str(centrocostos_val).strip()
                    key_cd = (cuenta_str, anio, mes, nombre_cuenta, centrocostos_str)
                    processed_groups_cd.setdefault(key_cd, Decimal('0.00'))
                    processed_groups_cd[key_cd] += total_decimal
                    detailed_records_cd.setdefault(key_cd, []).append({
                        'cuenta': cuenta_str,
                        'nombre_cuenta': nombre_cuenta,
                        'fecha': fecha_val,
                        'total': total_decimal,
                        'mes': mes,
                        'centrocostos': centrocostos_str
                    })
                    matched_rows_cd += 1
            
            except Exception as e:
                logger.error(f"Error procesando fila {total_rows}: {row}. Error: {e}")
                continue
        
        # Detección de duplicados para GA
        duplicate_keys_ga = []
        try:
            for key in processed_groups_ga.keys():
                cuenta, anio, mes, nombre_cuenta = key
                try:
                    if GastosAdministrativos.objects.filter(codigo=cuenta, anio=anio, mes=mes).exists():
                        duplicate_keys_ga.append(key)
                except Exception as e:
                    logger.error(f"Error comprobando duplicado GA para {key}: {e}")
        except Exception as e:
            logger.error("Error en detección de duplicados GA: " + str(e))
        
        # Detección de duplicados para CD
        duplicate_keys_cd = []
        try:
            for key in processed_groups_cd.keys():
                cuenta, anio, mes, nombre_cuenta, centrocostos_str = key
                try:
                    assetcost_obj = AssetCost.objects.filter(codigo=centrocostos_str).first()
                    if assetcost_obj and CostoDirecto.objects.filter(codigo=cuenta, anio=anio, mes=mes, assetcost=assetcost_obj).exists():
                        duplicate_keys_cd.append(key)
                except Exception as e:
                    logger.error(f"Error comprobando duplicado CD para {key}: {e}")
        except Exception as e:
            logger.error("Error en detección de duplicados CD: " + str(e))
        
        if (duplicate_keys_ga or duplicate_keys_cd) and not overwrite:
            return JsonResponse({
                'success': False,
                'error': 'Ya existen registros para algunos grupos. ¿Desea sobrescribir los datos?',
                'overwrite_required': True,
            })
        
        # Procesar grupos GA: Crear o actualizar registros en GastosAdministrativos
        created_ga = 0
        updated_ga = 0
        for key, total_sum in processed_groups_ga.items():
            try:
                cuenta, anio, mes, nombre_cuenta = key
                qs = GastosAdministrativos.objects.filter(codigo=cuenta, anio=anio, mes=mes)
                if qs.exists():
                    obj = qs.first()
                    obj.total = total_sum
                    obj.save()
                    updated_ga += 1
                else:
                    GastosAdministrativos.objects.create(
                        codigo=cuenta,
                        descripcion=nombre_cuenta,
                        anio=anio,
                        mes=mes,
                        total=total_sum
                    )
                    created_ga += 1
            except Exception as e:
                logger.error(f"Error procesando grupo GA {key}: {e}")
                continue
        
        # Procesar grupos CD: Crear o actualizar registros en CostoDirecto
        created_cd = 0
        updated_cd = 0
        for key, total_sum in processed_groups_cd.items():
            try:
                cuenta, anio, mes, nombre_cuenta, centrocostos_str = key
                assetcost_obj = AssetCost.objects.filter(codigo=centrocostos_str).first()
                if not assetcost_obj:
                    continue  # Si no hay coincidencia, se omite el grupo
                qs = CostoDirecto.objects.filter(codigo=cuenta, anio=anio, mes=mes, assetcost=assetcost_obj)
                if qs.exists():
                    obj = qs.first()
                    obj.total = total_sum
                    obj.save()
                    updated_cd += 1
                else:
                    CostoDirecto.objects.create(
                        codigo=cuenta,
                        descripcion=nombre_cuenta,
                        total=total_sum,
                        assetcost=assetcost_obj,
                        mes=mes,
                        anio=anio
                    )
                    created_cd += 1
            except Exception as e:
                logger.error(f"Error procesando grupo CD {key}: {e}")
                continue
        
        message = (f"Archivo procesado correctamente: "
                   f"Gastos Administrativos - {created_ga} creados, {updated_ga} actualizados; "
                   f"Costos Directos - {created_cd} creados, {updated_cd} actualizados. "
                   f"Se leyeron {total_rows} filas, de las cuales {matched_rows_ga} para GA y {matched_rows_cd} para CD.")
        
        detailed_records_serializable = {
            'ga': {str(key): value for key, value in detailed_records_ga.items()},
            'cd': {str(key): value for key, value in detailed_records_cd.items()},
        }
        
        return JsonResponse({
            'success': True,
            'message': message,
            'details': detailed_records_serializable,
        })
    else:
        return JsonResponse({'success': False, 'errors': form.errors})
