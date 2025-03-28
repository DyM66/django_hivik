# inventory_management/views.py
import base64
import io
import qrcode
import qrcode.image.svg
import re
import requests

from io import BytesIO
from itertools import groupby
from operator import attrgetter

from django.contrib import messages
from django.contrib.auth.decorators import permission_required, login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.files.base import ContentFile
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models import Q, Sum, Max
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.views import generic, View
from django.views.decorators.csrf import csrf_exempt
from django.views.generic.edit import CreateView, UpdateView
from django.utils.http import urlencode
from django.db import transaction, IntegrityError



from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from got.models import Asset, System, Equipo, Suministro, Item#, Image
from got.models import Image as DjangoImage
from got.utils import *
from got.forms import ItemForm, UploadImages
from inv.utils import enviar_correo_transferencia
from .models import DarBaja
from .forms import *
from ope.models import Operation


class EquipoCreateView(LoginRequiredMixin, CreateView):
    model = Equipo
    form_class = EquipoForm
    template_name = 'got/systems/equipo_form.html'
    http_method_names = ['get', 'post']

    def dispatch(self, request, *args, **kwargs):
        self.next_url = request.GET.get('next') or request.POST.get('next') or ''
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['system'] = get_object_or_404(System, pk=self.kwargs['pk']) # Se inyecta el objeto system en los argumentos del formulario
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Se instancia el formulario de carga de imágenes
        context['upload_form'] = UploadImages(self.request.POST, self.request.FILES) if self.request.method == 'POST' else UploadImages()
        context['sys'] = get_object_or_404(System, pk=self.kwargs['pk'])
        return context

    def post(self, request, *args, **kwargs):
        self.object = None
        form = self.get_form()
        upload_form = UploadImages(request.POST, request.FILES)
        if form.is_valid() and upload_form.is_valid():
            return self.form_valid(form, upload_form)
        else:
            return self.form_invalid(form, upload_form)

    def form_valid(self, form, upload_form):
        system = get_object_or_404(System, pk=self.kwargs['pk'])
        form.instance.system = system
        asset_abbreviation = system.asset.abbreviation
        tipo = form.cleaned_data['tipo'].upper()

        try:
            with transaction.atomic():
                form.instance.code = generate_equipo_code(asset_abbreviation, tipo)
                form.instance.modified_by = self.request.user
                response = super().form_valid(form)
        except IntegrityError:
            form.add_error('code', 'Error al generar un código único. Por favor, inténtalo de nuevo.')
            return self.form_invalid(form, upload_form)

        for file in self.request.FILES.getlist('file_field'):
            DjangoImage.objects.create(image=file, equipo=self.object)
        messages.success(self.request, "Equipo creado exitosamente.")
        return response

    def form_invalid(self, form, upload_form):
        context = self.get_context_data(form=form, upload_form=upload_form)
        return self.render_to_response(context)

    def get_success_url(self):
        "Redirige a la URL 'next' si está presente; de lo contrario redirige a la vista de detalle del sistema."
        if self.next_url:
            return self.next_url
        else:
            return reverse('got:sys-detail', kwargs={'pk': self.object.system.pk})


class EquipoUpdate(UpdateView):
    model = Equipo
    form_class = EquipoForm
    template_name = 'got/systems/equipo_form.html'
    http_method_names = ['get', 'post']

    def dispatch(self, request, *args, **kwargs):
        try:
            obj = self.get_object()
        except Equipo.DoesNotExist:
            messages.error(request, "El equipo que intentas actualizar ya no existe. Por favor, actualiza la URL.")
            return redirect('got:asset-list')
        
        if str(obj.pk) != kwargs.get('pk'):
            return redirect(obj.get_absolute_url())
        
        self.object = obj
        self.next_url = request.GET.get('next') or request.POST.get('next') or ''
        self.next_to = request.GET.get('next_to') or request.POST.get('next') or ''
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        equipo = self.get_object()
        kwargs['system'] = get_object_or_404(System, equipos=equipo)
        return kwargs
    
    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        self.old_tipo = obj.tipo  
        self.old_code = obj.code
        return obj

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Formulario para carga de imágenes
        context['upload_form'] = UploadImages(self.request.POST, self.request.FILES) if self.request.method == 'POST' else UploadImages()
        ImageFormset = forms.modelformset_factory(Image, fields=('image',), extra=0)
        context['image_formset'] = ImageFormset(queryset=self.object.images.all())
        context['image_count'] = self.object.images.count()
        context['next_url'] = self.next_url
        context['next_to'] = self.next_to
        return context
    
    def form_valid(self, form):
        form.instance.modified_by = self.request.user
        response = super().form_valid(form)
        new_tipo = self.object.tipo  # el "tipo" que el usuario seleccionó 
        new_code = self.object.code 

        # if new_tipo != self.old_tipo:
        #     # Llamar a la función update_equipo_code con el code viejo
        #     update_equipo_code(self.old_code)
        #     messages.info(
        #         self.request, 
        #         f"El tipo cambió de '{self.old_tipo}' a '{new_tipo}'. Se actualizó el código del equipo."
        #     )
        
        upload_form = self.get_context_data()['upload_form']
        if upload_form.is_valid():
            for file in self.request.FILES.getlist('file_field'):
                Image.objects.create(image=file, equipo=self.object)
                print(file)
        return response
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        if "delete_image" in request.POST: # Opción para eliminar imagen vía formulario tradicional (fallback)
            image_id = request.POST.get("delete_image")
            Image.objects.filter(id=image_id, equipo=self.object).delete()
            return redirect(self.get_success_url())
        return super().post(request, *args, **kwargs)
    
    def get_success_url(self):
        url = reverse('got:equipo-detail', kwargs={'pk': self.object.pk})
        if self.next_to:
            params = {'previous_url': self.next_to}
            url = f"{url}?{urlencode(params)}"
        elif self.next_url:
            return self.next_url
        return url



from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
class ActivoEquipmentListView(View):
    template_name = 'inventory_management/asset_equipment_list.html'

    def get(self, request, abbreviation):
        activo = get_object_or_404(Asset, abbreviation=abbreviation)
        sistemas_ids = get_full_systems_ids(activo, request.user)
        equipos = Equipo.objects.filter(system__in=sistemas_ids).select_related('system').prefetch_related('images')
        suministros = Suministro.objects.filter(asset=activo).select_related('item')
        all_items = Item.objects.all().only('id', 'name', 'reference', 'seccion', 'presentacion').order_by('name')

        context = {
            'activo': activo,
            'equipos': equipos,
            'fecha_actual': timezone.now().date(),
            'suministros': suministros,
            'all_items': all_items,
        }
        return render(request, self.template_name, context)
    

class AllAssetsEquipmentListView(View):
    template_name = 'inventory_management/all_equipment_list.html'

    def get(self, request):
        all_activos = Asset.objects.filter(show=True).select_related('supervisor', 'capitan').order_by('name')
        all_systems = System.objects.filter(asset__in=all_activos)
        equipos = Equipo.objects.filter(system__in=all_systems).order_by('name')
        suministros = Suministro.objects.filter(asset__in=all_activos).select_related('item')
        all_items = Item.objects.all().order_by('name')

        context = {
            'all_activos': all_activos,
            'equipos': equipos,
            'suministros': suministros,
            'all_items': all_items,
            'fecha_actual': timezone.now().date(),
        }
        return render(request, self.template_name, context)


@csrf_exempt
def public_equipo_detail(request, eq_code):
    """
    Vista pública (sin login) con la info completa de un equipo.
    Sin barra de navegación ni estilos de la app "base".
    """
    equipo = get_object_or_404(Equipo, code=eq_code)
    system = equipo.system  # model: System
    asset = system.asset    # model: Asset

    images = equipo.images.all().order_by('id')

    context = {
        'equipo': equipo,
        'system': system,
        'asset': asset,
        'images': images,
        # 'rutas': rutas, 'daily_fuel': daily_fuel, etc.
    }
    return render(request, 'inventory_management/public_equipo_detail.html', context)


def export_equipment_supplies(request, abbreviation):
    asset = get_object_or_404(Asset, abbreviation=abbreviation)
    systems = asset.system_set.all()
    equipos_qs = Equipo.objects.filter(system__in=systems).order_by('name')
    suministros_qs = Suministro.objects.filter(asset=asset).exclude(cantidad=0).select_related('item')

    equipment_data = []  # Construimos las filas con la información de cada equipo
    counter = 1
    for eq in equipos_qs:
        system_name = eq.system.name
        max_feature_length = 50  # define el límite deseado
        feature_display = (eq.feature[:max_feature_length] + '...') if eq.feature and len(eq.feature) > max_feature_length else (eq.feature or "")

        row = [
            counter,
            eq.code,
            eq.ubicacion or system_name,
            eq.get_tipo_display(),
            eq.name,
            eq.marca or "",
            eq.serial or "",
            eq.model or "",
            eq.estado.upper(),
            1,
            system_name,
            feature_display
        ]
        equipment_data.append(row)
        counter += 1


    for s in suministros_qs:
        equipment_data.append([
            counter,
            s.item.code,
            "",
            s.item.get_seccion_display(),
            s.item.name,
            s.item.reference or "",
            "",
            "",
            s.item.presentacion or "",
            str(s.cantidad), 
            "",
            ""
        ])
        counter += 1

    # 4. Crear el workbook y las hojas
    wb = Workbook()
    ws_equips = wb.active
    ws_equips.title = "Equipos"

    thin_border = Border(left=Side(style='thin', color="000000"), right=Side(style='thin', color="000000"), top=Side(style='thin', color="000000"), bottom=Side(style='thin', color="000000"))

    # Fila 1: Título
    ws_equips.merge_cells('A1:C3')
    ws_equips.merge_cells('D1:J3')
    ws_equips['D1'] = "ACTA DE INVENTARIO FÍSICO"
    ws_equips['D1'].font = Font(bold=True, size=16, name='Arial')
    ws_equips['D1'].alignment = Alignment(horizontal="center")

    # Fila 2: Formato, versión y fecha de actualización
    ws_equips.merge_cells('K1:L1')
    ws_equips['K1'] = "FORMATO: FR-SP-CM-25"
    ws_equips['K1'].font = Font(bold=True, name='Arial')
    ws_equips['K1'].alignment = Alignment(horizontal="center")

    # Fila 2: Formato, versión y fecha de actualización
    ws_equips.merge_cells('K2:L2')
    ws_equips['K2'] = "VERSION: 009"
    ws_equips['K2'].font = Font(bold=True, name='Arial')
    ws_equips['K2'].alignment = Alignment(horizontal="center")

    # Fila 2: Formato, versión y fecha de actualización
    ws_equips.merge_cells('K3:L3')
    ws_equips['K3'] = "FECHA DE ACTUALIZACIÓN: 13/09/2024"
    ws_equips['K3'].font = Font(bold=True, name='Arial')
    ws_equips['K3'].alignment = Alignment(horizontal="center")

    ws_equips.row_dimensions[1].height = 20
    ws_equips.row_dimensions[2].height = 20
    ws_equips.row_dimensions[3].height = 20

    # Fila 4-5: Datos del acta (por ejemplo, NRO DE ACTA, FECHA/HORA INICIO, CIUDAD, DEPENDENCIA)
    ws_equips.merge_cells('A4:F4')
    ws_equips['A4'] = "NRO DE ACTA:"
    ws_equips['A4'].font = Font(bold=True, name='Arial')
    ws_equips.merge_cells('G4:L4')
    ws_equips['G4'] = "FECHA Y HORA INICIO DE LA INSPECCIÓN:"
    ws_equips['G4'].font = Font(bold=True, name='Arial')
    ws_equips.merge_cells('A5:F5')
    ws_equips['A5'] = "CIUDAD:"
    ws_equips['A5'].font = Font(bold=True, name='Arial')
    ws_equips.merge_cells('G5:L5')
    ws_equips['G5'] = "DEPENDENCIA Y/O ÁREA:"
    ws_equips['G5'].font = Font(bold=True, name='Arial')

    ws_equips.merge_cells('A6:L6')
    ws_equips['A6'] = "LISTADO DE INVENTARIO"
    ws_equips['A6'].font = Font(bold=True, color="ffffff", name='Arial')
    ws_equips['A6'].fill = PatternFill(fill_type="solid", fgColor="0070c0")
    ws_equips['A6'].alignment = Alignment(horizontal="center", vertical="center")
    ws_equips.row_dimensions[6].height = 20

    # --- Agregar bordes al área de cabecera (título y datos generales) ---
    header_ranges = ["A1:C3", "D1:J3", "K1:L1", "K2:L2", "K3:L3", "A4:G4", "H4:L4", "A5:G5", "H5:L5", "A6:L6"]
    for merged_range in header_ranges:
        for row in ws_equips[merged_range]:
            for cell in row:
                cell.border = thin_border

    # === TABLA DE INVENTARIO (Equipos) ===
    # Encabezados originales para Equipos
    equipment_headers = [
        'ITEM', 'CÓDIGO', 'UBICACIÓN', 'TIPO', 'NOMBRE', 'MARCA',
        'SERIAL', 'MODELO', 'ESTADO', 'CANTIDAD', 'GRUPO CONSTRUCTIVO', 'OBSERVACIONES'
    ]
    start_row = 7  # Ajusta según lo necesites
    # Escribir encabezados con fondo azul (color 4d93d9) y texto blanco
    for col_num, header in enumerate(equipment_headers, 1):
        cell = ws_equips.cell(row=start_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="ffffff", name='Arial')
        cell.fill = PatternFill(fill_type="solid", fgColor="4d93d9")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Escribir los datos y asignar el borde negro a cada celda de la tabla
    current_row = start_row + 1
    for row_data in equipment_data:
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws_equips.cell(row=current_row, column=col_num)
            cell.value = cell_value
            cell.font = Font(name='Arial')
            cell.border = thin_border
        current_row += 1

    # Ajuste de anchos para la hoja de Equipos
    for idx, column_cells in enumerate(ws_equips.columns, 1):
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        col_letter = get_column_letter(idx)
        ws_equips.column_dimensions[col_letter].width = length + 1

    # === Sección de Observaciones y Cierre ===
    current_row = ws_equips.max_row + 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = " "

    for col in range(1, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    # === Sección de Observaciones y Cierre ===
    current_row = ws_equips.max_row + 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = "¿DURANTE LA INSPECCIÓN DEL INVENTARIO ANTERIORMENTE RELACIONADO SE PRESENTARON INCONSISTENCIAS?"
    cell.font = Font(bold=True, name='Arial')

    for col in range(1, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    current_row += 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = "DESCRIPCIÓN DE LAS INCONSISTENCIAS ENCONTRADAS (SI APLICA):"
    cell.alignment = Alignment(horizontal="left", vertical="top")
    cell.font = Font(bold=True, name='Arial')
    ws_equips.row_dimensions[current_row].height = 90

    for col in range(1, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    current_row += 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = "OBSERVACIONES GENERALES: (Describir hechos o situaciones relevantes a tener en cuenta)"
    cell.alignment = Alignment(horizontal="left", vertical="top")
    cell.font = Font(bold=True, name='Arial')
    ws_equips.row_dimensions[current_row].height = 90

    for col in range(1, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    current_row += 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = "CIERRE DE LA TOMA DE INVENTARIO FISICO"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(fill_type="solid", fgColor="4d93d9")
    cell.font = Font(bold=True, name='Arial')

    for col in range(1, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    current_row += 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = """Siendo las                              del día                    del mes                       del año, se da por finalizado la toma de inventario. En uso de la palabra los responsables que firman a continuación, manifiestan que las unidades anteriormente inspeccionadas existen en la ubicación y fueron contados durante el periodo de la toma fisica del inventario, y son en cantidad y estado, todos los que se encontraban. En caso de cualquier aclaración posterior a la toma del inventario se contaran con dichos documentos.  
    
    Se entrega copia firmada de la presente acta al responsable o delegado de la Dependencia Inspeccionada (Capitan, Jefe de Area, Etc)"
    """
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.font = Font(name='Arial')
    ws_equips.row_dimensions[current_row].height = 80

    for col in range(1, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    current_row += 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = "RESPONSABLES DEL CONTEO FISICO"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(fill_type="solid", fgColor="4d93d9")
    cell.font = Font(bold=True, name='Arial')

    for col in range(1, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border
    
    # Sección de firmas
    current_row = ws_equips.max_row + 1
    ws_equips.row_dimensions[current_row].height = 90

    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = "Firma de quien realiza la toma física"
    cell.alignment = Alignment(horizontal="center", vertical="bottom")
    cell.font = Font(bold=True, name='Arial')

    ws_equips.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=7)
    cell.value = "Firma - Responsable o delegado que acompaña en el conteo físico"
    cell.alignment = Alignment(horizontal="center", vertical="bottom")
    cell.font = Font(bold=True, name='Arial')

    for col in range(1, 7):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    for col in range(7, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    current_row = ws_equips.max_row + 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = "Nombre Completo"
    cell.alignment = Alignment(horizontal="left")
    cell.font = Font(bold=True, name='Arial')

    ws_equips.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=7)
    cell.value = "Nombre Completo"
    cell.alignment = Alignment(horizontal="left")
    cell.font = Font(bold=True, name='Arial')

    for col in range(1, 7):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    for col in range(7, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    current_row = ws_equips.max_row + 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = "Cédula No."
    cell.alignment = Alignment(horizontal="left")
    cell.font = Font(bold=True, name='Arial')

    ws_equips.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=7)
    cell.value = "Cédula No."
    cell.alignment = Alignment(horizontal="left")
    cell.font = Font(bold=True, name='Arial')

    for col in range(1, 7):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    for col in range(7, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    current_row += 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = "RESPONSABLES DE LAS DEPENDENCIAS"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(fill_type="solid", fgColor="4d93d9")
    cell.font = Font(bold=True, name='Arial')

    for col in range(1, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    current_row = ws_equips.max_row + 1
    ws_equips.row_dimensions[current_row].height = 90

    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = "Firma del Aprobador (Jefe de Abastecimiento y Logistica)"
    cell.alignment = Alignment(horizontal="center", vertical="bottom")
    cell.font = Font(bold=True, name='Arial')

    ws_equips.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=7)
    cell.value = "Firma - Responsable de la Dependencia Inspeccionada (Capitan, Jefe Area, Etc)"
    cell.alignment = Alignment(horizontal="center", vertical="bottom")
    cell.font = Font(bold=True, name='Arial')

    for col in range(1, 7):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    for col in range(7, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    current_row = ws_equips.max_row + 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = "Nombre Completo"
    cell.alignment = Alignment(horizontal="left")
    cell.font = Font(bold=True, name='Arial')

    ws_equips.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=7)
    cell.value = "Nombre Completo"
    cell.alignment = Alignment(horizontal="left")
    cell.font = Font(bold=True, name='Arial')

    for col in range(1, 7):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    for col in range(7, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    current_row = ws_equips.max_row + 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = "Cédula No."
    cell.alignment = Alignment(horizontal="left")
    cell.font = Font(bold=True, name='Arial')

    ws_equips.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=7)
    cell.value = "Cédula No."
    cell.alignment = Alignment(horizontal="left")
    cell.font = Font(bold=True, name='Arial')

    for col in range(1, 7):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    for col in range(7, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    current_row += 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = '“Este documento es propiedad de "SERPORT S.A.S" Se prohíbe su reproducción parcial o total.”'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(fill_type="solid", fgColor="4d93d9")
    cell.font = Font(name='Arial', italic=True, size=10)

    for col in range(1, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    # === Insertar la imagen desde la URL ===
    url = "https://hivik.s3.us-east-2.amazonaws.com/static/Logo.png"
    img_response = requests.get(url)
    if img_response.status_code == 200:
        image_data = BytesIO(img_response.content)
        # from openpyxl.drawing.image import Image
        image_data.name = "Logo.png"  # Asignar un nombre con la extensión adecuada
        image_data.seek(0)           # Reiniciar el puntero del archivo
        img = ExcelImage(image_data)
        img.width = 300  # Asignar dimensiones a la imagen
        img.height = 80
        ws_equips.add_image(img, "A1")
    else:
        print("No se pudo descargar la imagen.")

    output = io.BytesIO()  # Exportar a HttpResponse
    wb.save(output)
    output.seek(0)

    filename = f"{asset.abbreviation}_Equipos_y_Suministros.xlsx"
    response = HttpResponse(output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename={filename}'
    return response


class DarBajaCreateView(LoginRequiredMixin, CreateView):
    model = DarBaja
    form_class = DarBajaForm
    template_name = 'inventory_management/dar_baja_form.html'

    def get_equipo(self):
        equipo_code = self.kwargs['equipo_code']
        return get_object_or_404(Equipo, code=equipo_code)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        equipo = self.get_equipo()
        context['equipo'] = equipo

        if self.request.method == 'POST':  # Instanciar el segundo formulario (evidencias + firmas)
            context['upload_form'] = UploadEvidenciasYFirmasForm(self.request.POST, self.request.FILES)
        else:
            context['upload_form'] = UploadEvidenciasYFirmasForm()
        return context

    def form_valid(self, form):
        equipo = self.get_equipo()
        user = self.request.user
        form.instance.equipo = equipo
        form.instance.reporter = user.get_full_name() or user.username
        form.instance.activo = f"{equipo.system.asset.name} - {equipo.system.name}"

        # Guardamos de forma normal
        response = super().form_valid(form)

        # 2) Procesar upload_form
        upload_form = self.get_context_data()['upload_form']
        if not upload_form.is_valid():
            messages.error(self.request, "Error en el formulario de subida de archivos.")
            self.object.delete()
            return self.form_invalid(form)

        # Tomar subidas de firma
        fr_file = upload_form.cleaned_data.get('firma_responsable_file')
        fa_file = upload_form.cleaned_data.get('firma_autorizado_file')

        # Tomar data base64 de canvas
        firma_responsable_data = self.request.POST.get('firma_responsable_data', '')
        firma_autorizado_data = self.request.POST.get('firma_autorizado_data', '')

        # Chequeo => O suben archivo o dibujan la firma
        if not fr_file and not firma_responsable_data:
            messages.error(self.request, "Debe proporcionar la firma Responsable (archivo o dibujo).")
            self.object.delete()
            return self.form_invalid(form)

        if not fa_file and not firma_autorizado_data:
            messages.error(self.request, "Debe proporcionar la firma Autorizado (archivo o dibujo).")
            self.object.delete()
            return self.form_invalid(form)

        # Firma Responsable
        if fr_file:
            if not fr_file.content_type.startswith('image/'):
                messages.error(self.request, "Archivo de firma responsable no es válido.")
                self.object.delete()
                return self.form_invalid(form)
            self.object.firma_responsable = fr_file
            self.object.save(update_fields=['firma_responsable'])
        else:
            # Canvas base64
            fmt, imgstr = firma_responsable_data.split(';base64,')
            ext = fmt.split('/')[-1]
            data = ContentFile(base64.b64decode(imgstr))
            self.object.firma_responsable.save(f'firma_responsable_{self.object.pk}.{ext}', data, save=True)

        # Firma Autorizado
        if fa_file:
            if not fa_file.content_type.startswith('image/'):
                messages.error(self.request, "Archivo de firma autorizado no es válido.")
                self.object.delete()
                return self.form_invalid(form)
            self.object.firma_autorizado = fa_file
            self.object.save(update_fields=['firma_autorizado'])
        else:
            fmt, imgstr = firma_autorizado_data.split(';base64,')
            ext = fmt.split('/')[-1]
            data = ContentFile(base64.b64decode(imgstr))
            self.object.firma_autorizado.save(f'firma_autorizado_{self.object.pk}.{ext}', data, save=True)

        # 3) Evidencias => multiple
        evidences = upload_form.cleaned_data['file_field']  # lista de archivos
        for file_obj in evidences:
            if not file_obj.content_type.startswith('image/'):
                messages.error(self.request, "Uno de los archivos de evidencia no es una imagen válida.")
                self.object.delete()
                return self.form_invalid(form)
            Image.objects.create(image=file_obj, darbaja=self.object)

        # 4) Eliminar rutinas/hours/etc
        self.do_equipo_cleanup(equipo)

        # 5) Mover equipo a system 445
        new_system = System.objects.get(id=445)
        old_system = equipo.system
        equipo.system = new_system
        equipo.save()
        messages.success(self.request, f'El equipo ha sido trasladado al sistema {new_system.name}.')

        # 6) Generar PDF (opcional)
        self.generate_pdf()
        return redirect(self.get_success_url())  # 7) Redirigir => 'activos/<str:abbreviation>/equipos/'

    def do_equipo_cleanup(self, equipo):
        # Eliminar rutinas, hours, etc.
        ...
        # (Idéntico a tu lógica actual)

    def generate_pdf(self):
        context = {
            'dar_baja': self.object,
            'equipo': self.object.equipo
        }
        pdf = render_to_pdf('inventory_management/dar_baja_pdf.html', context)
        if pdf:
            # Podrías guardarlo en disco, o en la base de datos, 
            # o ignorarlo si solo deseas generarlo internamente
            filename = f'dar_baja_{self.object.pk}.pdf'
            # Ejemplo: guardarlo en un directorio local (opcional)
            # with open(f'/tmp/{filename}', 'wb') as f:
            #     f.write(pdf.getvalue())
        else:
            messages.warning(self.request, 'No se pudo generar el PDF.')

    def get_success_url(self):
        # Redirigir a: path('activos/<str:abbreviation>/equipos/', name='asset_equipment_list'),
        # old_system.asset.abbreviation => => 'inv:asset_equipment_list'
        abbreviation = self.object.equipo.system.asset.abbreviation
        return reverse('inv:asset_equipment_list', kwargs={'abbreviation': abbreviation})


def create_supply_view(request, abbreviation):
    """
    Crea un nuevo Suministro asociado a un Activo.
    Opcionalmente crea un nuevo Item si el usuario marcó "is_new_item == 1".
    """
    asset = get_object_or_404(Asset, abbreviation=abbreviation)

    if request.method == 'POST':
        is_new_item = request.POST.get('is_new_item', '0')
        cantidad_str = request.POST.get('cantidad')

        # Validar cantidad
        if not cantidad_str:
            messages.error(request, "Debe especificar la cantidad.")
            return redirect('inv:asset_equipment_list', abbreviation=abbreviation)

        try:
            cantidad = float(cantidad_str)
        except ValueError:
            messages.error(request, "Cantidad inválida.")
            return redirect('inv:asset_equipment_list', abbreviation=abbreviation)

        # Caso A: Crear artículo nuevo
        if is_new_item == '1':
            form_data = {
                'name': request.POST.get('new_item_name'),
                'reference': request.POST.get('new_item_reference'),
                'presentacion': request.POST.get('new_item_presentacion'),
                'code': request.POST.get('new_item_code'),
                'seccion': request.POST.get('new_item_seccion'),
                'unit_price': request.POST.get('new_item_unit_price') or 0.00,
            }
            files_data = {}
            if 'new_item_imagen' in request.FILES:
                files_data['imagen'] = request.FILES['new_item_imagen']

            item_form = ItemForm(form_data, files_data)
            if item_form.is_valid():
                new_item = item_form.save(commit=False)
                if request.user.is_authenticated:
                    new_item.modified_by = request.user
                new_item.save()
                item = new_item
                messages.success(request, f"Artículo '{item.name}' creado correctamente.")
            else:
                messages.error(request, f"Error al crear el artículo: {item_form.errors}")
                return redirect('inv:asset_equipment_list', abbreviation=abbreviation)

        # Caso B: Se usa un artículo existente
        else:
            item_id = request.POST.get('item_id')
            if not item_id:
                messages.error(request, "Debe seleccionar un artículo existente o crear uno nuevo.")
                return redirect('inv:asset_equipment_list', abbreviation=abbreviation)

            try:
                item = Item.objects.get(id=item_id)
            except Item.DoesNotExist:
                messages.error(request, "El artículo seleccionado no existe.")
                return redirect('inv:asset_equipment_list', abbreviation=abbreviation)

        # Crear el Suministro
        Suministro.objects.create(
            item=item,
            cantidad=cantidad,
            asset=asset
        )
        messages.success(request, f"Se ha creado un nuevo suministro de '{item.name}' para {asset.name}.")
        return redirect('inv:asset_equipment_list', abbreviation=abbreviation)

    # Si no es POST => redirigir sin hacer nada
    return redirect('inv:asset_equipment_list', abbreviation=abbreviation)


class AssetInventoryBaseView(LoginRequiredMixin, View):
    template_name = 'got/assets/asset_inventory_report.html'
    keyword_filter = None
    keyword_filter2 = None

    def get(self, request, abbreviation):
        asset = get_object_or_404(Asset, abbreviation=abbreviation)
        suministros = get_suministros(asset, self.keyword_filter)

        transacciones_historial = self.get_transacciones_historial(asset)
        ultima_fecha_transaccion = transacciones_historial.aggregate(Max('fecha'))['fecha__max'] or "---"
        context = self.get_context_data(request, asset, suministros, transacciones_historial, ultima_fecha_transaccion)
        return render(request, self.template_name, context)

    def post(self, request, abbreviation):
        asset = get_object_or_404(Asset, abbreviation=abbreviation)
        suministros = get_suministros(asset, self.keyword_filter)
        action = request.POST.get('action', '')

        if action == 'download_excel':
            headers_mapping = self.get_headers_mapping()
            filename = self.get_filename()
            transacciones_historial = self.get_transacciones_historial(asset)
            return generate_excel(transacciones_historial, headers_mapping, filename)

        elif action == 'transfer_suministro':
            suministro_id = request.POST.get('transfer_suministro_id')
            suministro = get_object_or_404(Suministro, id=suministro_id, asset=asset)
            transfer_fecha_str = request.POST.get('transfer_fecha', '')

            if not re.match(r'^\d{4}-\d{2}-\d{2}$', transfer_fecha_str):
                messages.error(request, "Fecha inválida de transferencia: usa el formato YYYY-MM-DD.")
                return self.redirect_with_next(request)
            
            try:
                transfer_fecha = datetime.strptime(transfer_fecha_str, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, "Fecha inválida: no se pudo interpretar el valor ingresado.")
                return self.redirect_with_next(request)
            
            result = handle_transfer(
                request,
                asset,
                suministro,
                request.POST.get('transfer_cantidad', '0') or '0',
                request.POST.get('destination_asset_id'),
                request.POST.get('transfer_motivo', ''),
                transfer_fecha_str 
            )
            if isinstance(result, HttpResponse):
                return result
            else:
                return self.redirect_with_next(request)

        elif action == 'update_inventory':
            motivo_global = ''
            fecha_reporte_str = request.POST.get('fecha_reporte', timezone.now().date().strftime('%Y-%m-%d'))

            print(fecha_reporte_str)
            # 1) Verificar si coincide con el patrón YYYY-MM-DD
            if not re.match(r'^\d{4}-\d{2}-\d{2}$', fecha_reporte_str):
                messages.error(request, "Fecha inválida: usa el formato YYYY-MM-DD.")
                return self.redirect_with_next(request)
            
            try:
                fecha_reporte = datetime.strptime(fecha_reporte_str, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, "Fecha inválida: no se pudo interpretar el valor ingresado.")
                return self.redirect_with_next(request)

            # 3) Comparar con la fecha actual
            if fecha_reporte > timezone.now().date():
                messages.error(request, 'La fecha del reporte no puede ser mayor a la fecha actual.')
                return self.redirect_with_next(request)

            operation = Operation.objects.filter(asset=asset, confirmado=True, start__lte=fecha_reporte, end__gte=fecha_reporte).first()

            if operation:
                motivo_global = operation.proyecto

            result = handle_inventory_update(request, asset, suministros, motivo_global)
            if isinstance(result, HttpResponse):
                return result
            else:
                return self.redirect_with_next(request)

        elif action == 'add_suministro' and request.user.has_perm('got.can_add_supply'):
            item_id = request.POST.get('item_id')
            item = get_object_or_404(Item, id=item_id)
            suministro_existente = Suministro.objects.filter(asset=asset, item=item).exists()
            if suministro_existente:
                messages.error(request, f'El suministro para el artículo "{item.name}" ya existe en este asset.')
            else:
                Suministro.objects.create(item=item, cantidad=Decimal('0.00'), asset=asset)
                messages.success(request, f'Suministro para "{item.name}" creado exitosamente.')
            return self.redirect_with_next(request)

        else:
            messages.error(request, 'Acción no reconocida.')
            return self.redirect_with_next(request)

    def get_context_data(self, request, asset, suministros, transacciones_historial, ultima_fecha_transaccion):
        motonaves = Asset.objects.filter(show=True)
        available_items = Item.objects.exclude(id__in=suministros.values_list('item_id', flat=True))
        users_en_historial = transacciones_historial.values_list('user', flat=True).distinct()
        users_unicos = User.objects.filter(username__in=users_en_historial).order_by('username')
        context = {
            'asset': asset,
            'suministros': suministros,
            'ultima_fecha_transaccion': ultima_fecha_transaccion,
            'transacciones_historial': transacciones_historial,
            'fecha_actual': timezone.now().date(),
            'motonaves': motonaves,
            'available_items': available_items,
            'articulos_unicos': self.get_articulos_unicos(transacciones_historial),
            'users_unicos': users_unicos, 
        }
        return context
    
    def get_articulos_unicos(self, transacciones_historial):
        items_en_historial = set()
        for t in transacciones_historial:
            if t.suministro and t.suministro.item:
                items_en_historial.add(t.suministro.item)
        articulos_unicos = sorted(items_en_historial, key=lambda x: (x.name.lower(), x.reference.lower() if x.reference else ""))
        return articulos_unicos

    def get_headers_mapping(self):
        return {}

    def get_filename(self):
        return 'export.xlsx'
    
    def get_transacciones_historial(self, asset):
        transacciones = Transaction.objects.filter(Q(suministro__asset=asset) | Q(suministro_transf__asset=asset))
        if self.keyword_filter:
            transacciones = transacciones.filter(self.keyword_filter2)
        transacciones_historial = transacciones.order_by('-fecha')
        return transacciones_historial
    
    def redirect_with_next(self, request):
        next_url = request.GET.get('next', '')
        if next_url:
            return redirect(next_url)
        return redirect(request.path)


class AssetSuministrosReportView(AssetInventoryBaseView):
    keyword_filter = Q(item__name__icontains='Combustible') | Q(item__name__icontains='Aceite') | Q(item__name__icontains='Filtro')

    keyword_filter2 = (
        Q(suministro__item__name__icontains='Combustible')
        | Q(suministro__item__name__icontains='Aceite')
        | Q(suministro__item__name__icontains='Filtro')
    )

    def get_context_data(self, request, asset, suministros, transacciones_historial, ultima_fecha_transaccion):
        context = super().get_context_data(request, asset, suministros, transacciones_historial, ultima_fecha_transaccion)
        # Agrupar suministros por presentación
        suministros = suministros.order_by('item__presentacion')
        grouped_suministros = {}
        for key, group in groupby(suministros, key=attrgetter('item.presentacion')):
            grouped_suministros[key] = list(group)
        context['grouped_suministros'] = grouped_suministros
        context['group_by'] = 'presentacion'

        items_en_historial = set()
        for t in transacciones_historial:
            if t.suministro and t.suministro.item:
                items_en_historial.add(t.suministro.item)

        # Convertir a lista y ordenar (opcional, según la preferencia):
        articulos_unicos = sorted(items_en_historial, key=lambda x: (x.name.lower(), x.reference.lower() if x.reference else ""))

        context['articulos_unicos'] = articulos_unicos
        return context

    def get_headers_mapping(self):
        return {
            'fecha': 'Fecha',
            'suministro__item__presentacion': 'Presentación',
            'suministro__item__name': 'Artículo',
            'cant': 'Cantidad',
            'tipo': 'Tipo',
            'user': 'Usuario',
            'motivo': 'Motivo',
            'cant_report': 'Cantidad Reportada',
        }

    def get_filename(self):
        return 'historial_suministros.xlsx'


class AssetInventarioReportView(AssetInventoryBaseView):
    keyword_filter = ~(Q(item__name__icontains='Combustible') | Q(item__name__icontains='Aceite') | Q(item__name__icontains='Filtro'))

    keyword_filter2 = ~(
        Q(suministro__item__name__icontains='Combustible')
        | Q(suministro__item__name__icontains='Aceite')
        | Q(suministro__item__name__icontains='Filtro')
    )

    def get_headers_mapping(self):
        return {
            'fecha': 'Fecha',
            'suministro__item__seccion': 'Categoría',
            'suministro__item__name': 'Artículo',
            'cant': 'Cantidad',
            'tipo': 'Tipo',
            'user': 'Usuario',
            'motivo': 'Motivo',
            'cant_report': 'Cantidad Reportada',
        }

    def get_filename(self):
        return 'historial_inventario.xlsx'

    def get_context_data(self, request, asset, suministros, transacciones_historial, ultima_fecha_transaccion):
        context = super().get_context_data(request, asset, suministros, transacciones_historial, ultima_fecha_transaccion)
        suministros = suministros.order_by('item__seccion')
        grouped_suministros = {}
        for key, group in groupby(suministros, key=attrgetter('item.seccion')):
            grouped_suministros[key] = list(group)
        context['grouped_suministros'] = grouped_suministros
        context['group_by'] = 'seccion'
        secciones_dict = dict(Item.SECCION)
        context['secciones_dict'] = secciones_dict

        items_en_historial = set()
        for t in transacciones_historial:
            if t.suministro and t.suministro.item:
                items_en_historial.add(t.suministro.item)

        # Convertir a lista y ordenar (opcional, según la preferencia):
        articulos_unicos = sorted(items_en_historial, key=lambda x: (x.name.lower(), x.reference.lower() if x.reference else ""))

        context['articulos_unicos'] = articulos_unicos
        return context
    

@permission_required('got.delete_transaction', raise_exception=True)
def delete_transaction(request, transaction_id):
    transaccion = get_object_or_404(Transaction, id=transaction_id)
    next_url = request.POST.get('next', request.META.get('HTTP_REFERER', '/'))

    if request.method == 'POST':
        result = handle_delete_transaction(request, transaccion)
        if isinstance(result, HttpResponse):
            return result
        else:
            return redirect(next_url)
    else:
        return redirect(next_url)
    

# @permission_required('got.delete_transaction', raise_exception=True)
def delete_sumi(request, sumi_id):
    sumi = get_object_or_404(Suministro, id=sumi_id)
    next_url = request.POST.get('next', request.META.get('HTTP_REFERER', '/'))

    if request.method == 'POST':
        sumini = sumi.item
        sumi.delete()
        messages.success(request, f"Eliminado correctamente el articulo {sumini}")
        return redirect(next_url)
    else:
        messages.success(request, f"se intento y no fuciono")
        return redirect(next_url)
    

def export_historial_pdf(request, abbreviation):
    """
    Genera un PDF con los registros de transacciones,
    filtrados por rango de fechas y por artículos seleccionados,
    usando la plantilla base de PDF (`pdf_template.html`) y la función render_to_pdf.
    """
    asset = get_object_or_404(Asset, abbreviation=abbreviation)

    if request.method == 'POST':
        # 1) Obtener fechas (inicio y fin)
        fecha_inicio_str = request.POST.get('fecha_inicio', '')
        fecha_fin_str = request.POST.get('fecha_fin', '')

        # Parsear a objeto date, si no vienen => None
        fecha_inicio = None
        fecha_fin = None
        if fecha_inicio_str:
            try:
                fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            except:
                pass
        if fecha_fin_str:
            try:
                fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
            except:
                pass

        # 2) Artículos seleccionados (lista de IDs)
        items_seleccionados = request.POST.getlist('items_seleccionados')
        if not items_seleccionados:
            messages.error(request, "Debe seleccionar al menos un artículo para generar el PDF.")
            return redirect('inv:asset_inventario_report', abbreviation=abbreviation)

        # 3) Obtener transacciones base
        transacciones = Transaction.objects.filter(
            Q(suministro__asset=asset) | Q(suministro_transf__asset=asset)
        ).order_by('-fecha')

        # 4) Filtrar por rango de fechas
        if fecha_inicio and fecha_fin:
            transacciones = transacciones.filter(
                fecha__range=[fecha_inicio, fecha_fin]
            )
        elif fecha_inicio:
            transacciones = transacciones.filter(fecha__gte=fecha_inicio)
        elif fecha_fin:
            transacciones = transacciones.filter(fecha__lte=fecha_fin)

        # 5) Filtrar por items
        #   items_seleccionados => ID del item
        transacciones = transacciones.filter(
            Q(suministro__item__id__in=items_seleccionados) |
            Q(suministro_transf__item__id__in=items_seleccionados)
        )
        # 6) Obtener los suministros seleccionados
        suministros_seleccionados = Suministro.objects.filter(
            item__id__in=items_seleccionados,
            asset=asset
        )

        # 7) Calcular resúmenes por suministro
        suministros_summary = []
        for suministro in suministros_seleccionados:
            if fecha_inicio:
                trans_before_start = Transaction.objects.filter(suministro=suministro, fecha__lte=fecha_inicio).order_by('-fecha').first()
                cantidad_inicial = trans_before_start.cant_report if trans_before_start and trans_before_start.cant_report else Decimal('0.00')
            else:
                cantidad_inicial = Decimal('0.00')

            total_consumido = transacciones.filter(suministro=suministro, tipo='c').aggregate(total=Sum('cant'))['total'] or Decimal('0.00') # Total consumido: sum of 'c' transactions in the period
            total_ingresado_base = transacciones.filter(suministro=suministro, tipo='i').aggregate(total=Sum('cant'))['total'] or Decimal('0.00') # Total ingresado: suma de transacciones de tipo 'i'
            incoming_transfer = transacciones.filter(suministro_transf=suministro, tipo='t').aggregate(total=Sum('cant'))['total'] or Decimal('0.00') # Sumar también las transferencias entrantes (donde este suministro es el destino)
            total_ingresado = total_ingresado_base + incoming_transfer
            total_transfer_out = transacciones.filter(suministro=suministro, tipo='t').aggregate(total=Sum('cant'))['total'] or Decimal('0.00') # Total transferido a otros: suma de transferencias salientes (donde este suministro es el emisor)

            if fecha_fin: # Cantidad final: latest 'cant_report' before 'fecha_fin'
                trans_before_end = Transaction.objects.filter(suministro=suministro, fecha__lte=fecha_fin).order_by('-fecha').first()
                cantidad_final = trans_before_end.cant_report if trans_before_end and trans_before_end.cant_report else Decimal('0.00')
            else:
                # Si no hay fecha_fin, usar la última cantidad reportada hasta hoy
                trans_before_end = Transaction.objects.filter(suministro=suministro, fecha__lte=date.today()).order_by('-fecha').first()
                cantidad_final = trans_before_end.cant_report if trans_before_end and trans_before_end.cant_report else Decimal('0.00')

            suministros_summary.append({
                'suministro': suministro,
                'cantidad_inicial': cantidad_inicial,
                'total_consumido': total_consumido,
                'total_ingresado': total_ingresado,
                'total_transfer_out': total_transfer_out,
                'cantidad_final': cantidad_final,
            })

        # 8) Obtener los artículos seleccionados para el resumen en el header
        articulos_seleccionados = Item.objects.filter(id__in=items_seleccionados)

        data_context = {
            'asset': asset,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'transacciones': transacciones,
            'fecha_hoy': date.today(),
            'items': articulos_seleccionados,
            'suministros_summary': suministros_summary,
        }

        # 7) Renderizar con la plantilla PDF
        pdf = render_to_pdf('inventory_management/historial_pdf.html', data_context)
        if pdf:
            filename = f"Historial_{asset.abbreviation}.pdf"
            # inline => para abrir en navegador, attachment => para forzar descarga
            content = f"inline; filename='{filename}'"
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = content
            return response
        else:
            return HttpResponse("Error al generar el PDF", status=400)

    # Si no es POST, redirigir
    return redirect('inv:asset_inventario_report', abbreviation=abbreviation)


@login_required
def transferir_equipo(request, equipo_id):
    equipo = get_object_or_404(Equipo, pk=equipo_id)  # Equipo a trasladar
    next_url = request.GET.get('next') or request.POST.get('next') or None
    related_equipos = equipo.related_with.all()

    if request.method == 'POST':
        form = TransferenciaForm(request.POST)

        transfer_related_ids = request.POST.getlist('transfer_related')
        if form.is_valid():
            nuevo_sistema = form.cleaned_data['destino']
            observaciones = form.cleaned_data['observaciones']
            receptor = form.cleaned_data['receptor']

            sistema_origen = equipo.system      
            equipo.system = nuevo_sistema  # Actualiza el sistema del equipo principal   
            equipo.save()

            rutas = Ruta.objects.filter(equipo=equipo)  # Actualiza las rutas asociadas al equipo
            for ruta in rutas:
                ruta.system = nuevo_sistema
                ruta.save()

            transferred_related_names = []  # Procesa los equipos relacionados
            for rel in related_equipos:
                if str(rel.code) in transfer_related_ids:
                    rel.system = nuevo_sistema  # Se transfiere el equipo relacionado
                    rel.save()
                    transferred_related_names.append(rel.name)
                else:
                    rel.related = None  # Se rompe la relación (se desvincula)
                    rel.save()

            # Anexa la lista de equipos relacionados transferidos a las observaciones
            if transferred_related_names:
                observaciones += "\nEquipos relacionados transferidos: " + ", ".join(transferred_related_names)

            # Crea el registro de Transferencia
            # transferencia = Transferencia.objects.create(
            transferencia = Transference.objects.create(
                equipo=equipo,
                responsable=f"{request.user.first_name} {request.user.last_name}",
                receptor=receptor,
                origen=sistema_origen,
                destino=nuevo_sistema,
                observaciones=observaciones
            )

            enviar_correo_transferencia(transferencia, equipo, sistema_origen, nuevo_sistema, observaciones)  # Enviar correo de notificación (ver sección siguiente)

            messages.success(request, "Transferencia realizada exitosamente.")
            if next_url:
                return redirect(next_url)
            else:
                return redirect(sistema_origen.get_absolute_url())
    else:
        form = TransferenciaForm()

    context = {
        'form': form,
        'equipo': equipo,
        'related_equipos': related_equipos,
        'next_url': next_url
    }
    return render(request, 'inventory_management/transferencias.html', context)


'EXPERIMENTAL VIEWS'
class ItemManagementView(generic.TemplateView):
    template_name = 'inventory_management/item_management.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['items'] = Item.objects.all()
        context['form'] = ItemForm()
        return context

    def post(self, request, *args, **kwargs):
        form = ItemForm(request.POST, request.FILES)
        if form.is_valid():
            item = form.save(commit=False)
            item.modified_by = request.user
            form.save()
            return redirect(reverse('inv:item_management'))

        context = self.get_context_data(**kwargs)
        context['form'] = form
        return self.render_to_response(context)


def edit_item(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    if request.method == 'POST':
        form = ItemForm(request.POST, request.FILES, instance=item)
        if form.is_valid():
            form.save()
            return redirect(reverse('inv:item_management'))
    else:
        form = ItemForm(instance=item)

    return render(request, 'got/solicitud/edit_item.html', {'form': form, 'item': item})