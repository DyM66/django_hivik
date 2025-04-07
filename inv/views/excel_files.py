from django.shortcuts import get_object_or_404
from django.http import HttpResponseRedirect, HttpResponse

import io
import requests
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as ExcelImage

from got.models import Asset, Equipo
from inv.models import Suministro


def export_equipment_supplies(request, abbreviation):
    asset = get_object_or_404(Asset, abbreviation=abbreviation)
    systems = asset.system_set.all()
    equipos_qs = Equipo.objects.filter(system__in=systems).order_by('name')
    suministros_qs = Suministro.objects.filter(asset=asset).exclude(cantidad=0).select_related('item')

    equipment_data = []  # Construimos las filas con la información de cada equipo
    counter = 1
    for eq in equipos_qs:
        system_name = eq.system.name
        max_feature_length = 50  # define el límite deseado
        feature_display = (eq.feature[:max_feature_length] + '...') if eq.feature and len(eq.feature) > max_feature_length else (eq.feature or "")

        row = [
            counter,
            eq.code,
            eq.ubicacion or system_name,
            eq.get_tipo_display(),
            eq.name,
            eq.marca or "",
            eq.serial or "",
            eq.model or "",
            eq.estado.upper(),
            1,
            system_name,
            feature_display
        ]
        equipment_data.append(row)
        counter += 1


    for s in suministros_qs:
        equipment_data.append([
            counter,
            s.item.code,
            "",
            s.item.get_seccion_display(),
            s.item.name,
            s.item.reference or "",
            "",
            "",
            s.item.presentacion or "",
            str(s.cantidad), 
            "",
            ""
        ])
        counter += 1

    # 4. Crear el workbook y las hojas
    wb = Workbook()
    ws_equips = wb.active
    ws_equips.title = "Equipos"

    thin_border = Border(left=Side(style='thin', color="000000"), right=Side(style='thin', color="000000"), top=Side(style='thin', color="000000"), bottom=Side(style='thin', color="000000"))

    # Fila 1: Título
    ws_equips.merge_cells('A1:C3')
    ws_equips.merge_cells('D1:J3')
    ws_equips['D1'] = "ACTA DE INVENTARIO FÍSICO"
    ws_equips['D1'].font = Font(bold=True, size=16, name='Arial')
    ws_equips['D1'].alignment = Alignment(horizontal="center")

    # Fila 2: Formato, versión y fecha de actualización
    ws_equips.merge_cells('K1:L1')
    ws_equips['K1'] = "FORMATO: FR-SP-CM-25"
    ws_equips['K1'].font = Font(bold=True, name='Arial')
    ws_equips['K1'].alignment = Alignment(horizontal="center")

    # Fila 2: Formato, versión y fecha de actualización
    ws_equips.merge_cells('K2:L2')
    ws_equips['K2'] = "VERSION: 009"
    ws_equips['K2'].font = Font(bold=True, name='Arial')
    ws_equips['K2'].alignment = Alignment(horizontal="center")

    # Fila 2: Formato, versión y fecha de actualización
    ws_equips.merge_cells('K3:L3')
    ws_equips['K3'] = "FECHA DE ACTUALIZACIÓN: 13/09/2024"
    ws_equips['K3'].font = Font(bold=True, name='Arial')
    ws_equips['K3'].alignment = Alignment(horizontal="center")

    ws_equips.row_dimensions[1].height = 20
    ws_equips.row_dimensions[2].height = 20
    ws_equips.row_dimensions[3].height = 20

    # Fila 4-5: Datos del acta (por ejemplo, NRO DE ACTA, FECHA/HORA INICIO, CIUDAD, DEPENDENCIA)
    ws_equips.merge_cells('A4:F4')
    ws_equips['A4'] = "NRO DE ACTA:"
    ws_equips['A4'].font = Font(bold=True, name='Arial')
    ws_equips.merge_cells('G4:L4')
    ws_equips['G4'] = "FECHA Y HORA INICIO DE LA INSPECCIÓN:"
    ws_equips['G4'].font = Font(bold=True, name='Arial')
    ws_equips.merge_cells('A5:F5')
    ws_equips['A5'] = "CIUDAD:"
    ws_equips['A5'].font = Font(bold=True, name='Arial')
    ws_equips.merge_cells('G5:L5')
    ws_equips['G5'] = "DEPENDENCIA Y/O ÁREA:"
    ws_equips['G5'].font = Font(bold=True, name='Arial')

    ws_equips.merge_cells('A6:L6')
    ws_equips['A6'] = "LISTADO DE INVENTARIO"
    ws_equips['A6'].font = Font(bold=True, color="ffffff", name='Arial')
    ws_equips['A6'].fill = PatternFill(fill_type="solid", fgColor="0070c0")
    ws_equips['A6'].alignment = Alignment(horizontal="center", vertical="center")
    ws_equips.row_dimensions[6].height = 20

    # --- Agregar bordes al área de cabecera (título y datos generales) ---
    header_ranges = ["A1:C3", "D1:J3", "K1:L1", "K2:L2", "K3:L3", "A4:G4", "H4:L4", "A5:G5", "H5:L5", "A6:L6"]
    for merged_range in header_ranges:
        for row in ws_equips[merged_range]:
            for cell in row:
                cell.border = thin_border

    # === TABLA DE INVENTARIO (Equipos) ===
    # Encabezados originales para Equipos
    equipment_headers = [
        'ITEM', 'CÓDIGO', 'UBICACIÓN', 'TIPO', 'NOMBRE', 'MARCA',
        'SERIAL', 'MODELO', 'ESTADO', 'CANTIDAD', 'GRUPO CONSTRUCTIVO', 'OBSERVACIONES'
    ]
    start_row = 7  # Ajusta según lo necesites
    # Escribir encabezados con fondo azul (color 4d93d9) y texto blanco
    for col_num, header in enumerate(equipment_headers, 1):
        cell = ws_equips.cell(row=start_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="ffffff", name='Arial')
        cell.fill = PatternFill(fill_type="solid", fgColor="4d93d9")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Escribir los datos y asignar el borde negro a cada celda de la tabla
    current_row = start_row + 1
    for row_data in equipment_data:
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws_equips.cell(row=current_row, column=col_num)
            cell.value = cell_value
            cell.font = Font(name='Arial')
            cell.border = thin_border
        current_row += 1

    # Ajuste de anchos para la hoja de Equipos
    for idx, column_cells in enumerate(ws_equips.columns, 1):
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        col_letter = get_column_letter(idx)
        ws_equips.column_dimensions[col_letter].width = length + 1

    # === Sección de Observaciones y Cierre ===
    current_row = ws_equips.max_row + 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = " "

    for col in range(1, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    # === Sección de Observaciones y Cierre ===
    current_row = ws_equips.max_row + 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = "¿DURANTE LA INSPECCIÓN DEL INVENTARIO ANTERIORMENTE RELACIONADO SE PRESENTARON INCONSISTENCIAS?"
    cell.font = Font(bold=True, name='Arial')

    for col in range(1, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    current_row += 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = "DESCRIPCIÓN DE LAS INCONSISTENCIAS ENCONTRADAS (SI APLICA):"
    cell.alignment = Alignment(horizontal="left", vertical="top")
    cell.font = Font(bold=True, name='Arial')
    ws_equips.row_dimensions[current_row].height = 90

    for col in range(1, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    current_row += 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = "OBSERVACIONES GENERALES: (Describir hechos o situaciones relevantes a tener en cuenta)"
    cell.alignment = Alignment(horizontal="left", vertical="top")
    cell.font = Font(bold=True, name='Arial')
    ws_equips.row_dimensions[current_row].height = 90

    for col in range(1, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    current_row += 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = "CIERRE DE LA TOMA DE INVENTARIO FISICO"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(fill_type="solid", fgColor="4d93d9")
    cell.font = Font(bold=True, name='Arial')

    for col in range(1, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    current_row += 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = """Siendo las                              del día                    del mes                       del año, se da por finalizado la toma de inventario. En uso de la palabra los responsables que firman a continuación, manifiestan que las unidades anteriormente inspeccionadas existen en la ubicación y fueron contados durante el periodo de la toma fisica del inventario, y son en cantidad y estado, todos los que se encontraban. En caso de cualquier aclaración posterior a la toma del inventario se contaran con dichos documentos.  
    
    Se entrega copia firmada de la presente acta al responsable o delegado de la Dependencia Inspeccionada (Capitan, Jefe de Area, Etc)"
    """
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.font = Font(name='Arial')
    ws_equips.row_dimensions[current_row].height = 80

    for col in range(1, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    current_row += 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = "RESPONSABLES DEL CONTEO FISICO"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(fill_type="solid", fgColor="4d93d9")
    cell.font = Font(bold=True, name='Arial')

    for col in range(1, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border
    
    # Sección de firmas
    current_row = ws_equips.max_row + 1
    ws_equips.row_dimensions[current_row].height = 90

    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = "Firma de quien realiza la toma física"
    cell.alignment = Alignment(horizontal="center", vertical="bottom")
    cell.font = Font(bold=True, name='Arial')

    ws_equips.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=7)
    cell.value = "Firma - Responsable o delegado que acompaña en el conteo físico"
    cell.alignment = Alignment(horizontal="center", vertical="bottom")
    cell.font = Font(bold=True, name='Arial')

    for col in range(1, 7):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    for col in range(7, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    current_row = ws_equips.max_row + 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = "Nombre Completo"
    cell.alignment = Alignment(horizontal="left")
    cell.font = Font(bold=True, name='Arial')

    ws_equips.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=7)
    cell.value = "Nombre Completo"
    cell.alignment = Alignment(horizontal="left")
    cell.font = Font(bold=True, name='Arial')

    for col in range(1, 7):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    for col in range(7, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    current_row = ws_equips.max_row + 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = "Cédula No."
    cell.alignment = Alignment(horizontal="left")
    cell.font = Font(bold=True, name='Arial')

    ws_equips.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=7)
    cell.value = "Cédula No."
    cell.alignment = Alignment(horizontal="left")
    cell.font = Font(bold=True, name='Arial')

    for col in range(1, 7):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    for col in range(7, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    current_row += 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = "RESPONSABLES DE LAS DEPENDENCIAS"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(fill_type="solid", fgColor="4d93d9")
    cell.font = Font(bold=True, name='Arial')

    for col in range(1, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    current_row = ws_equips.max_row + 1
    ws_equips.row_dimensions[current_row].height = 90

    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = "Firma del Aprobador (Jefe de Abastecimiento y Logistica)"
    cell.alignment = Alignment(horizontal="center", vertical="bottom")
    cell.font = Font(bold=True, name='Arial')

    ws_equips.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=7)
    cell.value = "Firma - Responsable de la Dependencia Inspeccionada (Capitan, Jefe Area, Etc)"
    cell.alignment = Alignment(horizontal="center", vertical="bottom")
    cell.font = Font(bold=True, name='Arial')

    for col in range(1, 7):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    for col in range(7, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    current_row = ws_equips.max_row + 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = "Nombre Completo"
    cell.alignment = Alignment(horizontal="left")
    cell.font = Font(bold=True, name='Arial')

    ws_equips.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=7)
    cell.value = "Nombre Completo"
    cell.alignment = Alignment(horizontal="left")
    cell.font = Font(bold=True, name='Arial')

    for col in range(1, 7):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    for col in range(7, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    current_row = ws_equips.max_row + 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = "Cédula No."
    cell.alignment = Alignment(horizontal="left")
    cell.font = Font(bold=True, name='Arial')

    ws_equips.merge_cells(start_row=current_row, start_column=7, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=7)
    cell.value = "Cédula No."
    cell.alignment = Alignment(horizontal="left")
    cell.font = Font(bold=True, name='Arial')

    for col in range(1, 7):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    for col in range(7, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    current_row += 1
    ws_equips.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    cell = ws_equips.cell(row=current_row, column=1)
    cell.value = '“Este documento es propiedad de "SERPORT S.A.S" Se prohíbe su reproducción parcial o total.”'
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(fill_type="solid", fgColor="4d93d9")
    cell.font = Font(name='Arial', italic=True, size=10)

    for col in range(1, 13):
        ws_equips.cell(row=current_row, column=col).border = thin_border

    # === Insertar la imagen desde la URL ===
    url = "https://hivik.s3.us-east-2.amazonaws.com/static/Logo.png"
    img_response = requests.get(url)
    if img_response.status_code == 200:
        image_data = BytesIO(img_response.content)
        image_data.name = "Logo.png"  # Asignar un nombre con la extensión adecuada
        image_data.seek(0)           # Reiniciar el puntero del archivo
        img = ExcelImage(image_data)
        img.width = 300  # Asignar dimensiones a la imagen
        img.height = 80
        ws_equips.add_image(img, "A1")
    else:
        print("No se pudo descargar la imagen.")

    output = io.BytesIO()  # Exportar a HttpResponse
    wb.save(output)
    output.seek(0)

    filename = f"{asset.abbreviation}_Equipos_y_Suministros.xlsx"
    response = HttpResponse(output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename={filename}'
    return response
