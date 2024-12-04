from django.views.generic import ListView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import permission_required, login_required
from django.utils.http import url_has_allowed_host_and_scheme
from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from datetime import date
from itertools import groupby
from dateutil.relativedelta import relativedelta
from operator import attrgetter
from django.core.paginator import Paginator
from .models import *
from .forms import *
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime, time, date
import holidays
from django.contrib import messages
from datetime import time
import json

# Horarios de trabajo
WEEKDAY_START = time(7, 30)
WEEKDAY_END = time(17, 0)

SATURDAY_START = time(8, 0)
SATURDAY_END = time(12, 0)

colombia_holidays = holidays.Colombia()

class OvertimeListView(LoginRequiredMixin, ListView):
    model = Overtime
    template_name = 'overtime/overtime_list.html'
    context_object_name = 'overtime_entries'
    paginate_by = 25

    def get_queryset(self):
        user = self.request.user
        person_name = self.request.GET.get('person_name', '')
        asset_name = self.request.GET.get('asset', '')
        date_filter = self.request.GET.get('date', '')
        aprobado_filter = self.request.GET.get('aprobado', 'all')

        today = date.today()

        if today.day < 24:
            start_date = (today - relativedelta(months=1)).replace(day=24)
            end_date = today.replace(day=24)
        else:
            start_date = today.replace(day=24)
            end_date = (today + relativedelta(months=1)).replace(day=24)

        overtime_entries = Overtime.objects.filter(fecha__gte=start_date, fecha__lt=end_date)

        if user.groups.filter(name='maq_members').exists():
            asset = Asset.objects.filter(models.Q(supervisor=user) | models.Q(capitan=user)).first()
            if asset:
                overtime_entries = overtime_entries.filter(asset=asset)
            else:
                overtime_entries = overtime_entries.filter(reportado_por=user)

        if person_name:
            overtime_entries = overtime_entries.filter(nombre_completo__icontains=person_name)

        if asset_name:
            overtime_entries = overtime_entries.filter(asset__name__icontains=asset_name)

        if date_filter:
            overtime_entries = overtime_entries.filter(fecha=date_filter)

        if aprobado_filter == 'aprobado':
            overtime_entries = overtime_entries.filter(approved=True)
        elif aprobado_filter == 'no_aprobado':
            overtime_entries = overtime_entries.filter(approved=False)

        overtime_entries = overtime_entries.order_by('-fecha', 'asset', 'justificacion')

        return overtime_entries

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        page_obj = context['page_obj']

        date_asset_justification_groups = []

        for fecha, fecha_entries in groupby(page_obj.object_list, key=attrgetter('fecha')):
            asset_groups = []
            fecha_entries = list(fecha_entries)
            for asset, asset_entries in groupby(fecha_entries, key=attrgetter('asset')):
                asset_entries = list(asset_entries)
                for justificacion, justificacion_entries in groupby(asset_entries, key=attrgetter('justificacion')):
                    justificacion_entries = list(justificacion_entries)
                    asset_groups.append({
                        'asset': asset,
                        'justificacion': justificacion,
                        'entries': justificacion_entries
                    })
            date_asset_justification_groups.append({
                'fecha': fecha,
                'assets': asset_groups
            })

        context['date_asset_groups'] = date_asset_justification_groups
        return context


@login_required
@permission_required('got.can_approve_overtime', raise_exception=True)
def approve_overtime(request, pk):
    if request.method == 'POST':
        overtime_entry = get_object_or_404(Overtime, pk=pk)
        # Cambiar el estado de aprobación
        overtime_entry.approved = not overtime_entry.approved
        overtime_entry.save()
        
        next_url = request.POST.get('next')
        if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
            return redirect(next_url)
        else:
            return redirect('overtime:overtime_list')
    return redirect('overtime:overtime_list')


@login_required
def edit_overtime(request, pk):
    overtime_entry = get_object_or_404(Overtime, pk=pk)
    if request.method == 'POST':
        form = OvertimeEditForm(request.POST, instance=overtime_entry)
        if form.is_valid():
            form.save()
            return redirect('overtime:overtime_list')
    return redirect('overtime:overtime_list')


@login_required
def delete_overtime(request, pk):
    overtime_entry = get_object_or_404(Overtime, pk=pk)
    if request.method == 'POST':
        overtime_entry.delete()
        return redirect('overtime:overtime_list')
    return redirect('overtime:overtime_list')


@login_required
def export_overtime_excel(request):

    today = date.today()
    if today.day < 24:
        start_date = (today - relativedelta(months=1)).replace(day=24)
        end_date = today.replace(day=24)
    else:
        start_date = today.replace(day=24)
        end_date = (today + relativedelta(months=1)).replace(day=24)

    overtime_entries = Overtime.objects.filter(fecha__gte=start_date, fecha__lt=end_date)

    person_name = request.GET.get('person_name', '')
    asset_name = request.GET.get('asset', '')
    date_filter = request.GET.get('date', '')

    if person_name:
        overtime_entries = overtime_entries.filter(nombre_completo__icontains=person_name)

    if asset_name:
        overtime_entries = overtime_entries.filter(asset__name__icontains=asset_name)

    if date_filter:
        overtime_entries = overtime_entries.filter(fecha=date_filter)

    overtime_entries = overtime_entries.order_by('fecha', 'asset', 'justificacion')
    wb = openpyxl.Workbook()
    ws = wb.active

    last_month = today - relativedelta(months=1)
    report_title = f"REPORTE 24 {last_month.strftime('%B')} hasta 24 {today.strftime('%B')} {today.year}"
    ws.append([report_title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(row=1, column=1).font = openpyxl.styles.Font(size=14, bold=True)

    headers = ['Nombre completo', 'Cédula', 'Fecha', 'Justificación', 'Hora inicio', 'Hora fin', 'Aprueba', 'Transporte']
    ws.append(headers)

    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions[get_column_letter(col_num)].width = 20  # Ajustar ancho de columna

    # Añadir los datos
    for entry in overtime_entries:
        # Calcular transporte
        if entry.hora_fin >= time(19, 0):  # 7:00 PM
            transporte = 'Sí'
        else:
            transporte = 'No'

        row = [
            entry.nombre_completo,
            entry.cedula,
            entry.fecha.strftime('%d/%m/%Y') if entry.fecha else '',
            entry.justificacion,
            entry.hora_inicio.strftime('%I:%M %p') if entry.hora_inicio else '',
            entry.hora_fin.strftime('%I:%M %p') if entry.hora_fin else '',
            'Sí' if entry.approved else 'No',
            transporte,
        ]
        ws.append(row)

    # Configurar la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Reporte_Horas_Extras_{today.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response


@login_required
def overtime_report(request):
    if request.method == 'POST':
        common_form = OvertimeCommonForm(request.POST)
        person_formset = OvertimePersonFormSet(request.POST)

        if common_form.is_valid() and person_formset.is_valid():

            try:
                asset = Asset.objects.filter(models.Q(supervisor=request.user) | models.Q(capitan=request.user)).first()
            except Asset.DoesNotExist:
                asset = None 
                
            # Guardar los datos comunes
            fecha = common_form.cleaned_data['fecha']
            hora_inicio = common_form.cleaned_data['hora_inicio']
            hora_fin = common_form.cleaned_data['hora_fin']
            justificacion = common_form.cleaned_data['justificacion']
            overtime_periods = calcular_horas_extras(fecha, hora_inicio, hora_fin)

            if not overtime_periods:
                messages.warning(request, 'Las horas reportadas no califican como horas extras.')
                return redirect('overtime:overtime_report')
            
            def subtract_time_ranges(desired_start, desired_end, existing_ranges):
                intervals = [(desired_start, desired_end)]
                for ex_start, ex_end in existing_ranges:
                    new_intervals = []
                    for interval_start, interval_end in intervals:
                        # No hay solapamiento
                        if ex_end <= interval_start or ex_start >= interval_end:
                            new_intervals.append((interval_start, interval_end))
                        else:
                            # Hay solapamiento, ajustamos los intervalos
                            if ex_start > interval_start:
                                new_intervals.append((interval_start, ex_start))
                            if ex_end < interval_end:
                                new_intervals.append((ex_end, interval_end))
                    intervals = new_intervals
                return intervals

            for person_form in person_formset:
                nombre_completo = person_form.cleaned_data.get('nombre_completo')
                cedula = person_form.cleaned_data.get('cedula')
                cargo = person_form.cleaned_data.get('cargo')

                if not nombre_completo and not cargo:
                    continue

                # Obtener los registros existentes para la cédula y fecha
                existing_entries = Overtime.objects.filter(
                    cedula=cedula,
                    fecha=fecha,
                ).order_by('hora_inicio')

                existing_ranges = [(entry.hora_inicio, entry.hora_fin) for entry in existing_entries]

                created_any = False  # Para verificar si se creó al menos un registro

                for period_start, period_end in overtime_periods:
                    # Restar los intervalos existentes
                    non_overlapping_intervals = subtract_time_ranges(period_start, period_end, existing_ranges)
                    if not non_overlapping_intervals:
                        messages.warning(request, f'Todas las horas reportadas para {nombre_completo} ({cedula}) en el periodo {period_start.strftime("%I:%M %p")} - {period_end.strftime("%I:%M %p")} ya han sido registradas.')
                        continue
                    for interval_start, interval_end in non_overlapping_intervals:
                        Overtime.objects.create(
                            fecha=fecha,
                            hora_inicio=interval_start,
                            hora_fin=interval_end,
                            justificacion=justificacion,
                            nombre_completo=nombre_completo,
                            cedula=cedula,
                            cargo=cargo,
                            reportado_por=request.user,
                            asset=asset,
                            approved=False 
                        )
                        # Añadir el nuevo intervalo a la lista de existentes
                        existing_ranges.append((interval_start, interval_end))
                        existing_ranges.sort(key=lambda x: x[0])  # Ordenar por hora de inicio
                        created_any = True
                    if len(non_overlapping_intervals) < len(overtime_periods):
                        messages.warning(request, f'Algunas horas reportadas para {nombre_completo} ({cedula}) en el periodo {period_start.strftime("%I:%M %p")} - {period_end.strftime("%I:%M %p")} ya han sido registradas y fueron excluidas.')
                if created_any:
                    messages.success(request, f'Reporte de horas extras para {nombre_completo} ({cedula}) ha sido registrado.')
                else:
                    messages.warning(request, f'No se pudo registrar ninguna hora extra para {nombre_completo} ({cedula}).')

            return redirect('overtime:overtime_success')
    else:
        common_form = OvertimeCommonForm()
        person_formset = OvertimePersonFormSet()

    current_year = date.today().year
    next_year = current_year + 1
    colombia_holidays = holidays.Colombia(years=[current_year, next_year])
    # Obtener las fechas en formato 'YYYY-MM-DD'
    holiday_dates = [h.strftime('%Y-%m-%d') for h in colombia_holidays.keys()]

    context = {
        'common_form': common_form,
        'person_formset': person_formset,
        'holiday_dates': json.dumps(holiday_dates),
    }
    return render(request, 'overtime/overtime_report.html', context)


def overtime_success(request):
    return render(request, 'overtime/overtime_success.html')


def calcular_horas_extras(fecha, hora_inicio, hora_fin):
    overtime_periods = []
    day_type = None  # 'weekday', 'saturday', 'sunday', 'holiday'

    if fecha in colombia_holidays:
        day_type = 'holiday'
    elif fecha.weekday() == 6:  # Domingo
        day_type = 'sunday'
    elif fecha.weekday() == 5:  # Sábado
        day_type = 'saturday'
    else:
        day_type = 'weekday'

    # Definir el horario laboral según el tipo de día
    if day_type == 'weekday':
        start_work = WEEKDAY_START
        end_work = WEEKDAY_END
    elif day_type == 'saturday':
        start_work = SATURDAY_START
        end_work = SATURDAY_END
    else:
        # Domingos y festivos: todo es horas extras
        if hora_inicio < hora_fin:
            overtime_periods.append((hora_inicio, hora_fin))
        return overtime_periods

    # Horas antes del inicio de la jornada laboral
    if hora_inicio < start_work:
        overtime_periods.append((hora_inicio, min(hora_fin, start_work)))

    # Horas después del fin de la jornada laboral
    if hora_fin > end_work:
        overtime_periods.append((max(hora_inicio, end_work), hora_fin))

    # Horas completamente fuera del horario laboral
    if hora_inicio >= hora_fin:
        return overtime_periods

    if hora_inicio >= end_work or hora_fin <= start_work:
        overtime_periods.append((hora_inicio, hora_fin))

    return overtime_periods
