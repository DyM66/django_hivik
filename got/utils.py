from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from django.apps import apps
from django.db import transaction as db_transaction
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import get_template
from django.utils import timezone
import calendar
from io import BytesIO
from xhtml2pdf import pisa
from django.http import HttpResponse
from .forms import RutinaFilterForm
from django.db.models import Sum
from collections import defaultdict
from datetime import date
import openpyxl
from openpyxl.styles import Alignment, Font
from django.contrib.auth.models import Group
import pandas as pd
from openpyxl.utils import get_column_letter
from megger_app.models import Megger
from django.db.models import Prefetch, Sum
from django.core.management.base import BaseCommand
from inv.models import EquipoCodeCounter
from django.db import transaction
from preoperacionales.models import *

from .models import *
from inv.models import DarBaja, Transferencia
from dth.models import UserProfile

from django.db.models import Func

class DayInterval(Func):
    template = "(%(expressions)s * interval '1 day')"


def update_compliance_all():
    """
    Recalcula y guarda el compliance de todos los Assets.
    Llamable desde shell con:
        >>> from got.management.commands.update_compliance import update_compliance_all
        >>> update_compliance_all()
    """
    assets = Asset.objects.all()
    for asset in assets:
        asset.update_maintenance_compliance_cache()
    print("Proceso completado. Se ha actualizado el compliance de todos los Assets.")


def actualizar_rutas_dependientes(ruta):
    '''
    Utilizada en: views/OtDetailView
    '''
    ruta.intervention_date = timezone.now()
    ruta.save()
    if ruta.dependencia is not None:
        actualizar_rutas_dependientes(ruta.dependencia)


def render_to_pdf(template_src, context_dict={}):
    """
    Función para renderizar un template HTML a PDF.
    """
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return HttpResponse('Error al generar el PDF', status=400)


def generate_equipo_code(asset_abbr, tipo):
    """
    Retorna el siguiente código único con el formato:
        <asset_abbr>-<tipo>-<seq>
    donde <seq> es un número incremental con 3 dígitos.
    """
    with transaction.atomic():
        # Bloquear la fila correspondiente para evitar condiciones de carrera
        counter, created = EquipoCodeCounter.objects.select_for_update().get_or_create(
            asset_abbr=asset_abbr,
            tipo=tipo,
            defaults={'last_seq': 0}
        )
        new_seq = counter.increment_seq()
        seq_str = str(new_seq).zfill(3)
        return f"{asset_abbr}-{tipo}-{seq_str}"


def get_filtered_rutas(asset, user, request_data=None):
    """
    Retorna las rutas de mantenimiento filtradas.

    - Por defecto: devuelve las rutas en que r.percentage_remaining < 10 
      o (r.ot existe y r.ot.state == 'x').
    - Si se utiliza el formulario (es decir, request_data es válido), se exige además que:
         (r.next_date.year < year) or (r.next_date.year == year and r.next_date.month <= month)
    
    Devuelve además el nombre del mes en español.
    """
    # Valores por defecto: el mes y el año actual.
    month = datetime.now().month
    year = datetime.now().year
    
    # Instanciamos el formulario; se asume que RutinaFilterForm está importado.
    form = RutinaFilterForm(request_data, asset=asset)
    extra_filter = False
    if form.is_valid():
        month = int(form.cleaned_data['month'])
        year = int(form.cleaned_data['year'])
        selected_locations = form.cleaned_data.get('locations', [])
        extra_filter = True
        print('si')
    else:
        selected_locations = []
        print('no')

    print(extra_filter)

    # Filtrar las rutas según los sistemas del asset (se asume que get_full_systems_ids está definida)
    rutas_qs = Ruta.objects.filter(
        system__in=get_full_systems_ids(asset, user)
    ).exclude(system__state__in=['x', 's']).order_by('-nivel', 'frecuency')

    if selected_locations:
        rutas_qs = rutas_qs.filter(
            models.Q(equipo__ubicacion__in=selected_locations) |
            models.Q(equipo__isnull=True)
        )

    all_rutas = list(rutas_qs)

    # Definir la función de filtrado
    if extra_filter:
        def match_criteria(r):
            # Solo se incluyen rutas que tengan next_date y que cumplan la condición extra
            if not r.next_date:
                return False
            print(r.next_date)
            next_date_ok = (r.next_date.year < year) or (r.next_date.year == year and r.next_date.month <= month)
            base_ok = r.percentage_remaining < 10 or (r.ot and r.ot.state == 'x')
            return base_ok or next_date_ok
    else:
        def match_criteria(r):
            return (r.percentage_remaining < 10 or (r.ot and r.ot.state == 'x'))

    filtered_rutas = [ruta for ruta in all_rutas if match_criteria(ruta)]
    current_month_name_es = traductor(calendar.month_name[month])
    return filtered_rutas, current_month_name_es


def pro_export_to_excel(model, headers, data, filename='export.xlsx', sheet_name=None, table_title=None):
    """
    Exporta datos a un archivo Excel.

    :param model: Modelo de Django o nombre del modelo (str).
    :param headers: Lista de nombres de encabezados de columnas.
    :param data: Lista de listas con los datos de cada fila.
    :param filename: Nombre del archivo Excel a generar.
    :param sheet_name: Nombre de la hoja (por defecto, el nombre del modelo).
    :param table_title: Título de la tabla que abarcará todas las columnas.
    """
    if sheet_name is None:
        sheet_name = model.__name__ if hasattr(model, '__name__') else str(model)

    # Crear un libro de trabajo y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Agregar el título de la tabla si se proporciona
    current_row = 1
    if table_title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        cell = ws.cell(row=1, column=1)
        cell.value = table_title
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True, size=14)
        current_row += 1

    # Agregar los encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Agregar los datos
    for row_data in data:
        current_row += 1
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = cell_value

    # Ajustar el ancho de las columnas automáticamente
    for idx, column_cells in enumerate(ws.columns, 1):
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        column_letter = get_column_letter(idx)
        ws.column_dimensions[column_letter].width = length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Preparar la respuesta HTTP
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response


def get_suministros(asset, keyword_filter):
    """
    Obtiene los suministros filtrados por un filtro de palabras clave.

    :param asset: Instancia de Asset.
    :param keyword_filter: Filtro Q para filtrar los artículos.
    :return: QuerySet de suministros filtrados.
    """
    suministros = Suministro.objects.filter(asset=asset).filter(keyword_filter).select_related('item').order_by('item__seccion')
    return suministros


def handle_transfer(request, asset, suministro, transfer_cantidad_str, destination_asset_id, transfer_motivo, transfer_fecha_str):
    """
    Maneja la transferencia de suministros entre activos.

    :param request: Objeto HttpRequest.
    :param asset: Asset de origen.
    :param suministro: Suministro de origen.
    :param transfer_cantidad_str: Cantidad a transferir en formato string.
    :param destination_asset_id: ID del asset de destino.
    :param transfer_motivo: Motivo de la transferencia.
    :param transfer_fecha_str: Fecha de la transferencia en formato string.
    :return: None o renderiza una plantilla de confirmación.
    """
    try:
        transfer_fecha = datetime.strptime(transfer_fecha_str, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'Fecha inválida.')
        return redirect(request.path)

    try:
        transfer_cantidad = Decimal(transfer_cantidad_str)
    except InvalidOperation:
        messages.error(request, 'Cantidad inválida.')
        return redirect(request.path)

    if transfer_cantidad <= 0:
        messages.error(request, 'La cantidad a transferir debe ser mayor a cero.')
        return redirect(request.path)

    if transfer_cantidad > suministro.cantidad:
        messages.error(request, 'La cantidad a transferir no puede ser mayor a la cantidad disponible.')
        return redirect(request.path)

    existing_transaction = Transaction.objects.filter(
        suministro=suministro,
        fecha=transfer_fecha,
        tipo='t'
    ).first()

    confirm_overwrite = request.POST.get('confirm_overwrite', 'no')

    if existing_transaction and confirm_overwrite != 'yes':
        context = {
            'asset': asset,
            'overwriting_transactions': [('Transferencia', existing_transaction)],
            'post_data': request.POST,
        }
        return render(request, 'got/assets/confirm_overwrite.html', context)

    destination_asset = get_object_or_404(Asset, abbreviation=destination_asset_id)

    with db_transaction.atomic():
        suministro_destino, _ = Suministro.objects.get_or_create(
            asset=destination_asset,
            item=suministro.item,
            defaults={'cantidad': Decimal('0.00')}
        )

        if existing_transaction:
            # Revertir cantidades previas
            suministro.cantidad += existing_transaction.cant
            suministro_destino.cantidad -= existing_transaction.cant

        suministro.cantidad -= transfer_cantidad
        suministro_destino.cantidad += transfer_cantidad

        suministro.save()
        suministro_destino.save()

        if existing_transaction:
            # Actualizar transacción existente
            existing_transaction.cant = transfer_cantidad
            existing_transaction.user = request.user.get_full_name()
            existing_transaction.motivo = transfer_motivo
            existing_transaction.cant_report = suministro.cantidad
            existing_transaction.cant_report_transf = suministro_destino.cantidad
            existing_transaction.suministro_transf = suministro_destino
            existing_transaction.save()
        else:
            # Crear nueva transacción
            Transaction.objects.create(
                suministro=suministro,
                cant=transfer_cantidad,
                fecha=transfer_fecha,
                user=request.user.get_full_name(),
                motivo=transfer_motivo,
                tipo='t',
                cant_report=suministro.cantidad,
                suministro_transf=suministro_destino,
                cant_report_transf=suministro_destino.cantidad
            )

    messages.success(request, f'Se ha transferido {transfer_cantidad} {suministro.item.presentacion} de {suministro.item.name} al activo {destination_asset.name}.')

def handle_inventory_update(request, asset, suministros, motivo_global=''):
    """
    +:param request: Objeto HttpRequest.
    :param asset: Instancia de Asset.
    :param suministros: QuerySet de suministros a actualizar.
    :param motivo_global: Motivo general para las transacciones.
    :return: None o renderiza una plantilla de confirmación.

    Procesa la actualización del inventario basado en los suministros y las entradas de consumo/ingreso.
    Asegura que no se generen cantidades negativas.

    """
    fecha_reporte_str = request.POST.get('fecha_reporte', timezone.now().date().strftime('%Y-%m-%d'))
    confirm_overwrite = request.POST.get('confirm_overwrite', 'no')

    try:
        fecha_reporte = datetime.strptime(fecha_reporte_str, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'Fecha inválida.')
        return redirect(request.path)

    overwriting_transactions = []

    for suministro in suministros:
        cantidad_consumida_str = request.POST.get(f'consumido_{suministro.id}', '0') or '0'
        cantidad_ingresada_str = request.POST.get(f'ingresado_{suministro.id}', '0') or '0'

        try:
            cantidad_consumida = Decimal(cantidad_consumida_str)
        except InvalidOperation:
            cantidad_consumida = Decimal('0')

        try:
            cantidad_ingresada = Decimal(cantidad_ingresada_str)
        except InvalidOperation:
            cantidad_ingresada = Decimal('0')

        if cantidad_consumida == Decimal('0') and cantidad_ingresada == Decimal('0'):
            continue

        existing_transactions = []

        # Calcular la cantidad total después de los cambios
        cantidad_actual = suministro.cantidad
        nueva_cantidad = cantidad_actual + cantidad_ingresada - cantidad_consumida

        # Validar que la nueva cantidad no sea negativa
        if nueva_cantidad < 0:
            messages.error(request, f'La cantidad consumida para "{suministro.item.name}" excede la cantidad disponible.')
            return redirect(request.path)

        if cantidad_ingresada > Decimal('0'):
            existing_ingreso = Transaction.objects.filter(
                suministro=suministro,
                fecha=fecha_reporte,
                tipo='i'
            ).first()
            if existing_ingreso:
                existing_transactions.append(('Ingreso', existing_ingreso))

        if cantidad_consumida > Decimal('0'):
            existing_consumo = Transaction.objects.filter(
                suministro=suministro,
                fecha=fecha_reporte,
                tipo='c'
            ).first()
            if existing_consumo:
                existing_transactions.append(('Consumo', existing_consumo))

        if existing_transactions and confirm_overwrite != 'yes':
            context = {
                'asset': asset,
                'overwriting_transactions': existing_transactions,
                'post_data': request.POST,
            }
            return render(request, 'got/assets/confirm_overwrite.html', context)

    with db_transaction.atomic():
        for suministro in suministros:
            cantidad_consumida_str = request.POST.get(f'consumido_{suministro.id}', '0') or '0'
            cantidad_ingresada_str = request.POST.get(f'ingresado_{suministro.id}', '0') or '0'

            try:
                cantidad_consumida = Decimal(cantidad_consumida_str)
            except InvalidOperation:
                cantidad_consumida = Decimal('0')

            try:
                cantidad_ingresada = Decimal(cantidad_ingresada_str)
            except InvalidOperation:
                cantidad_ingresada = Decimal('0')

            if cantidad_consumida == Decimal('0') and cantidad_ingresada == Decimal('0'):
                continue

            # Handle Ingreso
            if cantidad_ingresada > Decimal('0'):
                suministro, _ = Suministro.objects.get_or_create(asset=asset, item=suministro.item)
                transaccion_ingreso = Transaction.objects.filter(
                    suministro=suministro,
                    fecha=fecha_reporte,
                    tipo='i'
                ).first()
                if transaccion_ingreso:
                    # Revertir efecto previo
                    suministro.cantidad -= transaccion_ingreso.cant
                    # Actualizar transacción
                    transaccion_ingreso.cant = cantidad_ingresada
                    transaccion_ingreso.user = request.user.get_full_name()
                    transaccion_ingreso.motivo = motivo_global or request.POST.get('motivo', '')
                    transaccion_ingreso.cant_report = suministro.cantidad + cantidad_ingresada
                    transaccion_ingreso.save()
                else:
                    transaccion_ingreso = Transaction.objects.create(
                        suministro=suministro,
                        cant=cantidad_ingresada,
                        fecha=fecha_reporte,
                        user=request.user.get_full_name(),
                        motivo=motivo_global or request.POST.get('motivo', ''),
                        tipo='i',
                        cant_report=suministro.cantidad + cantidad_ingresada
                    )
                suministro.cantidad += cantidad_ingresada
                suministro.save()

            # Handle Consumo
            if cantidad_consumida > Decimal('0'):
                suministro, _ = Suministro.objects.get_or_create(asset=asset, item=suministro.item)
                transaccion_consumo = Transaction.objects.filter(
                    suministro=suministro,
                    fecha=fecha_reporte,
                    tipo='c'
                ).first()
                if transaccion_consumo:
                    # Revertir efecto previo
                    suministro.cantidad += transaccion_consumo.cant
                    # Actualizar transacción
                    transaccion_consumo.cant = cantidad_consumida
                    transaccion_consumo.user = request.user.get_full_name()
                    transaccion_consumo.motivo = motivo_global or request.POST.get('motivo', '')
                    transaccion_consumo.cant_report = suministro.cantidad - cantidad_consumida
                    transaccion_consumo.save()
                else:
                    transaccion_consumo = Transaction.objects.create(
                        suministro=suministro,
                        cant=cantidad_consumida,
                        fecha=fecha_reporte,
                        user=request.user.get_full_name(),
                        motivo=motivo_global or request.POST.get('motivo', ''),
                        tipo='c',
                        cant_report=suministro.cantidad - cantidad_consumida
                    )
                suministro.cantidad -= cantidad_consumida
                suministro.save()

    messages.success(request, 'Inventario actualizado exitosamente.')
    return redirect(request.path)


def handle_delete_transaction(request, transaccion):
    with db_transaction.atomic():
        suministro_origen = transaccion.suministro
        suministro_destino = transaccion.suministro_transf

        # Revertir las cantidades según el tipo de transacción
        if transaccion.tipo == 'i':
            suministro_origen.cantidad -= transaccion.cant
        elif transaccion.tipo == 'c':
            suministro_origen.cantidad += transaccion.cant
        elif transaccion.tipo == 't':
            suministro_origen.cantidad += transaccion.cant
            suministro_destino.cantidad -= transaccion.cant
            suministro_destino.save()
        else:
            # Manejar otros tipos si es necesario
            pass

        suministro_origen.save()
        transaccion.delete()

    messages.success(request, 'La transacción ha sido eliminada y las cantidades han sido actualizadas.')


def generate_excel(transacciones_historial, headers_mapping, filename):
    """
    Genera un archivo Excel a partir del historial de transacciones.

    :param transacciones_historial: QuerySet de transacciones.
    :param headers_mapping: Diccionario de mapeo de campos a encabezados.
    :param filename: Nombre del archivo a generar.
    :return: HttpResponse con el archivo Excel.
    """
    df = pd.DataFrame(list(transacciones_historial.values(*headers_mapping.keys())))
    df.rename(columns=headers_mapping, inplace=True)
    data = df.values.tolist()
    headers = list(headers_mapping.values())
    return pro_export_to_excel(Transaction, headers, data, filename=filename)


def copiar_rutas_de_sistema(system_id):
    sistema_origen = get_object_or_404(System, id=system_id)
    asset = sistema_origen.asset
    sistemas_destino = System.objects.filter(asset=asset).exclude(id=system_id)
    
    with db_transaction.atomic():
        for sistema in sistemas_destino:
            for ruta in sistema_origen.rutas.all():
                ruta_nueva = Ruta.objects.create(
                    name=ruta.name,
                    control=ruta.control,
                    frecuency=ruta.frecuency,
                    intervention_date=ruta.intervention_date,
                    astillero=ruta.astillero,
                    system=sistema,
                    # equipo=ruta.equipo.code,
                    dependencia=ruta.dependencia,
                )
                
                for tarea in ruta.task_set.all():
                    Task.objects.create(
                        ot=tarea.ot,
                        ruta=ruta_nueva,
                        responsible=tarea.responsible,
                        description=tarea.description,
                        procedimiento=tarea.procedimiento,
                        hse=tarea.hse,
                        news=tarea.news,
                        evidence=tarea.evidence,
                        start_date=tarea.start_date,
                        men_time=tarea.men_time,
                        finished=tarea.finished
                    )
    print(f"Rutas y tareas copiadas exitosamente a otros sistemas en el mismo activo {asset.name}.")


def calculate_executions(ruta, period_start, period_end):
    initial_next_date = ruta.next_date
    system = ruta.system

    if not initial_next_date:
        return 0

    # Para rutinas vencidas: si initial_next_date < period_start significa que ya debería haberse hecho,
    # igualmente contaremos desde esa fecha (overdue).
    current_date = initial_next_date

    executions = 0
    if ruta.control == 'd':
        # Frecuencia en días
        freq_days = ruta.frecuency
        # Avanzar en intervalos de freq_days desde current_date mientras esté dentro del periodo
        while current_date <= period_end:
            # Chequear si cae dentro del periodo
            if current_date >= period_start and current_date <= period_end:
                executions += 1
            current_date = current_date + timedelta(days=freq_days)

    elif ruta.control in ['h', 'k']:
        # Para horas o kilómetros, next_date ya calcula un ndays aproximado y retorna una fecha
        # Supondremos que la frecuencia es en las unidades originales, pero dado que next_date
        # devuelve una fecha final, trataremos la frecuencia como si fuera recurrente también en días.

        # Nota: en el modelo next_date para horas/kilómetros calcula ndays con base en prom_hours.
        # Entonces, asumimos que cada repetición ocurre cada 'frecuency' horas, que se traducen
        # en un intervalo aproximado de tiempo en días. Debemos recalcular el intervalo en días
        # similar a como se hace en next_date.

        # Reutilizamos la lógica: si se quisiera ser exacto, deberíamos replicar el cálculo
        # de días para cada iteración, pero eso puede ser demasiado complejo.
        # Aquí haremos una suposición simplificada:
        # - Obtenemos ndays inicial desde next_date (ya calculado).
        # - Para las siguientes ejecuciones, asumimos el mismo patrón:
        #   es decir, cada frecuencia en horas corresponde al mismo número de días (la razón
        #   es que no tenemos un cálculo dinámico de horas consumidas, solo uno inicial).

        # Como en next_date se calcula ndays = int(inv / prom_hours) para h/k,
        # debemos replicar esa lógica. Obtenemos ndays inicial (diferencia entre next_date y date.today())
        if system.state == 's':  # 's' representa 'Stand by'
            return 1
        
        ndays_iniciales = (initial_next_date - date.today()).days
        if ndays_iniciales < 1:
            ndays_iniciales = 1  # por si acaso, evitar division por cero u otros casos

        # Cada frecuencia en horas se traduce en freq_hours / prom_hours días aproximadamente.
        equipo = ruta.equipo
        prom_hours = equipo.prom_hours if equipo.prom_hours else 1
        # Si prom_hours es 0 se tomó 2 por defecto en next_date.

        # Calc days_per_execution = freq_hours / prom_hours (en días)
        # Si freq_hours es ruta.frecuency
        freq_hours = ruta.frecuency
        try:
            days_per_execution = int(freq_hours / prom_hours)
            if days_per_execution < 1:
                days_per_execution = 1
        except ZeroDivisionError:
            days_per_execution = 1

        # Ahora desde current_date sumamos days_per_execution cada vez
        while current_date <= period_end:
            if current_date <= period_end:
                executions += 1
            current_date = current_date + timedelta(days=days_per_execution)

    else:
        # Si no es d, h o k, retornar 0 por ahora
        return 0

    return executions


def calcular_repeticiones2(ruta, periodo='anual'):
    periodos = {
        'trimestral': 90,
        'semestral': 180,
        'anual': 365,
        'quinquenal': 1825,  # 5 años
    }

    dias_periodo = periodos.get(periodo, 0)
    today = datetime.now().date()
    meses_ejecucion = []
    
    if ruta.control == 'd':
        frecuencia_dias = ruta.frecuency
        next_date = ruta.next_date

        repeticiones = 0
        if next_date and next_date <= today + timedelta(days=dias_periodo):
            repeticiones += 1
            meses_ejecucion.append(next_date.strftime('%B %Y'))
            remaining_days = (today + timedelta(days=dias_periodo) - next_date).days
            for _ in range(remaining_days // frecuencia_dias):
                next_date += timedelta(days=frecuencia_dias)
                meses_ejecucion.append(next_date.strftime('%B %Y'))
            repeticiones += remaining_days // frecuencia_dias
    
    elif ruta.get_control_display() == 'Horas':
        diferencia_dias = (ruta.next_date - ruta.intervention_date).days
        repeticiones = 0

        if diferencia_dias > 0:
            next_date = ruta.next_date
            if next_date and next_date <= (today + timedelta(days=dias_periodo)):
                repeticiones += 1
                meses_ejecucion.append(next_date.strftime('%B %Y'))
                remaining_days = (today + timedelta(days=dias_periodo) - next_date).days
                for _ in range(remaining_days // diferencia_dias):
                    next_date += timedelta(days=diferencia_dias)
                    meses_ejecucion.append(next_date.strftime('%B %Y'))
                repeticiones += remaining_days // diferencia_dias

    return repeticiones, meses_ejecucion


def traductor(word):
    context = {
        "January": "Enero",
        "February": "Febrero",
        "March": "Marzo",
        "April": "Abril",
        "May": "Mayo",
        "June": "Junio",
        "July": "Julio",
        "August": "Agosto",
        "September": "Septiembre",
        "October": "Octubre",
        "November": "Noviembre",
        "December": "Diciembre"
    }
    return context[word]


def truncate_text(text, length=45):
    if len(text) > length:
        return text[:length] + '...'
    return text


def calculate_status_code(t):
    ot_tasks = Task.objects.filter(ot=t)

    earliest_start_date = min(t.start_date for t in ot_tasks)
    latest_final_date = max(t.final_date for t in ot_tasks)

    today = date.today()

    if latest_final_date < today:
        return 0
    elif earliest_start_date < today < latest_final_date:
        return 1
    elif earliest_start_date > today:
        return 2
    return None


def update_equipo_code(old_code):
    try:
        with transaction.atomic():
            equipo = Equipo.objects.select_for_update().get(code=old_code)
            system = equipo.system
            asset_abbreviation = system.asset.abbreviation
            tipo = equipo.tipo.upper()
 
            generated_code = generate_equipo_code(asset_abbreviation, tipo)

           # Actualizar el código del equipo
            equipo.code = generated_code
            equipo.save()

            # Actualizar todas las relaciones donde esté relacionado el equipo
            print(f"Actualizando código de equipo de '{old_code}' a '{generated_code}'...")

            # HistoryHour
            updated_history_hours = HistoryHour.objects.filter(component__code=old_code).update(component=equipo)
            print(f"HistoryHour actualizado: {updated_history_hours} registros")

            # Suministro
            updated_suministros = Suministro.objects.filter(equipo__code=old_code).update(equipo=equipo)
            print(f"Suministro actualizado: {updated_suministros} registros")

            # Ruta
            updated_rutas = Ruta.objects.filter(equipo__code=old_code).update(equipo=equipo)
            print(f"Ruta actualizado: {updated_rutas} registros")

            # DailyFuelConsumption
            updated_daily_fuel = DailyFuelConsumption.objects.filter(equipo__code=old_code).update(equipo=equipo)
            print(f"DailyFuelConsumption actualizado: {updated_daily_fuel} registros")

            # FailureReport
            updated_failure_reports = FailureReport.objects.filter(equipo__code=old_code).update(equipo=equipo)
            print(f"FailureReport actualizado: {updated_failure_reports} registros")

            # Megger
            updated_megger = Megger.objects.filter(equipo__code=old_code).update(equipo=equipo)
            print(f"Megger actualizado: {updated_megger} registros")

            # Preoperacional
            updated_preoperacional = Preoperacional.objects.filter(vehiculo__code=old_code).update(vehiculo=equipo)
            print(f"Preoperacional actualizado: {updated_preoperacional} registros")

            # PreoperacionalDiario
            updated_preoperacional_diario = PreoperacionalDiario.objects.filter(vehiculo__code=old_code).update(vehiculo=equipo)
            print(f"PreoperacionalDiario actualizado: {updated_preoperacional_diario} registros")

            # Transferencia
            updated_transferencia_origen = Transferencia.objects.filter(origen__equipos__code=old_code).update(origen=equipo.system)
            updated_transferencia_destino = Transferencia.objects.filter(destino__equipos__code=old_code).update(destino=equipo.system)
            print(f"Transferencia actualizado: {updated_transferencia_origen + updated_transferencia_destino} registros")

            # DarBaja
            updated_darbaja = DarBaja.objects.filter(equipo__code=old_code).update(equipo=equipo)
            print(f"DarBaja actualizado: {updated_darbaja} registros")

            # Image
            updated_images = Image.objects.filter(equipo__code=old_code).update(equipo=equipo)
            print(f"Image actualizado: {updated_images} registros")

            # Document
            updated_documents = Document.objects.filter(equipo__code=old_code).update(equipo=equipo)
            print(f"Document actualizado: {updated_documents} registros")

            Equipo.objects.filter(code=old_code).exclude(pk=equipo.pk).delete()

            print(f"El código del equipo '{old_code}' ha sido actualizado a '{generated_code}' en todas las tablas relacionadas.")
    
    except Equipo.DoesNotExist:
        print(f"El equipo con código '{old_code}' no existe.")
    except Exception as e:
        print(f"Ha ocurrido un error: {str(e)}")


def operational_users(current_user):
    if current_user.groups.filter(name__in=['maq_members', 'buzos_members']).exists():
        talleres = Group.objects.get(name='serport_members')
        taller_list = list(talleres.user_set.all())
        taller_list.append(current_user)
        return User.objects.filter(id__in=[user.id for user in taller_list])
    elif current_user.groups.filter(name='super_members').exists():
        return User.objects.exclude(groups__name='gerencia')


def fechas_range():
    hoy = timezone.now().date() - timedelta(days=1)
    hace_30_dias = hoy - timedelta(days=30)
    return [(hace_30_dias + timedelta(days=i)).strftime('%d/%m') for i in range(31)]


def consumos_combustible_asset(asset):
    hoy = timezone.now().date() - timedelta(days=1)
    hace_30_dias = hoy - timedelta(days=30)
    consumos = Transaction.objects.filter(
        suministro__asset=asset,
        suministro__item__id=132,
        fecha__range=[hace_30_dias, hoy],
        tipo='c'
    ).values('fecha').annotate(total_consumido=Sum('cant')).order_by('fecha')
    fechas = [(hace_30_dias + timedelta(days=i)).strftime('%d/%m') for i in range(31)]
    consumos_dict = {consumo['fecha'].strftime('%d/%m'): consumo['total_consumido'] for consumo in consumos}
    consumos_grafica = [consumos_dict.get(fecha, 0) for fecha in fechas]
    return consumos_grafica


def horas_total_asset(asset):
    hoy = timezone.now().date() - timedelta(days=1)
    hace_30_dias = hoy - timedelta(days=30)
    systems = System.objects.filter(asset=asset)
    equipos = Equipo.objects.filter(system__in=systems)
    horas_operacion = HistoryHour.objects.filter(
        component__in=equipos,
        report_date__range=[hace_30_dias, hoy]
    ).values('report_date').annotate(total_horas=Sum('hour')).order_by('report_date')
    horas_dict = {hora['report_date'].strftime('%d/%m'): hora['total_horas'] for hora in horas_operacion}
    horas_grafica = [horas_dict.get(fecha, 0) for fecha in fechas_range()]
    return horas_grafica


def get_full_systems(asset, user):
    systems = asset.system_set.all()
    other_asset_systems = System.objects.filter(location=asset.name).exclude(asset=asset)

    combined = systems.union(other_asset_systems)
    return combined
    # return (systems.union(other_asset_systems)).order_by('group')


def get_full_systems_ids(asset, user):
    systems = asset.system_set.values_list('id', flat=True)
    other_asset_systems = System.objects.filter(location=asset.name).exclude(asset=asset).values_list('id', flat=True)
    all_systems_ids = systems.union(other_asset_systems)

    # if user.groups.filter(name='buzos_members').exists():
    #     station = user.profile.station
    #     if station:
    #         all_systems_ids = System.objects.filter(id__in=all_systems_ids, location__iexact=station).values_list('id', flat=True)
    #     else:
    #         all_systems_ids = System.objects.none().values_list('id', flat=True)
    return all_systems_ids



def consumibles_summary(asset):

    '''
    Función utilizada para calcular la cantidad de articulos de tipo consumible que se repiten dentro de un mismo
    Asset y asociarlo a un grupo Subsystem o General.
    '''

    equipos = Equipo.objects.filter(system__asset=asset).prefetch_related('suministros__item')
    item_subsystems = defaultdict(set)
    items_by_subsystem = defaultdict(set)
    item_cant = defaultdict(set)

    inventory_counts = defaultdict(int)
    suministros = Suministro.objects.filter(asset=asset)

    for suministro in suministros:
        inventory_counts[suministro.item] += suministro.cantidad

    for equipo in equipos:
        subsystem = equipo.subsystem if equipo.subsystem else "General"
        for suministro in equipo.suministros.all():
                item_subsystems[suministro.item].add(subsystem)
                if suministro.item in item_cant:
                    item_cant[suministro.item] += suministro.cantidad
                else:
                    item_cant[suministro.item] = suministro.cantidad
    
    duplicated_items = {item for item, subsystems in item_subsystems.items() if len(subsystems) > 1}

    for equipo in equipos:
        subsystem = equipo.subsystem if equipo.subsystem else "General"
        for suministro in equipo.suministros.all():
            item_tuple = (suministro.item, item_cant[suministro.item], inventory_counts[suministro.item], item_cant[suministro.item] * 2)
            if suministro.item in duplicated_items:
                items_by_subsystem["General"].add(item_tuple)
            else:
                items_by_subsystem[subsystem].add(item_tuple)
    
    items_by_subsystem = {k: list(v) for k, v in items_by_subsystem.items() if v}

    return items_by_subsystem



def create_transaction(data):
    try:
        # Use get_or_create to avoid duplicates
        obj, created = Transaction.objects.get_or_create(
            suministro=data['suministro'],
            fecha=data['fecha'],
            tipo=data['tipo'],
            defaults={
                'cant': data['cant'],
                'user': data['user'],
                'motivo': data['motivo'],
            }
        )
        return created
    except Exception as e:
        print(f'Error creating transaction: {e}')
        return False
    

def get_cargo(full_name):
    """
    Recibe el nombre completo del usuario y devuelve el cargo desde UserProfile.
    """
    try:
        # Separar el nombre completo en nombre y apellido
        nombre, apellido = full_name.split(' ', 1)
        # Buscar el usuario por nombre y apellido (case-insensitive)
        user = User.objects.get(first_name__iexact=nombre, last_name__iexact=apellido)
        # Retornar el cargo desde UserProfile
        return user.profile.cargo if hasattr(user, 'profile') else ''
    except (User.DoesNotExist, ValueError, UserProfile.DoesNotExist):
        return ''


from django.db.models import F, ExpressionWrapper, DateField
def filter_tasks_queryset(request, base_queryset=None):
    """
    Aplica el mismo filtrado de tareas que la vista AssignedTaskByUserListView y
    la vista assignedTasks_pdf.
    Se leen los siguientes parámetros GET:
      - asset_id
      - worker
      - start_date y end_date (para filtrar por rango de fechas)
      - show_finalizadas: si es "1" se muestran ambas (pendientes y finalizadas), 
        si no, se filtran las pendientes (finished=False)
    Devuelve el queryset filtrado.
    """
    if base_queryset is None:
        # Por defecto, solo tareas con OT y start_date definido
        queryset = Task.objects.filter(ot__isnull=False, start_date__isnull=False)
    else:
        queryset = base_queryset

    # Ordenar: en la ListView se ordena por start_date, en la PDF se ordena por asset y start_date.
    # Puedes definir un orden común; en este ejemplo, usaremos 'start_date'
    queryset = queryset.order_by('start_date')

    # Filtrar por asset
    asset_id = request.GET.get('asset_id')
    if asset_id:
        queryset = queryset.filter(ot__system__asset_id=asset_id)

    # Filtrar por responsable
    responsable_id = request.GET.get('worker')
    if responsable_id:
        queryset = queryset.filter(responsible=responsable_id)

    # Filtrado por rango de fechas (solapamiento)
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            queryset = queryset.annotate(
                calc_final_date=ExpressionWrapper(
                    F('start_date') + DayInterval(F('men_time')),
                    output_field=DateField()
                )
            ).filter(calc_final_date__gte=start_date, start_date__lte=end_date)
        except ValueError:
            pass  # Si hay error en el formato, no filtramos por fecha

    # Filtrado por estado utilizando el checkbox de finalizadas
    # Por defecto, se muestran pendientes (finished=False) y el parámetro a usar es "show_finalizadas"
    show_finalizadas = request.GET.get('show_finalizadas')
    if show_finalizadas != "1":
        queryset = queryset.filter(finished=False)

    # Filtrado adicional según el grupo del usuario
    current_user = request.user
    if current_user.groups.filter(name='serport_members').exists():
        queryset = queryset.filter(responsible=current_user)
    elif current_user.groups.filter(name='super_members').exists():
        pass  # Sin filtro adicional
    elif current_user.groups.filter(name__in=['maq_members', 'buzos_members']).exists():
        queryset = queryset.filter(ot__system__asset__supervisor=current_user)
    else:
        queryset = queryset.none()

    return queryset