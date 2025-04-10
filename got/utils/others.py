import openpyxl
import pandas as pd

from datetime import datetime, timedelta, date
from decimal import Decimal, InvalidOperation
from django.apps import apps
from django.db import transaction as db_transaction
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import get_template
from django.utils import timezone
from io import BytesIO
from xhtml2pdf import pisa
from django.http import HttpResponse
from django.db.models import Sum

from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from meg.models import Megger
from django.db.models import Prefetch, Sum
from got.models import EquipoCodeCounter
from django.db import transaction
from preoperacionales.models import *
from got.models import *
from inv.models import DarBaja, Transference
from dth.models import UserProfile
from mto.utils import record_execution
from django.db.models import Func
from inv.models import Transaction

from inv.models import Suministro

class DayInterval(Func):
    template = "(%(expressions)s * interval '1 day')"

def actualizar_rutas(ruta, fecha=timezone.now(), ot=None):
    '''
    Utilizada en: 
    got/views/OtDetailView
    got/views/rutina_form_view
    '''
    ruta.intervention_date = fecha
    if ot:
        ruta.ot = ot
    ruta.save()

    plan = ruta.maintenance_plans.filter(period_start__lte=fecha, period_end__gte=fecha).first()
    if plan:
        record_execution(plan, fecha)

    if ruta.dependencia is not None:
        actualizar_rutas(ruta.dependencia, fecha, ot)


def procesar_tasks_y_dependencias(request, ruta, formset_data):
    tasks = Task.objects.filter(ruta=ruta)
    for task in tasks:
        realizado = request.POST.get(f'realizado_{task.id}') == 'on'
        observaciones = request.POST.get(f'observaciones_{task.id}')
        evidencias = request.FILES.getlist(f'evidencias_{task.id}')
        user = request.POST.get(f'user_{task.id}')
            
        formset_data.append({
            'task': task,
            'realizado': realizado,
            'observaciones': observaciones,
            'evidencias': evidencias,
            'user': user,
        })
        
    if ruta.dependencia:
        procesar_tasks_y_dependencias(request, ruta.dependencia, formset_data)


def update_compliance_all():
    """
    Recalcula y guarda el compliance de todos los Assets.
    Llamable desde shell con:
        >>> from got.management.commands.update_compliance import update_compliance_all
        >>> update_compliance_all()
    """
    assets = Asset.objects.all()
    for asset in assets:
        asset.update_maintenance_compliance_cache()
    print("Proceso completado. Se ha actualizado el compliance de todos los Assets.")


def render_to_pdf(template_src, context_dict={}):
    """
    Función para renderizar un template HTML a PDF.
    """
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return HttpResponse('Error al generar el PDF', status=400)

from django.template.loader import render_to_string
from weasyprint import HTML, CSS
def pdf_render(request, template_src, context_dict={}, file_name="document"):
    html_string = render_to_string(template_src, context_dict)
    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    css = CSS(string='@page { size: A4; margin: 2cm; }')
    pdf_file = html.write_pdf(stylesheets=[css])

    response = HttpResponse(pdf_file, content_type='application/pdf')
    filename = file_name
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


def generate_equipo_code(asset_abbr, tipo):
    """
    Retorna el siguiente código único con el formato:
        <asset_abbr>-<tipo>-<seq>
    donde <seq> es un número incremental con 3 dígitos.
    """
    with transaction.atomic():
        # Bloquear la fila correspondiente para evitar condiciones de carrera
        counter, created = EquipoCodeCounter.objects.select_for_update().get_or_create(
            asset_abbr=asset_abbr,
            tipo=tipo,
            defaults={'last_seq': 0}
        )
        new_seq = counter.increment_seq()
        seq_str = str(new_seq).zfill(3)
        return f"{asset_abbr}-{tipo}-{seq_str}"




def pro_export_to_excel(model, headers, data, filename='export.xlsx', sheet_name=None, table_title=None):
    """
    Exporta datos a un archivo Excel.

    :param model: Modelo de Django o nombre del modelo (str).
    :param headers: Lista de nombres de encabezados de columnas.
    :param data: Lista de listas con los datos de cada fila.
    :param filename: Nombre del archivo Excel a generar.
    :param sheet_name: Nombre de la hoja (por defecto, el nombre del modelo).
    :param table_title: Título de la tabla que abarcará todas las columnas.
    """
    if sheet_name is None:
        sheet_name = model.__name__ if hasattr(model, '__name__') else str(model)

    # Crear un libro de trabajo y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Agregar el título de la tabla si se proporciona
    current_row = 1
    if table_title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        cell = ws.cell(row=1, column=1)
        cell.value = table_title
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True, size=14)
        current_row += 1

    # Agregar los encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Agregar los datos
    for row_data in data:
        current_row += 1
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = cell_value

    # Ajustar el ancho de las columnas automáticamente
    for idx, column_cells in enumerate(ws.columns, 1):
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        column_letter = get_column_letter(idx)
        ws.column_dimensions[column_letter].width = length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Preparar la respuesta HTTP
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response


def generate_excel(transacciones_historial, headers_mapping, filename):
    """
    Genera un archivo Excel a partir del historial de transacciones.

    :param transacciones_historial: QuerySet de transacciones.
    :param headers_mapping: Diccionario de mapeo de campos a encabezados.
    :param filename: Nombre del archivo a generar.
    :return: HttpResponse con el archivo Excel.
    """
    df = pd.DataFrame(list(transacciones_historial.values(*headers_mapping.keys())))
    df.rename(columns=headers_mapping, inplace=True)
    data = df.values.tolist()
    headers = list(headers_mapping.values())
    return pro_export_to_excel(Transaction, headers, data, filename=filename)


def copiar_rutas_de_sistema(system_id):
    sistema_origen = get_object_or_404(System, id=system_id)
    asset = sistema_origen.asset
    sistemas_destino = System.objects.filter(asset=asset).exclude(id=system_id)
    
    with db_transaction.atomic():
        for sistema in sistemas_destino:
            for ruta in sistema_origen.rutas.all():
                ruta_nueva = Ruta.objects.create(
                    name=ruta.name,
                    control=ruta.control,
                    frecuency=ruta.frecuency,
                    intervention_date=ruta.intervention_date,
                    astillero=ruta.astillero,
                    system=sistema,
                    # equipo=ruta.equipo.code,
                    dependencia=ruta.dependencia,
                )
                
                for tarea in ruta.task_set.all():
                    Task.objects.create(
                        ot=tarea.ot,
                        ruta=ruta_nueva,
                        responsible=tarea.responsible,
                        description=tarea.description,
                        procedimiento=tarea.procedimiento,
                        hse=tarea.hse,
                        news=tarea.news,
                        evidence=tarea.evidence,
                        start_date=tarea.start_date,
                        men_time=tarea.men_time,
                        finished=tarea.finished
                    )
    print(f"Rutas y tareas copiadas exitosamente a otros sistemas en el mismo activo {asset.name}.")


def calculate_executions(ruta, period_start, period_end):
    initial_next_date = ruta.next_date
    system = ruta.system

    if not initial_next_date:
        return 0

    # Para rutinas vencidas: si initial_next_date < period_start significa que ya debería haberse hecho,
    # igualmente contaremos desde esa fecha (overdue).
    current_date = initial_next_date

    executions = 0
    if ruta.control == 'd':
        # Frecuencia en días
        freq_days = ruta.frecuency
        # Avanzar en intervalos de freq_days desde current_date mientras esté dentro del periodo
        while current_date <= period_end:
            # Chequear si cae dentro del periodo
            if current_date >= period_start and current_date <= period_end:
                executions += 1
            current_date = current_date + timedelta(days=freq_days)

    elif ruta.control in ['h', 'k']:
        # Para horas o kilómetros, next_date ya calcula un ndays aproximado y retorna una fecha
        # Supondremos que la frecuencia es en las unidades originales, pero dado que next_date
        # devuelve una fecha final, trataremos la frecuencia como si fuera recurrente también en días.

        # Nota: en el modelo next_date para horas/kilómetros calcula ndays con base en prom_hours.
        # Entonces, asumimos que cada repetición ocurre cada 'frecuency' horas, que se traducen
        # en un intervalo aproximado de tiempo en días. Debemos recalcular el intervalo en días
        # similar a como se hace en next_date.

        # Reutilizamos la lógica: si se quisiera ser exacto, deberíamos replicar el cálculo
        # de días para cada iteración, pero eso puede ser demasiado complejo.
        # Aquí haremos una suposición simplificada:
        # - Obtenemos ndays inicial desde next_date (ya calculado).
        # - Para las siguientes ejecuciones, asumimos el mismo patrón:
        #   es decir, cada frecuencia en horas corresponde al mismo número de días (la razón
        #   es que no tenemos un cálculo dinámico de horas consumidas, solo uno inicial).

        # Como en next_date se calcula ndays = int(inv / prom_hours) para h/k,
        # debemos replicar esa lógica. Obtenemos ndays inicial (diferencia entre next_date y date.today())
        if system.state == 's':  # 's' representa 'Stand by'
            return 1
        
        ndays_iniciales = (initial_next_date - date.today()).days
        if ndays_iniciales < 1:
            ndays_iniciales = 1  # por si acaso, evitar division por cero u otros casos

        # Cada frecuencia en horas se traduce en freq_hours / prom_hours días aproximadamente.
        equipo = ruta.equipo
        prom_hours = equipo.prom_hours if equipo.prom_hours else 1
        # Si prom_hours es 0 se tomó 2 por defecto en next_date.

        # Calc days_per_execution = freq_hours / prom_hours (en días)
        # Si freq_hours es ruta.frecuency
        freq_hours = ruta.frecuency
        try:
            days_per_execution = int(freq_hours / prom_hours)
            if days_per_execution < 1:
                days_per_execution = 1
        except ZeroDivisionError:
            days_per_execution = 1

        # Ahora desde current_date sumamos days_per_execution cada vez
        while current_date <= period_end:
            if current_date <= period_end:
                executions += 1
            current_date = current_date + timedelta(days=days_per_execution)

    else:
        # Si no es d, h o k, retornar 0 por ahora
        return 0

    return executions


def calcular_repeticiones2(ruta, periodo='anual'):
    periodos = {
        'trimestral': 90,
        'semestral': 180,
        'anual': 365,
        'quinquenal': 1825,  # 5 años
    }

    dias_periodo = periodos.get(periodo, 0)
    today = datetime.now().date()
    meses_ejecucion = []
    
    if ruta.control == 'd':
        frecuencia_dias = ruta.frecuency
        next_date = ruta.next_date

        repeticiones = 0
        if next_date and next_date <= today + timedelta(days=dias_periodo):
            repeticiones += 1
            meses_ejecucion.append(next_date.strftime('%B %Y'))
            remaining_days = (today + timedelta(days=dias_periodo) - next_date).days
            for _ in range(remaining_days // frecuencia_dias):
                next_date += timedelta(days=frecuencia_dias)
                meses_ejecucion.append(next_date.strftime('%B %Y'))
            repeticiones += remaining_days // frecuencia_dias
    
    elif ruta.get_control_display() == 'Horas':
        diferencia_dias = (ruta.next_date - ruta.intervention_date).days
        repeticiones = 0

        if diferencia_dias > 0:
            next_date = ruta.next_date
            if next_date and next_date <= (today + timedelta(days=dias_periodo)):
                repeticiones += 1
                meses_ejecucion.append(next_date.strftime('%B %Y'))
                remaining_days = (today + timedelta(days=dias_periodo) - next_date).days
                for _ in range(remaining_days // diferencia_dias):
                    next_date += timedelta(days=diferencia_dias)
                    meses_ejecucion.append(next_date.strftime('%B %Y'))
                repeticiones += remaining_days // diferencia_dias

    return repeticiones, meses_ejecucion


def traductor(word):
    context = {
        "January": "Enero",
        "February": "Febrero",
        "March": "Marzo",
        "April": "Abril",
        "May": "Mayo",
        "June": "Junio",
        "July": "Julio",
        "August": "Agosto",
        "September": "Septiembre",
        "October": "Octubre",
        "November": "Noviembre",
        "December": "Diciembre"
    }
    return context[word]


def truncate_text(text, length=45):
    if len(text) > length:
        return text[:length] + '...'
    return text


def calculate_status_code(t):
    ot_tasks = Task.objects.filter(ot=t)

    earliest_start_date = min(t.start_date for t in ot_tasks)
    latest_final_date = max(t.final_date for t in ot_tasks)

    today = date.today()

    if latest_final_date < today:
        return 0
    elif earliest_start_date < today < latest_final_date:
        return 1
    elif earliest_start_date > today:
        return 2
    return None


def update_equipo_code(old_code):
    try:
        with transaction.atomic():
            equipo = Equipo.objects.select_for_update().get(code=old_code)
            system = equipo.system
            asset_abbreviation = system.asset.abbreviation
            tipo = equipo.tipo.upper()
 
            generated_code = generate_equipo_code(asset_abbreviation, tipo)

           # Actualizar el código del equipo
            equipo.code = generated_code
            equipo.save()

            # Actualizar todas las relaciones donde esté relacionado el equipo
            print(f"Actualizando código de equipo de '{old_code}' a '{generated_code}'...")

            # HistoryHour
            updated_history_hours = HistoryHour.objects.filter(component__code=old_code).update(component=equipo)
            print(f"HistoryHour actualizado: {updated_history_hours} registros")

            # Suministro
            updated_suministros = Suministro.objects.filter(equipo__code=old_code).update(equipo=equipo)
            print(f"Suministro actualizado: {updated_suministros} registros")

            # Ruta
            updated_rutas = Ruta.objects.filter(equipo__code=old_code).update(equipo=equipo)
            print(f"Ruta actualizado: {updated_rutas} registros")

            # DailyFuelConsumption
            updated_daily_fuel = DailyFuelConsumption.objects.filter(equipo__code=old_code).update(equipo=equipo)
            print(f"DailyFuelConsumption actualizado: {updated_daily_fuel} registros")

            # FailureReport
            updated_failure_reports = FailureReport.objects.filter(equipo__code=old_code).update(equipo=equipo)
            print(f"FailureReport actualizado: {updated_failure_reports} registros")

            # Megger
            updated_megger = Megger.objects.filter(equipo__code=old_code).update(equipo=equipo)
            print(f"Megger actualizado: {updated_megger} registros")

            # Preoperacional
            updated_preoperacional = Preoperacional.objects.filter(vehiculo__code=old_code).update(vehiculo=equipo)
            print(f"Preoperacional actualizado: {updated_preoperacional} registros")

            # PreoperacionalDiario
            updated_preoperacional_diario = PreoperacionalDiario.objects.filter(vehiculo__code=old_code).update(vehiculo=equipo)
            print(f"PreoperacionalDiario actualizado: {updated_preoperacional_diario} registros")

            # Transferencia
            updated_transferencia_origen = Transference.objects.filter(origen__equipos__code=old_code).update(origen=equipo.system)
            updated_transferencia_destino = Transference.objects.filter(destino__equipos__code=old_code).update(destino=equipo.system)
            print(f"Transferencia actualizado: {updated_transferencia_origen + updated_transferencia_destino} registros")

            # DarBaja
            updated_darbaja = DarBaja.objects.filter(equipo__code=old_code).update(equipo=equipo)
            print(f"DarBaja actualizado: {updated_darbaja} registros")

            # Image
            updated_images = Image.objects.filter(equipo__code=old_code).update(equipo=equipo)
            print(f"Image actualizado: {updated_images} registros")

            # Document
            updated_documents = Document.objects.filter(equipo__code=old_code).update(equipo=equipo)
            print(f"Document actualizado: {updated_documents} registros")

            Equipo.objects.filter(code=old_code).exclude(pk=equipo.pk).delete()

            print(f"El código del equipo '{old_code}' ha sido actualizado a '{generated_code}' en todas las tablas relacionadas.")
    
    except Equipo.DoesNotExist:
        print(f"El equipo con código '{old_code}' no existe.")
    except Exception as e:
        print(f"Ha ocurrido un error: {str(e)}")


def consumos_combustible_asset(asset):
    hoy = timezone.now().date() - timedelta(days=1)
    hace_30_dias = hoy - timedelta(days=30)
    consumos = Transaction.objects.filter(
        suministro__asset=asset,
        suministro__item__id=132,
        fecha__range=[hace_30_dias, hoy],
        tipo='c'
    ).values('fecha').annotate(total_consumido=Sum('cant')).order_by('fecha')
    fechas = [(hace_30_dias + timedelta(days=i)).strftime('%d/%m') for i in range(31)]
    consumos_dict = {consumo['fecha'].strftime('%d/%m'): consumo['total_consumido'] for consumo in consumos}
    consumos_grafica = [consumos_dict.get(fecha, 0) for fecha in fechas]
    return consumos_grafica


def create_transaction(data):
    try:
        # Use get_or_create to avoid duplicates
        obj, created = Transaction.objects.get_or_create(
            suministro=data['suministro'],
            fecha=data['fecha'],
            tipo=data['tipo'],
            defaults={
                'cant': data['cant'],
                'user': data['user'],
                'motivo': data['motivo'],
            }
        )
        return created
    except Exception as e:
        print(f'Error creating transaction: {e}')
        return False
    

def get_cargo(full_name):
    """
    Recibe el nombre completo del usuario y devuelve el cargo desde UserProfile.
    """
    try:
        # Separar el nombre completo en nombre y apellido
        nombre, apellido = full_name.split(' ', 1)
        # Buscar el usuario por nombre y apellido (case-insensitive)
        user = User.objects.get(first_name__iexact=nombre, last_name__iexact=apellido)
        # Retornar el cargo desde UserProfile
        return user.profile.cargo if hasattr(user, 'profile') else ''
    except (User.DoesNotExist, ValueError, UserProfile.DoesNotExist):
        return ''




#     return queryset


import requests
from io import BytesIO
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment

def excel_template(
        ws,
        num_cols,
        logo_url,
        title_text,
        formato_code="FR-SP-CM-25",
        version="009",
        fecha_update="13/09/2024"
    ):
    """
    Crea un encabezado en la hoja `ws`, adaptando merges según `num_cols`.
    
    :param ws: Worksheet de openpyxl
    :param num_cols: Número total de columnas efectivas en tu tabla
    :param logo_url: URL de la imagen del logo
    :param title_text: Texto del título principal (por ej. "ACTA DE INVENTARIO FÍSICO")
    :param formato_code: Ej. "FR-SP-CM-25"
    :param version: Ej. "009"
    :param fecha_update: Ej. "13/09/2024"
    """

    # 1) Calcular “breakpoints” en columnas (por ejemplo 1er tercio, 2do tercio, etc.)
    #    En este ejemplo dividimos en 3 secciones: logo, título, info de formato
    #    El “num_cols” se basa en cuántas columnas tendrá tu tabla “real”.

    # Evitar divisiones por 0
    if num_cols < 3:
        num_cols = 3

    # Indices (1-based) de dónde romper
    # Sección A = [1..logo_end]
    # Sección B = [logo_end+1..title_end]
    # Sección C = [title_end+1..num_cols]
    logo_end = max(1, num_cols // 3)        # 1/3
    title_end = max(logo_end + 1, 2*(num_cols // 3))  # 2/3

    # 2) Insertar Logo en la celda A1 (si se desea).
    #    O en la celda (fila=1, col=1).
    try:
        img_response = requests.get(logo_url)
        if img_response.status_code == 200:
            image_data = BytesIO(img_response.content)
            image_data.name = "Logo.png"
            image_data.seek(0)
            img = Image(image_data)
            img.width = 300
            img.height = 80
            ws.add_image(img, "A1")
        else:
            print("No se pudo descargar la imagen del logo.")
    except Exception as e:
        print(f"Error descargando logo: {e}")

    # 3) Merge para la sección del LOGO (fila 1 a 3, col 1 a logo_end)
    #    Nota: Las celdas se expresan en “A1” => (row=1, col=1)
    #    to “???” => (row=3, col=logo_end).
    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=logo_end)

    # 4) Merge para el TÍTULO (fila 1 a 3, col logo_end+1 a title_end)
    ws.merge_cells(
        start_row=1,
        start_column=logo_end+1,
        end_row=3,
        end_column=title_end
    )
    cell_title = ws.cell(row=1, column=logo_end+1)
    cell_title.value = title_text
    cell_title.font = Font(bold=True, size=16, name='Arial')
    cell_title.alignment = Alignment(horizontal="center", vertical="center")

    # 5) Merge para la sección del FORMATO (fila 1, col=title_end+1..num_cols)
    #    Vamos a ocupar 3 filas => Formato, Versión, Fecha
    #    Por ejemplo, la fila 1 => Formato
    ws.merge_cells(
        start_row=1,
        start_column=title_end+1,
        end_row=1,
        end_column=num_cols
    )
    cell_format = ws.cell(row=1, column=title_end+1)
    cell_format.value = f"FORMATO: {formato_code}"
    cell_format.font = Font(bold=True, name='Arial')
    cell_format.alignment = Alignment(horizontal="center", vertical="center")

    # Fila 2 => Versión
    ws.merge_cells(
        start_row=2,
        start_column=title_end+1,
        end_row=2,
        end_column=num_cols
    )
    cell_version = ws.cell(row=2, column=title_end+1)
    cell_version.value = f"VERSIÓN: {version}"
    cell_version.font = Font(bold=True, name='Arial')
    cell_version.alignment = Alignment(horizontal="center", vertical="center")

    # Fila 3 => Fecha
    ws.merge_cells(
        start_row=3,
        start_column=title_end+1,
        end_row=3,
        end_column=num_cols
    )
    cell_fecha = ws.cell(row=3, column=title_end+1)
    cell_fecha.value = f"FECHA DE ACTUALIZACIÓN: {fecha_update}"
    cell_fecha.font = Font(bold=True, name='Arial')
    cell_fecha.alignment = Alignment(horizontal="center", vertical="center")

    # Ajustar alto de filas
    for r in range(1, 4):
        ws.row_dimensions[r].height = 20

    # (Opcional) Ajustar ancho de columnas si deseas:
    # p.ej. ancho de las primeras column, etc.

    # Listo, retornamos la hoja si hace falta
    return ws



from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import requests

def export_inventario_excel(request, abbreviation):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Supongamos que tu tabla tendrá 12 columnas:
    total_columns = 2

    # Llamamos a la función para crear el encabezado
    excel_template(
        ws=ws,
        num_cols=total_columns,
        logo_url="https://hivik.s3.us-east-2.amazonaws.com/static/Logo.png",
        title_text="ACTA DE INVENTARIO FÍSICO",  # u otro
        formato_code="FR-SP-CM-25",
        version="009",
        fecha_update="13/09/2024"
    )

    # (Ahora en la fila 5 o 6 empiezas a poner los headers de tu tabla normal)
    # ...
    # Ejemplo:
    # ws.cell(row=5, column=1, value="CÓDIGO")
    # ...
    # Lógica de llenado.

    # Al final, generas la response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Inventario.xlsx"'
    wb.save(response)
    return response
