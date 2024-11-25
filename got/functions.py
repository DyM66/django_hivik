from django.apps import apps
from django.contrib.auth.models import Permission
from django.contrib.contenttypes.models import ContentType
from django.db import transaction
from django.shortcuts import get_object_or_404
from django.template.loader import get_template
from django.utils import timezone
from io import BytesIO
from xhtml2pdf import pisa
from django.http import HttpResponse
from got.models import *
from collections import defaultdict
from datetime import date
import openpyxl
from openpyxl.styles import Alignment, Font
from django.contrib.auth.models import Group
import pandas as pd
from django.contrib.admin.models import LogEntry
from django.contrib.contenttypes.models import ContentType
from openpyxl.utils import get_column_letter

def remove_invalid_permissions():
    """
    Elimina los permisos asociados a los modelos que ya no existen en el proyecto.
    """
    # Obtén todos los content types registrados
    existing_content_types = ContentType.objects.all()

    # Lista para almacenar los content types que ya no existen
    invalid_content_types = []

    # Recorremos todos los content types existentes
    for content_type in existing_content_types:
        model = content_type.model
        app_label = content_type.app_label
        
        # Verifica si el modelo aún existe en la app correspondiente
        if not apps.is_installed(app_label) or not apps.get_models(include_auto_created=True):
            invalid_content_types.append(content_type)

        # También verificar si el modelo ya no existe en las apps registradas
        try:
            apps.get_model(app_label, model)
        except LookupError:
            invalid_content_types.append(content_type)

    # Elimina los permisos asociados a esos content types que ya no existen
    for content_type in invalid_content_types:
        print(f"Eliminando permisos para el modelo {content_type.model} en la app {content_type.app_label}")
        
        # Elimina los permisos asociados a este content type
        Permission.objects.filter(content_type=content_type).delete()
        
        # Elimina el content type en sí
        content_type.delete()

    print(f"Se eliminaron {len(invalid_content_types)} content types inválidos y sus permisos asociados.")

    invalid_log_entries = LogEntry.objects.filter(
        content_type_id__isnull=False
    ).exclude(content_type_id__in=ContentType.objects.values_list('id', flat=True))

    print(f"Found {invalid_log_entries.count()} invalid log entries")

    invalid_log_entries.delete()
    print("Deleted invalid log entries.")


def actualizar_rutas_dependientes(ruta):
    ruta.intervention_date = timezone.now()
    ruta.save()
    if ruta.dependencia is not None:
        actualizar_rutas_dependientes(ruta.dependencia)


def generate_pdf_content(ot):

    template_path = 'got/pdf_template.html'
    context = {'ot': ot}
    template = get_template(template_path)
    html = template.render(context)
    pdf_content = BytesIO()

    pisa.CreatePDF(html, dest=pdf_content)

    return pdf_content.getvalue()


# def render_to_pdf(template_src, context_dict={}):
#     template = get_template(template_src)
#     html = template.render(context_dict)
#     response = HttpResponse(content_type='application/pdf')
#     pisa_status = pisa.CreatePDF(html, dest=response)
#     if pisa_status.err:
#         return HttpResponse('Se encontraron errores al generar el PDF <pre>' + html + '</pre>')
#     return response


def render_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None


def copiar_rutas_de_sistema(system_id):
    sistema_origen = get_object_or_404(System, id=system_id)
    asset = sistema_origen.asset
    sistemas_destino = System.objects.filter(asset=asset).exclude(id=system_id)
    
    with transaction.atomic():
        for sistema in sistemas_destino:
            for ruta in sistema_origen.rutas.all():
                ruta_nueva = Ruta.objects.create(
                    name=ruta.name,
                    control=ruta.control,
                    frecuency=ruta.frecuency,
                    intervention_date=ruta.intervention_date,
                    astillero=ruta.astillero,
                    system=sistema,
                    # equipo=ruta.equipo.code,
                    dependencia=ruta.dependencia,
                )
                
                for tarea in ruta.task_set.all():
                    Task.objects.create(
                        ot=tarea.ot,
                        ruta=ruta_nueva,
                        responsible=tarea.responsible,
                        description=tarea.description,
                        procedimiento=tarea.procedimiento,
                        hse=tarea.hse,
                        news=tarea.news,
                        evidence=tarea.evidence,
                        start_date=tarea.start_date,
                        men_time=tarea.men_time,
                        finished=tarea.finished
                    )
    print(f"Rutas y tareas copiadas exitosamente a otros sistemas en el mismo activo {asset.name}.")

from datetime import datetime, timedelta
def calcular_repeticiones(ruta, periodo='anual'):
    periodos = {
        'trimestral': 90,
        'semestral': 180,
        'anual': 365,
        'quinquenal': 1825,  # 5 años
    }

    dias_periodo = periodos.get(periodo, 0)
    today = datetime.now().date()
    if ruta.control == 'd':
        frecuencia_dias = ruta.frecuency
        next_date = ruta.next_date

        repeticiones = 0
        if next_date and next_date <= today + timedelta(days=dias_periodo):
            repeticiones += 1
            # Calcular cuántas veces más ocurrirá la rutina dentro del periodo
            remaining_days = (today + timedelta(days=dias_periodo) - next_date).days
            repeticiones += remaining_days // frecuencia_dias
    
    elif ruta.get_control_display() == 'Horas':
        diferencia_dias = (ruta.next_date - ruta.intervention_date).days
        repeticiones = 0

        if diferencia_dias > 0:
            # Calcular la primera repetición
            next_date = ruta.next_date
            if next_date and next_date <= (today + timedelta(days=dias_periodo)):
                repeticiones += 1
                remaining_days = (today + timedelta(days=dias_periodo) - next_date).days
                repeticiones += remaining_days // diferencia_dias
    
    return repeticiones



def calcular_repeticiones2(ruta, periodo='anual'):
    periodos = {
        'trimestral': 90,
        'semestral': 180,
        'anual': 365,
        'quinquenal': 1825,  # 5 años
    }

    dias_periodo = periodos.get(periodo, 0)
    today = datetime.now().date()
    meses_ejecucion = []
    
    if ruta.control == 'd':
        frecuencia_dias = ruta.frecuency
        next_date = ruta.next_date

        repeticiones = 0
        if next_date and next_date <= today + timedelta(days=dias_periodo):
            repeticiones += 1
            meses_ejecucion.append(next_date.strftime('%B %Y'))
            remaining_days = (today + timedelta(days=dias_periodo) - next_date).days
            for _ in range(remaining_days // frecuencia_dias):
                next_date += timedelta(days=frecuencia_dias)
                meses_ejecucion.append(next_date.strftime('%B %Y'))
            repeticiones += remaining_days // frecuencia_dias
    
    elif ruta.get_control_display() == 'Horas':
        diferencia_dias = (ruta.next_date - ruta.intervention_date).days
        repeticiones = 0

        if diferencia_dias > 0:
            next_date = ruta.next_date
            if next_date and next_date <= (today + timedelta(days=dias_periodo)):
                repeticiones += 1
                meses_ejecucion.append(next_date.strftime('%B %Y'))
                remaining_days = (today + timedelta(days=dias_periodo) - next_date).days
                for _ in range(remaining_days // diferencia_dias):
                    next_date += timedelta(days=diferencia_dias)
                    meses_ejecucion.append(next_date.strftime('%B %Y'))
                repeticiones += remaining_days // diferencia_dias

    return repeticiones, meses_ejecucion


def traductor(word):
    context = {
        "January": "Enero",
        "February": "Febrero",
        "March": "Marzo",
        "April": "Abril",
        "May": "Mayo",
        "June": "Junio",
        "July": "Julio",
        "August": "Agosto",
        "September": "Septiembre",
        "October": "Octubre",
        "November": "Noviembre",
        "December": "Diciembre"
    }
    return context[word]


def truncate_text(text, length=45):
    if len(text) > length:
        return text[:length] + '...'
    return text


def calculate_status_code(t):
    ot_tasks = Task.objects.filter(ot=t)

    earliest_start_date = min(t.start_date for t in ot_tasks)
    latest_final_date = max(t.final_date for t in ot_tasks)

    today = date.today()

    if latest_final_date < today:
        return 0
    elif earliest_start_date < today < latest_final_date:
        return 1
    elif earliest_start_date > today:
        return 2

    return None


def export_rutinas_to_excel(systems, filename="rutinas.xlsx"):
    data = []

    for system in systems:
        filtered_rutas = Ruta.objects.filter(system=system).exclude(system__state__in=['x', 's']).order_by('-nivel', 'frecuency')

        for ruta in filtered_rutas:
            days_left = (ruta.next_date - datetime.now().date()).days if ruta.next_date else '---'
            data.append({
                'asset': ruta.system.asset.name,  # Añadir el nombre del Asset
                'equipo_name': ruta.equipo.name if ruta.equipo else ruta.system.name,
                'location': ruta.equipo.system.location if ruta.equipo else ruta.system.location,
                'name': ruta.name,
                'frecuency': ruta.frecuency,
                'control': ruta.get_control_display(),
                'daysleft': days_left,
                'intervention_date': ruta.intervention_date.strftime('%d/%m/%Y') if ruta.intervention_date else '---',
                'next_date': ruta.next_date.strftime('%d/%m/%Y') if ruta.next_date else '---',
                'ot_num_ot': ruta.ot.num_ot if ruta.ot else '---',
                'activity': '---',  # Placeholder para la fila de la rutina en la columna de actividades
                'responsable': ''  # Columna vacía para la fila de la rutina
            })

            tasks = Task.objects.filter(ruta=ruta)
            for task in tasks:
                responsable_name = task.responsible.get_full_name() if task.responsible else '---'
                data.append({
                    'asset': ruta.system.asset.name,  # Mantener el nombre del Asset para las tareas
                    'equipo_name': '',
                    'location': '',
                    'name': '',
                    'frecuency': '',
                    'control': '',
                    'daysleft': '',
                    'intervention_date': '',
                    'next_date': '',
                    'ot_num_ot': '',
                    'activity': f'- {task.description}',
                    'responsable': responsable_name
                })

    # Convertimos la lista de diccionarios en un DataFrame
    df = pd.DataFrame(data)
    df.columns = ['Asset', 'Equipo', 'Ubicación', 'Código', 'Frecuencia', 'Control', 'Tiempo Restante', 'Última Intervención', 'Próxima Intervención', 'Orden de Trabajo', 'Actividad', 'Responsable']

    # Generamos la respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

    return response


def update_equipo_code(old_code):

    equipo = Equipo.objects.get(code=old_code)
    system = equipo.system
    asset_abbreviation = system.asset.abbreviation
    group_number = system.group
    tipo = equipo.tipo.upper()

    similar_equipments = Equipo.objects.filter(
        code__startswith=f"{asset_abbreviation}-{group_number}-{tipo}"
    )
    sequence_number = similar_equipments.count() + 1
    sequence_str = str(sequence_number).zfill(3) 
    generated_code = f"{asset_abbreviation}-{group_number}-{tipo}-{sequence_str}"

    try:
        # Iniciar una transacción para asegurar que todo se actualice correctamente
        with transaction.atomic():
            # Obtener el equipo cuyo código se va a cambiar
            equipo = Equipo.objects.get(code=old_code)
            equipo.code = generated_code
            equipo.save()

            # Actualizar todas las relaciones donde esté relacionado el equipo
            print(f"Actualizando código de equipo de '{old_code}' a '{generated_code}'...")

            # HistoryHour
            updated_history_hours = HistoryHour.objects.filter(component__code=old_code).update(component=equipo)
            print(f"HistoryHour actualizado: {updated_history_hours} registros")

            # Suministro
            updated_suministros = Suministro.objects.filter(equipo__code=old_code).update(equipo=equipo)
            print(f"Suministro actualizado: {updated_suministros} registros")

            # Ruta
            updated_rutas = Ruta.objects.filter(equipo__code=old_code).update(equipo=equipo)
            print(f"Ruta actualizado: {updated_rutas} registros")

            # DailyFuelConsumption
            updated_daily_fuel = DailyFuelConsumption.objects.filter(equipo__code=old_code).update(equipo=equipo)
            print(f"DailyFuelConsumption actualizado: {updated_daily_fuel} registros")

            # FailureReport
            updated_failure_reports = FailureReport.objects.filter(equipo__code=old_code).update(equipo=equipo)
            print(f"FailureReport actualizado: {updated_failure_reports} registros")

            # Megger
            updated_megger = Megger.objects.filter(equipo__code=old_code).update(equipo=equipo)
            print(f"Megger actualizado: {updated_megger} registros")

            # Preoperacional
            updated_preoperacional = Preoperacional.objects.filter(vehiculo__code=old_code).update(vehiculo=equipo)
            print(f"Preoperacional actualizado: {updated_preoperacional} registros")

            # PreoperacionalDiario
            updated_preoperacional_diario = PreoperacionalDiario.objects.filter(vehiculo__code=old_code).update(vehiculo=equipo)
            print(f"PreoperacionalDiario actualizado: {updated_preoperacional_diario} registros")

            # Transferencia
            updated_transferencia_origen = Transferencia.objects.filter(origen__equipos__code=old_code).update(origen=equipo.system)
            updated_transferencia_destino = Transferencia.objects.filter(destino__equipos__code=old_code).update(destino=equipo.system)
            print(f"Transferencia actualizado: {updated_transferencia_origen + updated_transferencia_destino} registros")

            # DarBaja
            updated_darbaja = DarBaja.objects.filter(equipo__code=old_code).update(equipo=equipo)
            print(f"DarBaja actualizado: {updated_darbaja} registros")

            # Image
            updated_images = Image.objects.filter(equipo__code=old_code).update(equipo=equipo)
            print(f"Image actualizado: {updated_images} registros")

            # Document
            updated_documents = Document.objects.filter(equipo__code=old_code).update(equipo=equipo)
            print(f"Document actualizado: {updated_documents} registros")

            print(f"El código del equipo '{old_code}' ha sido actualizado a '{generated_code}' en todas las tablas relacionadas.")
    
    except Equipo.DoesNotExist:
        print(f"El equipo con código '{old_code}' no existe.")
    except Exception as e:
        print(f"Ha ocurrido un error: {str(e)}")


def operational_users(current_user):
    if current_user.groups.filter(name__in=['maq_members', 'buzos_members']).exists():
        talleres = Group.objects.get(name='serport_members')
        taller_list = list(talleres.user_set.all())
        taller_list.append(current_user)
        return User.objects.filter(id__in=[user.id for user in taller_list])
    elif current_user.groups.filter(name='super_members').exists():
        return User.objects.exclude(groups__name='gerencia')


def fechas_range():
    hoy = timezone.now().date() - timedelta(days=1)
    hace_30_dias = hoy - timedelta(days=30)
    return [(hace_30_dias + timedelta(days=i)).strftime('%d/%m') for i in range(31)]


def consumos_combustible_asset(asset):
    hoy = timezone.now().date() - timedelta(days=1)
    hace_30_dias = hoy - timedelta(days=30)
    consumos = Transaction.objects.filter(
        suministro__asset=asset,
        suministro__item__id=132,
        fecha__range=[hace_30_dias, hoy],
        tipo='c'
    ).values('fecha').annotate(total_consumido=Sum('cant')).order_by('fecha')
    fechas = [(hace_30_dias + timedelta(days=i)).strftime('%d/%m') for i in range(31)]
    consumos_dict = {consumo['fecha'].strftime('%d/%m'): consumo['total_consumido'] for consumo in consumos}
    consumos_grafica = [consumos_dict.get(fecha, 0) for fecha in fechas]
    return consumos_grafica


def horas_total_asset(asset):
    hoy = timezone.now().date() - timedelta(days=1)
    hace_30_dias = hoy - timedelta(days=30)
    systems = System.objects.filter(asset=asset)
    equipos = Equipo.objects.filter(system__in=systems)
    horas_operacion = HistoryHour.objects.filter(
        component__in=equipos,
        report_date__range=[hace_30_dias, hoy]
    ).values('report_date').annotate(total_horas=Sum('hour')).order_by('report_date')
    horas_dict = {hora['report_date'].strftime('%d/%m'): hora['total_horas'] for hora in horas_operacion}
    horas_grafica = [horas_dict.get(fecha, 0) for fecha in fechas_range()]
    return horas_grafica


def get_full_systems(asset, user):
    systems = asset.system_set.all()
    if user.groups.filter(name='buzos_members').exists():
        station = user.profile.station
        if station:
            return systems.filter(location__iexact=station)
        else:
            return System.objects.none()
    other_asset_systems = System.objects.filter(location=asset.name).exclude(asset=asset)
    return (systems.union(other_asset_systems)).order_by('group')


def get_full_systems_ids(asset, user):
    systems = asset.system_set.values_list('id', flat=True)
    other_asset_systems = System.objects.filter(location=asset.name).exclude(asset=asset).values_list('id', flat=True)

    all_systems_ids = systems.union(other_asset_systems)

    if user.groups.filter(name='buzos_members').exists():
        station = user.profile.station
        if station:
            all_systems_ids = System.objects.filter(id__in=all_systems_ids, location__iexact=station).values_list('id', flat=True)
        else:
            all_systems_ids = System.objects.none().values_list('id', flat=True)

    return all_systems_ids



def consumibles_summary(asset):

    '''
    Función utilizada para calcular la cantidad de articulos de tipo consumible que se repiten dentro de un mismo
    Asset y asociarlo a un grupo Subsystem o General.
    '''

    equipos = Equipo.objects.filter(system__asset=asset).prefetch_related('suministros__item')
    item_subsystems = defaultdict(set)
    items_by_subsystem = defaultdict(set)
    item_cant = defaultdict(set)

    inventory_counts = defaultdict(int)
    suministros = Suministro.objects.filter(asset=asset)

    for suministro in suministros:
        inventory_counts[suministro.item] += suministro.cantidad

    for equipo in equipos:
        subsystem = equipo.subsystem if equipo.subsystem else "General"
        for suministro in equipo.suministros.all():
                item_subsystems[suministro.item].add(subsystem)
                if suministro.item in item_cant:
                    item_cant[suministro.item] += suministro.cantidad
                else:
                    item_cant[suministro.item] = suministro.cantidad
    
    duplicated_items = {item for item, subsystems in item_subsystems.items() if len(subsystems) > 1}

    for equipo in equipos:
        subsystem = equipo.subsystem if equipo.subsystem else "General"
        for suministro in equipo.suministros.all():
            item_tuple = (suministro.item, item_cant[suministro.item], inventory_counts[suministro.item], item_cant[suministro.item] * 2)
            if suministro.item in duplicated_items:
                items_by_subsystem["General"].add(item_tuple)
            else:
                items_by_subsystem[subsystem].add(item_tuple)
    
    items_by_subsystem = {k: list(v) for k, v in items_by_subsystem.items() if v}

    return items_by_subsystem



def create_transaction(data):
    try:
        # Use get_or_create to avoid duplicates
        obj, created = Transaction.objects.get_or_create(
            suministro=data['suministro'],
            fecha=data['fecha'],
            tipo=data['tipo'],
            defaults={
                'cant': data['cant'],
                'user': data['user'],
                'motivo': data['motivo'],
            }
        )
        return created
    except Exception as e:
        print(f'Error creating transaction: {e}')
        return False
    

def get_cargo(full_name):
    """
    Recibe el nombre completo del usuario y devuelve el cargo desde UserProfile.
    """
    try:
        # Separar el nombre completo en nombre y apellido
        nombre, apellido = full_name.split(' ', 1)
        # Buscar el usuario por nombre y apellido (case-insensitive)
        user = User.objects.get(first_name__iexact=nombre, last_name__iexact=apellido)
        # Retornar el cargo desde UserProfile
        return user.profile.cargo if hasattr(user, 'profile') else ''
    except (User.DoesNotExist, ValueError, UserProfile.DoesNotExist):
        return ''


from datetime import datetime
import calendar
from .models import Ruta
from .forms import RutinaFilterForm

def get_filtered_rutas(asset, user, request_data=None):
    form = RutinaFilterForm(request_data, asset=asset)
    current_month_name_es = traductor(datetime.now().strftime('%B'))

    if form.is_valid():
        month = int(form.cleaned_data['month'])
        year = int(form.cleaned_data['year'])
        show_execute = form.cleaned_data.get('execute', False)
        selected_locations = form.cleaned_data.get('locations')
        current_month_name_es = traductor(calendar.month_name[month])
        
        filtered_rutas = Ruta.objects.filter(
            system__in=get_full_systems_ids(asset, user),
            system__location__in=selected_locations
        ).exclude(system__state__in=['x', 's']).order_by('-nivel', 'frecuency')

        if show_execute == 'on':
            filtered_rutas = [
                ruta for ruta in filtered_rutas
                if (ruta.next_date.month <= month and ruta.next_date.year <= year)
                or (ruta.ot and ruta.ot.state == 'x')
                or (ruta.percentage_remaining < 15)
            ]
        else:
            filtered_rutas = [
                ruta for ruta in filtered_rutas
                if (ruta.next_date.month <= month and ruta.next_date.year <= year)
                or (ruta.percentage_remaining < 15)
            ]
    else:
        filtered_rutas = Ruta.objects.filter(
            system__in=get_full_systems_ids(asset, user)
        ).exclude(system__state__in=['x', 's']).order_by('-nivel', 'frecuency')
        filtered_rutas = [
            ruta for ruta in filtered_rutas
            if (ruta.percentage_remaining < 15)
        ]

    return filtered_rutas, current_month_name_es


def pro_export_to_excel(model, headers, data, filename='export.xlsx', sheet_name=None, table_title=None):
    """
    Exporta datos a un archivo Excel.

    :param model: Modelo de Django o nombre del modelo (str).
    :param headers: Lista de nombres de encabezados de columnas.
    :param data: Lista de listas con los datos de cada fila.
    :param filename: Nombre del archivo Excel a generar.
    :param sheet_name: Nombre de la hoja (por defecto, el nombre del modelo).
    :param table_title: Título de la tabla que abarcará todas las columnas.
    """
    if sheet_name is None:
        sheet_name = model.__name__ if hasattr(model, '__name__') else str(model)

    # Crear un libro de trabajo y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Agregar el título de la tabla si se proporciona
    current_row = 1
    if table_title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        cell = ws.cell(row=1, column=1)
        cell.value = table_title
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True, size=14)
        current_row += 1

    # Agregar los encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Agregar los datos
    for row_data in data:
        current_row += 1
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = cell_value

    # Ajustar el ancho de las columnas automáticamente
    for idx, column_cells in enumerate(ws.columns, 1):
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        column_letter = get_column_letter(idx)
        ws.column_dimensions[column_letter].width = length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Preparar la respuesta HTTP
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response
