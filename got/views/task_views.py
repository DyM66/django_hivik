from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.views import generic, View
from django.http import HttpResponse

from got.models import Task, Asset
from got.utils.others import pdf_render
from dth.utils import COLOMBIA_HOLIDAYS
from got.utils.task_utils import  operational_users, filter_tasks_queryset

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict
from io import BytesIO


class AssignedTaskByUserListView(LoginRequiredMixin, generic.ListView):
    model = Task
    template_name = 'got/task_templates/tasks_pendient.html'
    paginate_by = 25

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Cargar assets (solo si tiene permiso)
        if self.request.user.has_perm('got.access_all_assets'):
            context['assets'] = Asset.objects.filter(show=True)
        else:
            context['assets'] = None
        context['all_users'] = operational_users()
        context['total_tasks'] = self.get_queryset().count()
        context['start_date'] = self.request.GET.get('start_date', '')
        context['end_date'] = self.request.GET.get('end_date', '')
        context['estado'] = self.request.GET.get('estado', '0')
        context['asset_id'] = self.request.GET.get('asset_id', '')
        context['worker'] = self.request.GET.get('worker', '')
        context['holidays'] = [d.strftime('%Y-%m-%d') for d in sorted(COLOMBIA_HOLIDAYS)]
        return context

    def get_queryset(self):
        qs = Task.objects.filter(ot__isnull=False).select_related('ot__system', 'responsible')
        qs = filter_tasks_queryset(self.request, base_queryset=qs)
        qs = qs.order_by('start_date')
        return qs


class TaskPDFView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        queryset = filter_tasks_queryset(
            request
        ).order_by('ot__system__asset__name', 'start_date')

        # Obtener los parámetros de fecha
        start = request.GET.get('start_date', '')
        end = request.GET.get('end_date', '')
        # Formar el string con el rango de fechas (separado por coma)
        date_range = f"{start},{end}" if start != end else start

        context = {
            'tasks': queryset,
            'start': start,
            'end': end,
            'finalizados': request.GET.get('show_finalizadas', '0'),
            'date_range': date_range,
        }
        return pdf_render(
            request,
            'got/ots/assigned_tasks_pdf.html',
            context,
            "REPORTE DE ACTIVIDADES.pdf"
        )


@login_required
def assignedTasks_excel(request):
    # Usar el mismo queryset filtrado
    queryset = filter_tasks_queryset(request)
    start = request.GET.get('start_date', '')
    # Ordenar por asset, OT y start_date (la misma lógica que en la vista PDF)
    queryset = queryset.order_by('ot__system__asset__name', 'ot__num_ot', 'start_date')

    # Agrupar las tareas: grouped[asset][ot] = lista de tareas
    grouped = defaultdict(lambda: defaultdict(list))
    for task in queryset:
        asset = task.ot.system.asset
        ot = task.ot
        grouped[asset][ot].append(task)

    # Crear el workbook y la hoja de cálculo
    wb = Workbook()
    ws = wb.active
    ws.title = "Actividades"

    # Definir un borde delgado para las celdas
    thin_border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )

    # === Insertar el logo de la empresa ===
    url = "https://hivik.s3.us-east-2.amazonaws.com/static/Logo.png"
    img_response = requests.get(url)
    if img_response.status_code == 200:
        image_data = BytesIO(img_response.content)
        from openpyxl.drawing.image import Image as XLImage
        img = XLImage(image_data)
        img.width = 300   # Ajusta según tus necesidades
        img.height = 80   # Ajusta según tus necesidades
        ws.add_image(img, "A1")
    else:
        print("No se pudo descargar la imagen.")

    # --- Fusionar las primeras 4 filas para el título ---
    ws.merge_cells('A1:D4')
    title_cell = ws['A1']
    title_cell.value = "LISTADO DE ACTIVIDADES PROGRAMADAS" + (f" ({start})" if start else "")
    title_cell.font = Font(bold=True, size=20, color="FFFFFF")
    title_cell.fill = PatternFill(fill_type="solid", fgColor="4d93d9")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in range(1, 5):
        ws.row_dimensions[row].height = 30

    current_row = 5  # Comenzamos el contenido a partir de la fila 5

    # Recorrer cada asset (ordenado por nombre)
    for asset in sorted(grouped.keys(), key=lambda a: a.name):
        # Encabezado del asset: fusionar columnas 1 a 5
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=4
        )
        cell_asset = ws.cell(
            row=current_row,
            column=1,
            value=f"{asset.name} ({asset.abbreviation})"
        )
        ws.row_dimensions[current_row].height = 20
        cell_asset.font = Font(bold=True, size=14)
        cell_asset.alignment = Alignment(horizontal="left")
        # Aplicar borde a todas las celdas del rango fusionado
        for row_cells in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=1, max_col=4):
            for cell in row_cells:
                cell.border = thin_border
        current_row += 1

        # Recorrer cada OT dentro del asset (ordenado por num_ot)
        for ot in sorted(grouped[asset].keys(), key=lambda ot: ot.num_ot):
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            cell_ot = ws.cell(row=current_row, column=1, value=f"OT-{ot.num_ot}: {ot.description}")
            cell_ot.font = Font(bold=True, size=12)
            cell_ot.alignment = Alignment(horizontal="left")
            for row_cells in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=1, max_col=4):
                for cell in row_cells:
                    cell.border = thin_border
            current_row += 1

            # Encabezados para las actividades
            headers = ['Actividad', 'Responsable', 'Inicio', 'Fin']
            for col_num, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(fill_type="solid", fgColor="4d93d9")
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
            current_row += 1

            # Escribir cada actividad de esta OT
            for task in grouped[asset][ot]:
                ws.cell(
                    row=current_row,
                    column=1,
                    value=task.description
                ).border = thin_border
                responsable = (
                    f"{task.responsible.get_full_name()}"
                    if task.responsible else ""
                )
                ws.cell(
                    row=current_row, column=2,
                    value=responsable
                ).border = thin_border
                ws.cell(
                    row=current_row, column=3,
                    value=task.start_date.strftime('%d/%m/%Y')
                    if task.start_date else ""
                ).border = thin_border
                ws.cell(
                    row=current_row,
                    column=4,
                    value=task.final_date.strftime('%d/%m/%Y')
                    if task.final_date else ""
                ).border = thin_border
                current_row += 1

    # Ajustar el ancho de las columnas (evitando errores con celdas fusionadas)
    for i in range(1, ws.max_column + 1):
        col_letter = get_column_letter(i)
        max_length = 0
        for cell in ws[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    ws.column_dimensions['A'].width = 30

    for cell in ws['A']:
        cell.alignment = Alignment(wrap_text=True)

    # Exportar el workbook a BytesIO y retornar la respuesta
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = "actividades.xlsx"
    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename={filename}'
    return response
