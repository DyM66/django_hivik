import base64
import io
import json
import logging
import os
import requests
import uuid
import zipfile

from datetime import timedelta, date, datetime
from decimal import Decimal
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import permission_required, login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin, PermissionRequiredMixin
from django.contrib.auth.models import Group
from django.contrib.auth.views import PasswordResetView
from django.core.files.base import ContentFile
from django.core.mail import EmailMessage
from django.core.paginator import Paginator
from django.db.models import Count, Q, OuterRef, Subquery, F, ExpressionWrapper, DateField, Prefetch, Case, When, IntegerField, BooleanField, Value
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse, HttpResponseForbidden, HttpResponseNotFound
from django.shortcuts import render, get_object_or_404, redirect
from django.template.loader import render_to_string
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.translation import gettext as _
from django.views import generic, View
from django.views.generic import TemplateView
from django.views.generic.edit import CreateView, UpdateView, DeleteView

from .utils import *
from .models import *
from .forms import *
from inv.models import Transaction, Solicitud
from mto.utils import record_execution, get_filtered_rutas
from meg.models import Megger
from ope.models import Operation

from io import BytesIO
from weasyprint import HTML, CSS
import weasyprint
from taggit.models import Tag 
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)
TODAY = timezone.now().date()

AREAS = {
    'a': 'Barcos',
    'c': 'Barcazas',
    'o': 'Oceanografía',
    'l': 'Locativo',
    'v': 'Vehiculos',
    'x': 'Apoyo',
}

'ASSETS VIEWS'
class AssetsListView(LoginRequiredMixin, TemplateView):
    template_name = 'got/assets/asset_list.html'

    def dispatch(self, request, *args, **kwargs):
        """Redirecciones personalizadas según el grupo del usuario."""
        user = request.user
        if user.groups.filter(name__in=['maq_members', 'buzos_members']).exists():
            asset = Asset.objects.filter(models.Q(supervisor=request.user) | models.Q(capitan=request.user)).first()
            if asset:
                return redirect('got:asset-detail', pk=asset.abbreviation)
        elif user.groups.filter(name='serport_members').exists():
            return redirect('got:my-tasks')
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        assets = Asset.objects.filter(show=True).annotate(
                total_failures_open=Count('system__equipos__failurereport', filter=Q(system__equipos__failurereport__closed=False), distinct=True),
                total_failures_process=Count('system__equipos__failurereport', filter=Q(system__equipos__failurereport__closed=False, system__equipos__failurereport__related_ot__isnull=False), distinct=True),
                total_ots_execution=Count('system__ot', filter=Q(system__ot__state='x'), distinct=True)
            ).order_by('area', 'name')

        for asset in assets:
            current_ops = Operation.objects.filter(asset=asset, start__lte=TODAY, end__gte=TODAY).order_by('start')

            if current_ops.exists():
                operation = current_ops.first()
                asset.operation_name = operation.proyecto
                asset.is_current_operation = True
            else:
                upcoming_ops = Operation.objects.filter(asset=asset, start__gt=TODAY).order_by('start')

                if upcoming_ops.exists():
                    operation = upcoming_ops.first()
                    asset.operation_name = operation.proyecto
                    asset.is_current_operation = False
                else:
                    asset.operation_name = None
                    asset.is_current_operation = False

        context['assets_by_area'] = {area_name: [asset for asset in assets if asset.area == area_code] for area_code, area_name in AREAS.items()}
        context['area_icons'] = {'a': 'fa-ship', 'c': 'fa-solid fa-ferry', 'o': 'fa-water', 'l': 'fa-building', 'v': 'fa-car', 'x': 'fa-cogs'}
        context['places'] = Place.objects.all()
        context['supervisors'] = User.objects.filter(groups__name__in=['maq_members', 'mto_members'])
        context['capitanes'] = User.objects.filter(groups__name='maq_members')
        return context
    
    def post(self, request, *args, **kwargs):
        asset_pk = request.POST.get('asset_pk')
        if not asset_pk:
            messages.error(request, "No se ha especificado el asset a actualizar.")
            return redirect('got:asset-list')

        asset = get_object_or_404(Asset, pk=asset_pk)
        action = request.POST.get('action')

        if action == 'update_place':
            place_id = request.POST.get('selected_place')
            if place_id:
                place = get_object_or_404(Place, pk=place_id)
                asset.place = place
                asset.save()
                messages.success(request, "La ubicación del asset se ha actualizado correctamente.")
            else:
                messages.error(request, "No se seleccionó ninguna ubicación.")

        elif action == 'update_supervisor':
            supervisor = asset.supervisor
            if supervisor.groups.filter(name__in=['maq_members', 'buzos_members']).exists():
                first_name = request.POST.get('first_name', '').strip()
                last_name = request.POST.get('last_name', '').strip()
                if not first_name or not last_name:
                    messages.error(request, "El nombre y el apellido del supervisor son requeridos.")
                    return redirect('got:asset-list')
                supervisor.first_name = first_name
                supervisor.last_name = last_name
                supervisor.save()
                messages.success(request, "Supervisor actualizado correctamente.")
            else:
                messages.error(request, "No se puede cambiar el nombre de este usuario.")
                return redirect('got:asset-list')

        elif action == 'update_capitan':
            capitan = asset.capitan
            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            if not first_name or not last_name:
                messages.error(request, "El nombre y el apellido del capitán son requeridos.")
                return redirect('got:asset-list')
            capitan.first_name = first_name
            capitan.last_name = last_name
            capitan.save()
            messages.success(request, "Capitán actualizado correctamente.")
        else:
            messages.error(request, "Acción no reconocida.")
        return redirect('got:asset-list')


class AssetDetailView(LoginRequiredMixin, generic.DetailView):
    model = Asset
    template_name = 'got/assets/asset_detail.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        asset = self.get_object()
        user = self.request.user
        systems = get_full_systems(asset, user)
        rotativos = Equipo.objects.filter(system__in=get_full_systems_ids(asset, user), tipo='r').exists()
        
        context['rotativos'] = rotativos
        context['view_type'] = 'detail'
        context['consumibles'] = Suministro.objects.filter(asset=asset).exists()
        context['page_obj'] = systems.order_by('name')
        context['items_by_subsystem'] = consumibles_summary(asset)
        context['fechas'] = fechas_range()
        # context['consumos_grafica'] = consumos_combustible_asset(asset)
        context['horas_grafica'] = horas_total_asset(asset)
        context['sys_form'] = SysForm()
        return context

    def post(self, request, *args, **kwargs):
        asset = self.get_object()
        sys_form = SysForm(request.POST)
        
        if sys_form.is_valid():
            sys = sys_form.save(commit=False)
            sys.asset = asset
            sys.save()
            return redirect(request.path)
        else:
            context = {'asset': asset, 'sys_form': sys_form}
            return render(request, self.template_name, context)


'OTS VIEWS'
class OtListView(LoginRequiredMixin, generic.ListView):
    model = Ot
    paginate_by = 18
    template_name = 'got/ots/ot_list.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        asset_id = self.request.GET.get('asset_id')
        if asset_id:
            systems = System.objects.filter(asset_id=asset_id)
            context['systems'] = systems
            context['selected_asset_id'] = asset_id
        else:
            context['systems'] = System.objects.none()
            context['selected_asset_id'] = None

        user = self.request.user
        user_groups = user.groups.values_list('name', flat=True)

        if 'buzos_members' in user_groups:
            info_filter = Asset.objects.filter(area='b', show=True)
        elif 'maq_members' in user_groups:
            info_filter = Asset.objects.none() 
        else:
            info_filter = Asset.objects.filter(show=True)
        context['asset'] = info_filter

        super_group = Group.objects.get(name='mto_members')
        users_in_group = super_group.user_set.all()
        context['mto_members'] = users_in_group

        queryset = self.get_queryset()
        total_ots_en_ejecucion = queryset.filter(state='x').count()
        context['total_ots_en_ejecucion'] = total_ots_en_ejecucion
        return context

    def get_queryset(self):
        queryset = Ot.objects.all().select_related('system', 'system__asset')
        user = self.request.user
        user_groups = user.groups.values_list('name', flat=True)

        if 'maq_members' in user_groups:
            supervised_assets = Asset.objects.filter(Q(supervisor=user) | Q(capitan=user))
            queryset = queryset.filter(system__asset__in=supervised_assets)
        elif 'buzos_members' in user_groups:
            supervised_assets = Asset.objects.filter(area='b')
            queryset = queryset.filter(system__asset__in=supervised_assets)
        elif any(group in user_groups for group in ['mto_members', 'serport_members', 'gerencia', 'operaciones']):
            pass
        else:
            queryset = queryset.none()

        state = self.request.GET.get('state')
        asset_id = self.request.GET.get('asset_id')
        responsable = self.request.GET.get('responsable')
        num_ot = self.request.GET.get('num_ot')
        system_id = self.request.GET.get('system_id')
        keyword = self.request.GET.get('keyword')

        if state:
            queryset = queryset.filter(state=state)
        if asset_id:
            queryset = queryset.filter(system__asset_id=asset_id)
        if responsable:
            queryset = queryset.filter(supervisor__icontains=responsable)
        if num_ot:
            queryset = queryset.filter(num_ot=num_ot)
        if system_id:
            queryset = queryset.filter(system_id=system_id)
        if keyword:
            queryset = queryset.filter(description__icontains=keyword)

        queryset = queryset.annotate(
            unfinished_tasks=Count('task', filter=Q(task__finished=False)),
            is_pausado=Case(
                When(Q(state='x') & Q(unfinished_tasks=0), then=Value(True)),
                default=Value(False),
                output_field=BooleanField(),
            ),
            state_order=Case(
                When(state='x', then=1),            # En ejecución
                When(is_pausado=True, then=2),      # Pausado
                When(state='a', then=3),            # Abierto
                When(state='f', then=4),            # Finalizado
                When(state='c', then=5),            # Cancelado
                default=6,
                output_field=IntegerField(),
            )
        ).order_by('state_order', '-creation_date', '-num_ot')

        return queryset


def ot_pdf(request, num_ot):
    ot_info = get_object_or_404(Ot, num_ot=num_ot)
    rutas = Ruta.objects.filter(ot=ot_info)
    rutina = rutas.exists()
    fallas = FailureReport.objects.filter(related_ot=ot_info)
    failure = fallas.exists()

    tareas = Task.objects.filter(ot=ot_info).select_related('responsible', 'responsible__profile').order_by('start_date')
    usuarios_participacion_dict = {}

    for tarea in tareas:
        if tarea.responsible:
            user = tarea.responsible
            if user.id not in usuarios_participacion_dict:
                usuarios_participacion_dict[user.id] = {
                    'nombre': f"{user.first_name} {user.last_name}",
                    'cargo': user.profile.cargo if hasattr(user, 'profile') else '',
                    'earliest_start_date': tarea.start_date,
                    'latest_final_date': tarea.final_date,
                }
            else:
                # Actualizar earliest_start_date
                if tarea.start_date and (usuarios_participacion_dict[user.id]['earliest_start_date'] is None or tarea.start_date < usuarios_participacion_dict[user.id]['earliest_start_date']):
                    usuarios_participacion_dict[user.id]['earliest_start_date'] = tarea.start_date
                # Actualizar latest_final_date
                if tarea.final_date and (usuarios_participacion_dict[user.id]['latest_final_date'] is None or tarea.final_date > usuarios_participacion_dict[user.id]['latest_final_date']):
                    usuarios_participacion_dict[user.id]['latest_final_date'] = tarea.final_date
    

    usuarios_participacion = []
    for user_id, info in usuarios_participacion_dict.items():
        if info['earliest_start_date'] and info['latest_final_date']:
            total_execution_time = info['latest_final_date'] - info['earliest_start_date']
            total_execution_time_display = f"{total_execution_time.days} días"
        else:
            total_execution_time_display = "N/A"
        
        usuarios_participacion.append({
            'nombre': info['nombre'],
            'cargo': info['cargo'],
            'start_date': info['earliest_start_date'],
            'final_date': info['latest_final_date'],
            'total': total_execution_time_display
        })

    # Aquí se agrupan las imágenes por fecha
    images_qs = Image.objects.filter(task__ot=ot_info).order_by('creation')
    has_evidence = images_qs.exists()
    evidence_by_date = defaultdict(list)
    for image in images_qs:
        evidence_by_date[image.creation].append(image.image.url)
    
    # Obtener una lista de todas las URLs de las imágenes
    # evidence_images = [image.image.url for image in images_qs]
    context = {
        'ot': ot_info,
        'fallas': fallas,
        'failure': failure,
        'rutas': rutas,
        'rutina': rutina,
        'users': usuarios_participacion,
        'has_evidence': has_evidence,
        # 'evidence_images': evidence_images,
        'evidence_by_date': dict(evidence_by_date),
        'tareas': tareas
    }
    template_path = 'got/ots/ot_pdf.html'
    download = request.GET.get('download', None)
    response = render_to_pdf(template_path, context)
    filename = f'orden_de_trabajo_{num_ot}.pdf'
    if download:
        content = f'attachment; filename="{filename}"'
    else:
        content = f'inline; filename="{filename}"'
    response['Content-Disposition'] = content
    return response


class MaintenancePlanPDFExportView(View):
    def get(self, request, asset_abbr):
        # Obtener el asset a exportar
        asset = get_object_or_404(Asset, abbreviation=asset_abbr)
        # Obtener la información necesaria, similar a tu exportación Excel
        main, filtered_rutas, exec, realized, filter_date = get_filtered_rutas(asset, request.GET)

        # Preparar contexto: incluye todos los datos que necesitas
        context = {
            'asset': asset,
            'asset_abbr': asset_abbr,
            'filtered_rutas': filtered_rutas,
            'TODAY': date.today(),
            # Puedes agregar más datos (por ejemplo, datos de financiación, etc.)
        }
        
        # Renderizamos el template a una cadena HTML
        html_string = render_to_string("got/maintenance_plan_pdf.html", context)
        
        # Generamos el PDF usando WeasyPrint; base_url es importante para que se resuelvan correctamente rutas de imágenes o CSS
        pdf_file = weasyprint.HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()
        
        # Configuramos la respuesta como PDF
        response = HttpResponse(pdf_file, content_type="application/pdf")
        filename = f"maintenance_plan_{asset_abbr}_{date.today().strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class MaintenancePlanExcelExportView(View):
    def get(self, request, asset_abbr):
        asset = get_object_or_404(Asset, abbreviation=asset_abbr)
        main, filtered_rutas, exec, realized, filter_date = get_filtered_rutas(asset, request.GET)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rutinas de Mantenimiento"

        # Definir un borde delgado en color negro
        thin_border = Border(left=Side(style='thin', color="000000"), right=Side(style='thin', color="000000"), top=Side(style='thin', color="000000"), bottom=Side(style='thin', color="000000"))

        # Fila 1: Título
        ws.merge_cells('A1:B4')
        ws.merge_cells('C1:F2')
        ws['C1'] = "PROGRAMACIÓN RUTINAS DE MANTENIMIENTO"
        ws['C1'].font = Font(bold=True, size=16, name='Arial')
        ws['C1'].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells('C3:C4')
        ws['C3'] = "FORMATO:"
        ws['C3'].font = Font(bold=True, size=11, name='Arial')
        ws['C3'].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells('D3:D4')
        ws['D3'] = "VERSION 001"
        ws['D3'].font = Font(bold=True, size=11, name='Arial')
        ws['D3'].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells('E3:F4')
        ws['E3'] = "FECHA DE ACTUALIZACIÓN: 18/02/2025"
        ws['E3'].font = Font(bold=True, size=11, name='Arial')
        ws['E3'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 20

        ws.merge_cells('A5:F5')
        ws['A5'] = f"AREA: {asset.name}"
        ws['A5'].font = Font(bold=True, name='Arial')
        ws['A5'].fill = PatternFill(fill_type="solid", fgColor="92cddc")
        
        ws.merge_cells('A6:F6')
        ws['A6'] = f"FECHA: {TODAY.strftime('%A, %d de %B de %Y')}"
        ws['A6'].font = Font(bold=True, name='Arial') # 
        ws['A6'].fill = PatternFill(fill_type="solid", fgColor="92cddc")

        # --- Agregar bordes al área de cabecera (título y datos generales) ---
        header_ranges = ["A1:B4", "C1:F2", "C3:C4", "D3:D4", "E3:F4", "A5:F5", "A6:F6"]
        for merged_range in header_ranges:
            for row in ws[merged_range]:
                for cell in row:
                    cell.border = thin_border

        start_row = 7
        headers = [
            "ITEM", "EQUIPO", "CÓDIGO", "FRECUENCIA", "TIEMPO RESTANTE", "PRÓXIMA INTERVENCIÓN"
        ]
        ws.append(headers)

        # Estilos para el encabezado
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="ffffff", name='Arial')
            cell.fill = PatternFill(fill_type="solid", fgColor="4d93d9")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Construimos las filas con la información de cada equipo
        rutas_data = []
        counter = 1
        for ruta in filtered_rutas:
            equipo = ruta.equipo.name if ruta.equipo else ruta.system.name
            row = [
                counter,
                equipo,
                ruta.name,
                f"{ruta.frecuency} {ruta.get_control_display()}",
                f"{ruta.daysleft} {ruta.get_control_display()}",
                ruta.next_date,
            ]
            rutas_data.append(row)
            counter += 1

        # Escribir los datos y asignar el borde negro a cada celda de la tabla
        current_row = start_row + 1
        for i, row_data in enumerate(rutas_data):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = cell_value
                cell.font = Font(name='Arial')
                cell.border = thin_border
            current_row += 1
            
            ruta = filtered_rutas[i]
            for j, task in enumerate(Task.objects.filter(ruta=ruta), 1):
                task_data = [
                    float(f"{row_data[0]}.{j}"),
                    task.description
                ]
                cell = ws.cell(row=current_row, column=1)
                cell.value = task_data[0]
                cell.font = Font(name='Arial')
                cell.border = thin_border

                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=6)
                cell = ws.cell(row=current_row, column=2)
                cell.value = task_data[1]
                cell.font = Font(name='Arial', italic=True)
                cell.fill = PatternFill(fill_type="solid", fgColor="eeece1")
                cell.border = thin_border
                
                # Obtener la referencia del rango fusionado, por ejemplo "B{current_row}:F{current_row}"
                merged_range = f"{get_column_letter(2)}{current_row}:{get_column_letter(6)}{current_row}"

                # Iterar sobre cada celda en el rango fusionado y asignar el borde
                for row in ws[merged_range]:
                    for cell in row:
                        cell.border = thin_border

                current_row += 1


        for column in ws.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width         

        # === Sección de Observaciones y Cierre ===
        current_row = ws.max_row + 1
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        cell = ws.cell(row=current_row, column=1)
        cell.value = "¿DURANTE LA INSPECCIÓN SE PRESENTARON INCONSISTENCIAS?"
        cell.font = Font(bold=True, name='Arial')

        for col in range(1, 7):
            ws.cell(row=current_row, column=col).border = thin_border

        current_row += 1
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        cell = ws.cell(row=current_row, column=1)
        cell.value = "DESCRIPCIÓN DE LAS INCONSISTENCIAS ENCONTRADAS (SI APLICA):"
        cell.alignment = Alignment(horizontal="left", vertical="top")
        cell.font = Font(bold=True, name='Arial')
        ws.row_dimensions[current_row].height = 90

        for col in range(1, 7):
            ws.cell(row=current_row, column=col).border = thin_border


        current_row += 1
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        cell = ws.cell(row=current_row, column=1)
        cell.value = "RESPONSABLES"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(fill_type="solid", fgColor="4d93d9")
        cell.font = Font(bold=True, name='Arial')

        for col in range(1, 7):
            ws.cell(row=current_row, column=col).border = thin_border
        
        # Sección de firmas
        current_row = ws.max_row + 1
        ws.row_dimensions[current_row].height = 90

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        cell = ws.cell(row=current_row, column=1)
        cell.value = "Firma de receptor"
        cell.alignment = Alignment(horizontal="center", vertical="bottom")
        cell.font = Font(bold=True, name='Arial')

        ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
        cell = ws.cell(row=current_row, column=4)
        cell.value = "Firma de supervisor de Mantenimiento"
        cell.alignment = Alignment(horizontal="center", vertical="bottom")
        cell.font = Font(bold=True, name='Arial')

        for col in range(1, 3):
            ws.cell(row=current_row, column=col).border = thin_border

        for col in range(4, 7):
            ws.cell(row=current_row, column=col).border = thin_border

        current_row = ws.max_row + 1
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        cell = ws.cell(row=current_row, column=1)
        cell.value = "Nombre Completo"
        cell.alignment = Alignment(horizontal="left")
        cell.font = Font(bold=True, name='Arial')

        ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
        cell = ws.cell(row=current_row, column=4)
        cell.value = "Nombre Completo"
        cell.alignment = Alignment(horizontal="left")
        cell.font = Font(bold=True, name='Arial')

        for col in range(1, 4):
            ws.cell(row=current_row, column=col).border = thin_border

        for col in range(4, 7):
            ws.cell(row=current_row, column=col).border = thin_border

        current_row = ws.max_row + 1
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        cell = ws.cell(row=current_row, column=1)
        cell.value = "Cédula No."
        cell.alignment = Alignment(horizontal="left")
        cell.font = Font(bold=True, name='Arial')

        ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=6)
        cell = ws.cell(row=current_row, column=4)
        cell.value = "Cédula No."
        cell.alignment = Alignment(horizontal="left")
        cell.font = Font(bold=True, name='Arial')

        for col in range(1, 4):
            ws.cell(row=current_row, column=col).border = thin_border

        for col in range(4, 7):
            ws.cell(row=current_row, column=col).border = thin_border

        current_row += 1
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        cell = ws.cell(row=current_row, column=1)
        cell.value = '“Este documento es propiedad de "SERPORT S.A.S" Se prohíbe su reproducción parcial o total.”'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(fill_type="solid", fgColor="4d93d9")
        cell.font = Font(name='Arial', italic=True, size=10)

        for col in range(1, 7):
            ws.cell(row=current_row, column=col).border = thin_border

        current_row += 1
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        cell = ws.cell(row=current_row, column=1)
        cell.value = '“Este documento fue generado por el software GOT - SERPORT para el departamento de mantenimiento.”'
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(fill_type="solid", fgColor="4d93d9")
        cell.font = Font(name='Arial', italic=True, size=12)

        for col in range(1, 7):
            ws.cell(row=current_row, column=col).border = thin_border

        # === Insertar la imagen desde la URL ===
        url = "https://hivik.s3.us-east-2.amazonaws.com/static/Logo.png"
        img_response = requests.get(url)
        if img_response.status_code == 200:
            image_data = BytesIO(img_response.content)
            from openpyxl.drawing.image import Image  # Asegúrate de tener este import
            img = Image(image_data)
            # Asignar dimensiones a la imagen para evitar el error
            img.width = 300   # Ajusta según tus necesidades
            img.height = 80   # Ajusta según tus necesidades
            # Insertar la imagen en la hoja de Equipos, por ejemplo en la celda A1
            ws.add_image(img, "A1")
            # ws_equips.row_dimensions[1].height = 30
            print(ws.row_dimensions[1].height)

        else:
            print("No se pudo descargar la imagen.")

        # Configurar respuesta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"rutinas_{asset.abbreviation}_{date.today().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class AssetDocumentsView(View):
    form_class = DocumentForm
    template_name = 'got/assets/add-document.html'

    def get_context_data(self, request, asset, form=None, keyword=None):
        systems = asset.system_set.all()
        ots = Ot.objects.filter(system__in=systems)
        equipos = Equipo.objects.filter(system__in=systems)
        all_docs = Document.objects.filter( Q(asset=asset) | Q(ot__in=ots) | Q(equipo__in=equipos)).distinct()

        if keyword:
            all_docs = all_docs.filter(description__icontains=keyword)

        paginator = Paginator(all_docs, 50)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

        today = date.today()
        for doc in page_obj:
            doc.is_expired = False
            if doc.date_expiry and doc.date_expiry < today:
                doc.is_expired = True

        existing_tags = Tag.objects.annotate(num_docs=Count('taggit_taggeditem_items')).filter(num_docs__gt=0).order_by('name')
        if form is None:
            form = self.form_class()

        context = {
            'asset': asset,
            'documents': page_obj,
            'form': form,
            'today': today,
            'is_paginated': page_obj.has_other_pages(),
            'page_obj': page_obj,
            'existing_tags': existing_tags,
            'request': request  
        }
        return context

    def get(self, request, abbreviation):
        asset = get_object_or_404(Asset, abbreviation=abbreviation)
        keyword = request.GET.get('keyword', '').strip()
        context = self.get_context_data(request, asset, keyword=keyword)
        return render(request, self.template_name, context)

    def post(self, request, abbreviation):
        asset = get_object_or_404(Asset, abbreviation=abbreviation)
        form = self.form_class(request.POST, request.FILES)
        form.instance.asset = asset
        if form.is_valid():
            document = form.save(commit=False)
            document.asset = asset
            document.uploaded_by = request.user
            document.save()
            form.save_m2m()
            return redirect('got:asset-documents', abbreviation=abbreviation)
        else:
            messages.error(request, 'Error al agregar el documento. Por favor, revisa los campos.')
            keyword = request.GET.get('keyword', '').strip()
            context = self.get_context_data(request, asset, form=form, keyword=keyword)
            return render(request, self.template_name, context)


@login_required
def edit_document(request, pk):
    doc = get_object_or_404(Document, pk=pk)
    if request.method == 'POST':
        form = DocumentEditForm(request.POST, instance=doc)
        if form.is_valid():
            form.save()
            if doc.asset:
                abbr = doc.asset.abbreviation
            elif doc.ot:
                abbr = doc.ot.system.asset.abbreviation
            else:
                abbr = doc.equipo.system.asset.abbreviation
            return redirect('got:asset-documents', abbreviation=abbr)
    if doc.asset:
        abbr = doc.asset.abbreviation
    elif doc.ot:
        abbr = doc.ot.system.asset.abbreviation
    elif doc.equipo:
        abbr = doc.equipo.system.asset.abbreviation
    return redirect('got:asset-documents', abbreviation=abbr)


@login_required
def delete_document(request, pk):
    doc = get_object_or_404(Document, pk=pk)
    if request.method == 'POST':
        if doc.asset:
            abbr = doc.asset.abbreviation
        elif doc.ot:
            abbr = doc.ot.system.asset.abbreviation
        elif doc.equipo:
            abbr = doc.equipo.system.asset.abbreviation
        doc.delete()
        return redirect('got:asset-documents', abbreviation=abbr)
    return redirect('got:asset-documents', abbreviation='AAA')


'SYSTEMS VIEW'
class SysDetailView(LoginRequiredMixin, generic.DetailView):
    model = System
    template_name = "got/systems/system_base.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        system = self.get_object()

        next_url = self.request.GET.get('next')
        if not next_url:
            next_url = reverse('got:asset-detail', args=[system.asset.abbreviation])
        context['next_url'] = next_url

        main_equipos = Equipo.objects.filter(system=system, related__isnull=True)
        # Para cada equipo, obtenemos sus dependientes
        equipos_con_dependientes = []
        for eq in main_equipos:
            dependientes = eq.related_with.all()
            equipos_con_dependientes.append({'principal': eq, 'dependientes': dependientes,})
        context['equipos_con_dependientes'] = equipos_con_dependientes

        rutinas = Ruta.objects.filter(system=system)
        context['rutinas'] = rutinas
        return context


class SysUpdate(UpdateView):
    model = System
    form_class = SysForm
    template_name = 'got/systems/system_form.html'

    def form_valid(self, form):
        system = form.save(commit=False)
        system.modified_by = self.request.user
        system.save()
        return super().form_valid(form)


class SysDelete(DeleteView):
    model = System
    template_name = 'got/systems/system_confirm_delete.html'

    def delete(self, request, *args, **kwargs):
        system = self.get_object()
        system.modified_by = request.user
        return super().delete(request, *args, **kwargs)

    def get_success_url(self):
        asset_code = self.object.asset.abbreviation
        success_url = reverse_lazy(
            'got:asset-detail', kwargs={'pk': asset_code})
        return str(success_url)
    

class EquipoDetailView(LoginRequiredMixin, generic.DetailView):
    model = Equipo
    template_name = "got/systems/equipo_detail.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        equipo = self.get_object()
        system = equipo.system
        main_equipos = Equipo.objects.filter(system=system, related__isnull=True)
        equipos_con_dependientes = []
        for eq in main_equipos:
            dependientes = eq.related_with.all()
            equipos_con_dependientes.append({
                'principal': eq,
                'dependientes': dependientes,
            })
        context['equipos_con_dependientes'] = equipos_con_dependientes
        context['object'] = equipo
        context['suministros'] = Suministro.objects.filter(equipo=equipo)
        context['rutinas'] = Ruta.objects.filter(equipo=equipo)
        context['transferencias'] = Transferencia.objects.filter(equipo=equipo)

        previous_url = self.request.GET.get('previous_url') or self.request.META.get('HTTP_REFERER', '/')
        context['previous_url'] = previous_url
        return context
    

class EquipoPDFView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        equipo = get_object_or_404(Equipo, pk=self.kwargs['pk'])
        images = Image.objects.filter(equipo=equipo)
        related_equipos = equipo.related_with.all()
        related_related_equipos = Equipo.objects.none()
        for r in related_equipos:
            related_related_equipos = related_related_equipos | r.related_with.all()

        all_related_equipos = related_equipos | related_related_equipos

        rutinas = Ruta.objects.filter(
            Q(equipo=equipo) |
            Q(equipo__in=all_related_equipos) |
            Q(task__equipo=equipo)
        ).distinct()

        context = {
            'equipo': equipo,
            'related': all_related_equipos,
            'images': images,
            'suministros': Suministro.objects.filter(equipo=equipo),
            'rutinas': rutinas,
            'transferencias': Transferencia.objects.filter(equipo=equipo),
            'today': timezone.now().date(),  # Para mostrar la fecha
        }

        context = {
            'equipo': equipo,
            'related': related_equipos,
            'images': images,
            'suministros': Suministro.objects.filter(equipo=equipo),
            'rutinas': Ruta.objects.filter(Q(equipo=equipo) | Q(equipo__in=related_equipos) | Q(equipo__in=related_related_equipos) | Q(task__equipo=equipo)).distinct(),
            'transferencias': Transferencia.objects.filter(equipo=equipo),
            'today': timezone.now().date(),  # Para mostrar la fecha
        }
        
        # Renderizar la plantilla para el informe PDF
        html_string = render_to_string("got/systems/eq_datasheet_pdf.html", context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        
        response = HttpResponse(html.write_pdf(), content_type='application/pdf')
        filename = f"Datasheet_{equipo}.pdf"
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response


class EquipoDeleteImageView(LoginRequiredMixin, View):
    def post(self, request, pk):
        # Obtener el Equipo
        equipo = get_object_or_404(Equipo, pk=pk)
        # Obtener el ID de la imagen desde el POST
        image_id = request.POST.get('image_id')
        if not image_id:
            return JsonResponse({'error': 'No se proporcionó el ID de la imagen.'}, status=400)
        
        # Obtener la imagen y asegurarse de que pertenece al equipo
        image = get_object_or_404(Image, id=image_id, equipo=equipo)
        
        # Eliminar la imagen
        image.delete()
        
        # Devolver la cantidad actualizada de imágenes
        image_count = equipo.images.count()
        return JsonResponse({'success': True, 'image_count': image_count})
    

from django.utils.decorators import method_decorator
@method_decorator(login_required, name='dispatch')
class TaskDeleteImageView(View):
    def post(self, request, task_pk):
        # Obtener la tarea
        task = get_object_or_404(Task, pk=task_pk)
        # Obtener el ID de la imagen desde el POST
        image_id = request.POST.get("image_id")
        if not image_id:
            return JsonResponse({'error': 'No se proporcionó el ID de la imagen.'}, status=400)
        # Asegurarse de que la imagen pertenece a la tarea
        image = get_object_or_404(Image, id=image_id, task=task)
        image.delete()
        # Devolver la cantidad actualizada de imágenes asociadas a la tarea
        image_count = task.images.count()  # Asumiendo que en el modelo Task se definió related_name='images'
        return JsonResponse({'success': True, 'image_count': image_count})


class EquipoDelete(DeleteView):
    model = Equipo
    template_name = 'got/systems/equipo_confirm_delete.html'

    def dispatch(self, request, *args, **kwargs):
        # Capturar la next_url
        self.next_url = request.GET.get('next') or ''
        return super().dispatch(request, *args, **kwargs)
    
    def get_success_url(self):
        # Redirigir a next si existe
        if self.next_url:
            return self.next_url
        # fallback
        sys_code = self.object.system.id
        return reverse_lazy('got:sys-detail', kwargs={'pk': sys_code})


def add_supply_to_equipment(request, code):
    equipo = get_object_or_404(Equipo, code=code)
    if request.method == 'POST':
        form = SuministrosEquipoForm(request.POST)
        if form.is_valid():
            suministro = form.save(commit=False)
            suministro.equipo = equipo
            suministro.save()
            return redirect(reverse('got:sys-detail', args=[equipo.system.id, equipo.code]))


@login_required
def reportHoursAsset(request, asset_id):
    asset = get_object_or_404(Asset, pk=asset_id)
    today = date.today()
    dates = [today - timedelta(days=x) for x in range(90)]
    systems = get_full_systems_ids(asset, request.user)
    equipos_rotativos = Equipo.objects.filter(system__in=systems, tipo='r')
    rotativos = equipos_rotativos.exists()

    if request.method == 'POST':
        if "equipo_id" in request.POST and "report_date" in request.POST and "hour" in request.POST:
            equipo_id = request.POST.get('equipo_id')
            report_date = request.POST.get('report_date')
            hour = request.POST.get('hour')
            hist_id = request.POST.get('hist_id')

            try:
                hour_value = float(hour)
                fecha = datetime.strptime(report_date, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                messages.error(request, "Datos inválidos para actualizar horas (formato de fecha u horas).")
                return redirect(reverse('got:horas-asset', args=[asset_id]))

            # Validar rangos de hora
            if hour_value < 0 or hour_value > 24:
                messages.error(request, "El valor de horas debe estar entre 0 y 24.")
                return redirect(request.META.get(
                    'HTTP_REFERER',
                    reverse('got:horas-asset', args=[asset_id])
                ))
            
            # Buscar o crear HistoryHour
            equipo = get_object_or_404(Equipo, pk=equipo_id)
            # Verificamos si hist_id es un entero válido
            try:
                hist_id_int = int(hist_id)  # si hist_id es None, "" o "None", lanzará ValueError
                history_hour = get_object_or_404(HistoryHour, pk=hist_id_int)
            except (TypeError, ValueError):
                # hist_id no es un entero válido => creamos un nuevo registro
                history_hour = HistoryHour(component=equipo)

            history_hour.hour = hour_value
            history_hour.report_date = fecha
            history_hour.reporter = request.user
            history_hour.modified_by = request.user
            history_hour.save()

            messages.success(request, "Horas actualizadas correctamente.")
            return redirect(reverse('got:horas-asset', args=[asset_id]))

        else:
            # =============== LÓGICA DEL FORMULARIO TRADICIONAL ===============
            form = ReportHoursAsset(request.POST, equipos=equipos_rotativos, asset=asset)
            if form.is_valid():
                instance = form.save(commit=False)
                instance.reporter = request.user
                instance.modified_by = request.user
                instance.save()
                messages.success(request, "Formulario enviado correctamente.")
                return redirect(reverse('got:horas-asset', args=[asset_id]))
            else:
                # El form no es válido, se re-renderiza con errores
                messages.error(request, "Error en el formulario tradicional.")

    else:
        # GET: Simplemente creamos la instancia vacía del formulario tradicional
        form = ReportHoursAsset(equipos=equipos_rotativos, asset=asset)

    # --------------------------------------------------------------------------
    # 2. Construir la información (equipos_data y transposed_data) para la tabla
    # --------------------------------------------------------------------------
    hours = HistoryHour.objects.filter(component__system__asset=asset)[:90]

    equipos_data = []
    for equipo in equipos_rotativos:
        # Diccionario con fecha -> {hour, reporter, hist_id}
        horas_reportadas = {}
        for d in dates:
            horas_reportadas[d] = {
                'hour': 0,
                'reporter': None,
                'hist_id': None
            }

        # Consulta de HistoryHour existentes
        registros = HistoryHour.objects.filter(
            component=equipo,
            report_date__range=(dates[-1], today)
        ).select_related('reporter')

        for reg in registros:
            if reg.report_date in horas_reportadas:
                horas_reportadas[reg.report_date]['hour'] = reg.hour
                horas_reportadas[reg.report_date]['reporter'] = (reg.reporter.username if reg.reporter else None)
                horas_reportadas[reg.report_date]['hist_id'] = reg.id

        # Crear lista en el mismo orden que `dates`
        lista_horas = [horas_reportadas[d] for d in dates]

        equipos_data.append({
            'equipo': equipo,
            'horas': lista_horas,
        })

    # Transponer la data (filas=fechas, columnas=equipos)
    transposed_data = []
    for i, d in enumerate(dates):
        fila = {
            'date': d,
            'valores': []
        }
        for data in equipos_data:
            fila['valores'].append(data['horas'][i])
        transposed_data.append(fila)
    context = {
        'form': form,
        'horas': hours,
        'asset': asset,
        'equipos_data': equipos_data,
        'equipos_rotativos': equipos_rotativos,
        'dates': dates,
        'transposed_data': transposed_data,
        'rotativos': rotativos
    }
    return render(request, 'got/assets/hours_asset.html', context)


class RutaDetailView(LoginRequiredMixin, generic.DetailView):
    model = Ruta
    template_name = 'got/rutinas/ruta_detail.html'
    context_object_name = 'ruta'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ruta = self.get_object()

        all_items = Item.objects.all()
        all_services = Service.objects.all()
        context['all_items'] = all_items
        context['all_services'] = all_services
        context['tasks'] = ruta.task_set.all().order_by('-priority')
        requirements = ruta.requisitos.all()
        art_requirements = []
        svc_requirements = []
        material_types = ['m', 'h']

        for req in requirements:
            if req.tipo in material_types:
                if req.item:
                    name = str(req.item)
                else:
                    name = str(req.descripcion)
                art_requirements.append((name.lower(), req))
            elif req.tipo == 's':
                if req.service:
                    name = req.service.description.lower()
                else:
                    name = str(req.descripcion).lower()
                svc_requirements.append((name, req))

        art_requirements.sort(key=lambda x: x[0])
        svc_requirements.sort(key=lambda x: x[0])

        art_requirements = [req for _, req in art_requirements]
        svc_requirements = [req for _, req in svc_requirements]

        context['requirements_art'] = art_requirements
        context['requirements_svc'] = svc_requirements
        context['task_form'] = RutActForm(asset=ruta.system.asset)
        context['requirement_form'] = MaintenanceRequirementForm()
        context['service_form'] = ServiceForm()
        context['activity_logs'] = ActivityLog.objects.filter(model_name='Ruta', object_id=str(ruta.code)).order_by('-timestamp')
        context['task_edit_forms'] = {task.id: RutActForm(instance=task, asset=ruta.system.asset) for task in ruta.task_set.all()}
        context['requirement_edit_forms'] = {req.id: MaintenanceRequirementForm(instance=req) for req in requirements}
        context['material_types'] = material_types
        return context

    def post(self, request, *args, **kwargs):
        ruta = self.get_object()
        self.object = ruta
        action = request.POST.get('action')

        if action == 'create_task':
            task_form = RutActForm(request.POST, asset=ruta.system.asset)
            if task_form.is_valid():
                task = task_form.save(commit=False)
                task.ruta = ruta
                task.finished = False
                task.modified_by = request.user
                task.save()
                messages.success(request, 'Actividad creada exitosamente.')
                return redirect('got:ruta_detail', pk=ruta.pk)
            else:
                messages.error(request, 'Error al crear la actividad.')
                context = self.get_context_data()
                context['task_form'] = task_form
                return render(request, self.template_name, context)

        elif action == 'edit_task':
            task_id = request.POST.get('task_id')
            task = get_object_or_404(Task, id=task_id, ruta=ruta)
            task_form = RutActForm(request.POST, instance=task, asset=ruta.system.asset)
            if task_form.is_valid():
                task = task_form.save(commit=False)
                task.modified_by = request.user
                task.save()
                messages.success(request, 'Actividad actualizada exitosamente.')
            else:
                messages.error(request, 'Error al actualizar la actividad.')
            return redirect('got:ruta_detail', pk=ruta.pk)

        elif action == 'delete_task':
            task_id = request.POST.get('task_id')
            task = get_object_or_404(Task, id=task_id, ruta=ruta)
            task.delete()
            messages.success(request, 'Actividad eliminada exitosamente.')
            return redirect('got:ruta_detail', pk=ruta.pk)

        elif action == 'create_requirement':
            requirement_form = MaintenanceRequirementForm(request.POST)
            if requirement_form.is_valid():
                requirement = requirement_form.save(commit=False)
                requirement.ruta = ruta
                requirement.modified_by = request.user
                requirement.save()
                messages.success(request, 'Requerimiento creado exitosamente.')
                return redirect('got:ruta_detail', pk=ruta.pk)
            else:
                for field, errors in requirement_form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
                messages.error(request, 'Error al crear el requerimiento.')
                context = self.get_context_data()
                context['requirement_form'] = requirement_form
                return render(request, self.template_name, context)

        elif action == 'edit_requirement':
            requirement_id = request.POST.get('requirement_id')
            requirement = get_object_or_404(MaintenanceRequirement, id=requirement_id, ruta=ruta)
            requirement_form = MaintenanceRequirementForm(request.POST, instance=requirement)
            if requirement_form.is_valid():
                requirement = requirement_form.save(commit=False)
                requirement.modified_by = request.user
                requirement.save()
                messages.success(request, 'Requerimiento actualizado exitosamente.')
            else:
                messages.error(request, 'Error al actualizar el requerimiento.')
            return redirect('got:ruta_detail', pk=ruta.pk)

        elif action == 'delete_requirement':
            requirement_id = request.POST.get('requirement_id')
            requirement = get_object_or_404(MaintenanceRequirement, id=requirement_id, ruta=ruta)
            requirement.delete()
            messages.success(request, 'Requerimiento eliminado exitosamente.')
            return redirect('got:ruta_detail', pk=ruta.pk)

        elif action == 'create_service':
            service_form = ServiceForm(request.POST)
            if service_form.is_valid():
                service_form.save()
                messages.success(request, 'Servicio creado exitosamente.')
            else:
                for field, errors in service_form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
                messages.error(request, 'Error al crear el servicio.')
            return redirect('got:ruta_detail', pk=ruta.pk)

        else:
            messages.error(request, 'Acción no reconocida.')
            return redirect('got:ruta_detail', pk=ruta.pk)


class RutaCreate(CreateView):
    model = Ruta
    form_class = RutaForm
    template_name = 'got/rutinas/ruta_form.html'

    def form_valid(self, form):
        pk = self.kwargs['pk']
        system = get_object_or_404(System, pk=pk)
        form.instance.system = system
        form.instance.modified_by = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        ruta = self.object
        return reverse('got:sys-detail', args=[ruta.system.id])

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['system'] = System.objects.get(pk=self.kwargs['pk'])
        return kwargs


class RutaUpdate(UpdateView):
    model = Ruta
    form_class = RutaForm
    template_name = 'got/rutinas/ruta_form.html'

    def form_valid(self, form):
        form.instance.modified_by = self.request.user 
        return super().form_valid(form)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['system'] = self.object.system
        return kwargs


class RutaDelete(DeleteView):
    model = Ruta
    template_name = 'got/rutinas/ruta_confirm_delete.html'

    def get_success_url(self):
        sys_code = self.object.system.id
        success_url = reverse_lazy('got:sys-detail', kwargs={'pk': sys_code})
        return success_url


def acta_entrega_pdf(request, pk):
    equipo = get_object_or_404(Equipo, pk=pk)
    current_date = timezone.now().date()
    context = {
        'equipo': equipo,
        'current_date': current_date,
    }
    return render_to_pdf('got/systems/acta-entrega.html', context)
    

def asset_maintenance_pdf(request, asset_id):
    """
    Vista para generar un PDF con la información de todos los sistemas de un activo.
    """
    asset = get_object_or_404(Asset, pk=asset_id)
    systems = System.objects.filter(asset=asset).prefetch_related('equipos', 'rutas__requisitos', 'ot_set')
    start_date = timezone.now()

    sections = [
        {'title': 'Resumen', 'id': 'summary'},
        {'title': 'Información de los Sistemas', 'id': 'systems-info'},
        {'title': 'Equipos Asociados', 'id': 'associated-equipment'},
        {'title': 'Rutinas de Mantenimiento', 'id': 'maintenance-routines'},
        {'title': 'Bitácora de Mantenimientos', 'id': 'maintenance-log'},
    ]

    context = {
        'asset': asset,
        'systems': systems,
        'sections': sections,
        'current_date': start_date,
    }

    return render_to_pdf('got/systems/asset_pdf_template.html', context)


'FAILURE REPORTS VIEW'
class FailureListView(LoginRequiredMixin, generic.ListView):
    model = FailureReport
    paginate_by = 15
    template_name = 'got/fail/failurereport_list.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['assets'] = Asset.objects.filter(show=True)
        context['count_proceso'] = self.get_queryset().filter(closed=False, related_ot__isnull=False).count()
        context['count_abierto'] = self.get_queryset().filter(closed=False).count()
        return context

    def get_queryset(self):
        queryset = super().get_queryset()
        asset_id = self.request.GET.get('asset_id')
        state = self.request.GET.get('state')
        user = self.request.user

        if asset_id:
            queryset = queryset.filter(equipo__system__asset_id=asset_id)

        if state == 'abierto':
            queryset = queryset.filter(closed=False, related_ot__isnull=True)
        elif state == 'proceso':
            queryset = queryset.filter(closed=False, related_ot__isnull=False)
        elif state == 'cerrado':
            queryset = queryset.filter(closed=True)

        if self.request.user.groups.filter(name='maq_members').exists():
            supervised_assets = Asset.objects.filter(Q(supervisor=user) | Q(capitan=user))
            queryset = queryset.filter(equipo__system__asset__in=supervised_assets)
        elif self.request.user.groups.filter(name='buzos_members').exists():
            supervised_assets = Asset.objects.filter(area='b')
            queryset = queryset.filter(equipo__system__asset__in=supervised_assets)

        return queryset
    

class FailureDetailView(LoginRequiredMixin, generic.DetailView):
    model = FailureReport
    template_name = 'got/fail/failurereport_detail.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        failurereport = self.get_object()
        system = failurereport.equipo.system
        existing_ots = Ot.objects.filter(system=system)
        context['existing_ots'] = existing_ots
        return context


class FailureReportForm(LoginRequiredMixin, CreateView):
    model = FailureReport
    form_class = failureForm
    template_name = 'got/fail/failurereport_form.html'
    http_method_names = ['get', 'post']

    def send_email(self, context):
        """Sends an email compiled from the given context."""
        subject = 'Nuevo Reporte de Falla'
        email_template_name = 'emails/failure_report_email.txt'
        
        email_body_html = render_to_string(email_template_name, context)
        
        email = EmailMessage(
            subject,
            email_body_html,
            settings.EMAIL_HOST_USER,
            [user.email for user in Group.objects.get(name='mto_members').user_set.all()],
            reply_to=[settings.EMAIL_HOST_USER]
        )
        
        if self.object.evidence:
            mimetype = f'image/{self.object.evidence.name.split(".")[-1]}'
            email.attach(
                'Evidencia.' + self.object.evidence.name.split(".")[-1],
                self.object.evidence.read(),
                mimetype
            )
        
        email.send()

    def get_email_context(self):
        """Builds the context dictionary for the email."""
        impacts_display = [self.object.get_impact_display(code) for code in self.object.impact]
        return {
            'reporter': self.object.report,
            'moment': self.object.moment.strftime('%Y-%m-%d %H:%M'),
            'equipo': f'{self.object.equipo.system.asset}-{self.object.equipo.name}',
            'description': self.object.description,
            'causas': self.object.causas,
            'suggest_repair': self.object.suggest_repair,
            'impact': impacts_display, 
            'critico': 'Sí' if self.object.critico else 'No',
            'report_url': self.request.build_absolute_uri(self.object.get_absolute_url()),
        }


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        asset_id = self.kwargs.get('asset_id')
        asset = get_object_or_404(Asset, pk=asset_id)
        context['asset_main'] = asset
        if 'image_form' not in context:
            context['image_form'] = UploadImages()
        return context

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        asset_id = self.kwargs.get('asset_id')
        asset = get_object_or_404(Asset, pk=asset_id)
        form.fields['equipo'].queryset = Equipo.objects.filter(system__asset=asset)
        return form
    
    def post(self, request, *args, **kwargs):
        form = self.get_form()
        image_form = UploadImages(request.POST, request.FILES)
        
        if form.is_valid() and image_form.is_valid():
            full_name = request.user.get_full_name()
            if not full_name.strip():
                full_name = request.user.username
            form.instance.report = full_name
            form.instance.modified_by = request.user
            response = super().form_valid(form)
            # context = self.get_email_context()  
            # self.send_email(context)
            for file in request.FILES.getlist('file_field'):
                Image.objects.create(failure=self.object, image=file)
            return response
        else:
            return self.form_invalid(form)
    
    def form_invalid(self, form, **kwargs):
        context = self.get_context_data(form=form, **kwargs)
        return self.render_to_response(context)


class FailureReportUpdate(LoginRequiredMixin, UpdateView):
    model = FailureReport
    form_class = failureForm
    template_name = 'got/fail/failurereport_form.html'
    http_method_names = ['get', 'post']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        asset = self.get_object().equipo.system.asset
        context['asset_main'] = asset
        if 'image_form' not in context:
            context['image_form'] = UploadImages()
        return context

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        asset = self.get_object().equipo.system.asset
        form.fields['equipo'].queryset = Equipo.objects.filter(system__asset=asset)
        return form

    def form_valid(self, form):
        form.instance.modified_by = self.request.user
        if 'related_ot' in form.cleaned_data:
            form.instance.related_ot = form.cleaned_data['related_ot']
            if form.cleaned_data['related_ot'].state == 'f':
                form.instance.closed = True
            else:
                form.instance.closed = False
        return super().form_valid(form)
    

def fail_pdf(request, pk):
    registro = FailureReport.objects.get(pk=pk)
    context = {'fail': registro}
    return render_to_pdf('got/fail/fail_pdf.html', context)


@permission_required('got.can_see_completely')
def crear_ot_failure_report(request, fail_id):
    fail = get_object_or_404(FailureReport, pk=fail_id)
    nueva_ot = Ot(
        description=f"Reporte de falla - {fail.equipo.name}",
        state='x',
        supervisor=f"{request.user.first_name} {request.user.last_name}",
        tipo_mtto='c',
        system=fail.equipo.system,
        modified_by=request.user
    )
    nueva_ot.save()

    fail.related_ot = nueva_ot
    fail.save()
    return redirect('got:ot-detail', pk=nueva_ot.pk)


@permission_required('got.can_see_completely')
def asociar_ot_failure_report(request, fail_id):
    if request.method == 'POST':
        ot_id = request.POST.get('ot_id')
        fail = get_object_or_404(FailureReport, pk=fail_id)
        ot = get_object_or_404(Ot, num_ot=ot_id)
        fail.related_ot = ot
        if ot.state == 'f':
            fail.closed = True
        fail.save()

        return redirect('got:failure-report-detail', pk=fail_id)
    else:
        return redirect('got:failure-report-detail', pk=fail_id)


@login_required
def download_ot_task_images(request, ot_num):
    """
    Vista que recopila todas las imágenes asociadas a las actividades (tareas) de una OT
    y las empaqueta en un archivo ZIP para descarga.
    """
    ot = get_object_or_404(Ot, num_ot=ot_num)
    tasks = ot.task_set.all()
    images = Image.objects.filter(task__in=tasks)

    if not images.exists():
        return HttpResponseNotFound("No hay imágenes asociadas a las actividades de esta OT.")

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for image in images:
            try:
                response = requests.get(image.image.url)
                if response.status_code == 200:
                    file_data = response.content
                    file_name = f"{image.id}_{os.path.basename(image.image.name)}"
                    zip_file.writestr(file_name, file_data)
                else:
                    print(f"Error al descargar la imagen {image.id}: status {response.status_code}")
            except Exception as e:
                print(f"Error al procesar la imagen {image.id}: {e}")
                continue

    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer, content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="OT_{ot_num}_task_images.zip"'
    return response



class OtDetailView(LoginRequiredMixin, generic.DetailView):
    model = Ot
    template_name = 'got/ots/ot_detail.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['state_form'] = FinishOtForm()
        context['image_form'] = UploadImages()
        context['doc_form'] = DocumentForm()

        if self.request.user.groups.filter(name='mto_members').exists():
            context['task_form'] = ActForm()
        else:
            context['task_form'] = ActFormNoSup()

        context['all_tasks_finished'] = not self.get_object().task_set.filter(finished=False).exists()
        context['has_activities'] = self.get_object().task_set.exists()

        failure_report = FailureReport.objects.filter(related_ot=self.get_object())
        context['failure_report'] = failure_report
        context['failure'] = failure_report.exists()

        rutas = self.get_object().ruta_set.all()
        context['rutas'] = rutas
        context['equipos'] = set([ruta.equipo for ruta in rutas])

        system = self.get_object().system
        context['electric_motors'] = system.equipos.filter(Q(tipo='e') | Q(tipo='g'))
        context['has_electric_motors'] = system.equipos.filter(Q(tipo='e') | Q(tipo='g')).exists()
        context['megger_tests'] = Megger.objects.filter(ot=self.get_object())
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        ot = self.get_object()
        
        if 'add-doc' in request.POST:
            doc_form = DocumentForm(request.POST, request.FILES, ot=ot)
            if doc_form.is_valid():
                doc_form.save()
                return redirect(ot.get_absolute_url()) 
            else:
                print(doc_form.errors)

        if 'delete_task' in request.POST:
            task_id = request.POST.get('delete_task_id')
            task = Task.objects.get(id=task_id, ot=ot)
            task.modified_by = request.user 
            task.save()
            task.delete()
            return redirect(ot.get_absolute_url()) 
    
        task_form_class = ActForm if request.user.groups.filter(name='mto_members').exists() else ActFormNoSup
        task_form = task_form_class(request.POST, request.FILES)
        image_form = UploadImages(request.POST, request.FILES)

        if task_form.is_valid() and image_form.is_valid():
            task = task_form.save(commit=False)
            task.ot = ot
            task.user = task.responsible.get_full_name()
            task.modified_by = request.user 
            task.save()

            for file in request.FILES.getlist('file_field'):
                Image.objects.create(task=task, image=file)
        
        state_form = FinishOtForm(request.POST)
        if 'finish_ot' in request.POST and state_form.is_valid():
            # Cerrar la OT
            self.object.state = 'f'
            self.object.modified_by = request.user
            signature_image = request.FILES.get('signature_image', None)
            signature_data = request.POST.get('sign_supervisor', None)

            # Guardar firma del supervisor
            if signature_image:
                self.object.sign_supervision = signature_image
            elif signature_data:
                format, imgstr = signature_data.split(';base64,')
                ext = format.split('/')[-1]
                filename = f'supervisor_signature_{uuid.uuid4()}.{ext}'
                data = ContentFile(base64.b64decode(imgstr), name=filename)
                self.object.sign_supervision.save(filename, data, save=True)
            
            self.object.save()

            # Actualizar rutinas de mantenimiento relacionadas y registrar ejecuciones en el PM
            rutas = Ruta.objects.filter(ot=ot)
            for ruta in rutas:
                actualizar_rutas(ruta)
                plan = ruta.maintenance_plans.filter(period_start__lte=TODAY, period_end__gte=TODAY).first()
                if plan:
                    success = record_execution(plan, TODAY)
                    if not success:
                        self.stdout.write(f"No se pudo registrar la ejecución para la ruta {ruta.name} en la fecha {TODAY}")

            # Cerrar reportes de averias relacionados
            fallas = FailureReport.objects.filter(related_ot=ot)
            for fail in fallas:
                fail.closed = True
                fail.modified_by = request.user
                fail.save()

            return redirect(ot.get_absolute_url())

        context = {'ot': ot, 'task_form': task_form, 'state_form': state_form}
        return render(request, self.template_name, context)


class OtCreate(CreateView):
    model = Ot
    http_method_names = ['get', 'post']
    template_name = 'got/ots/ot_form.html'

    def get_form_class(self):
        if self.request.user.groups.filter(name='mto_members').exists():
            return OtForm
        else:
            return OtFormNoSup

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        asset_id = self.kwargs.get('pk')
        asset = Asset.objects.get(pk=asset_id)
        kwargs['asset'] = asset
        return kwargs

    def form_valid(self, form):
        ot = form.save(commit=False)
        ot.modified_by = self.request.user
        if isinstance(form, OtFormNoSup):
            ot.supervisor = self.request.user.get_full_name()
        ot.save()
        return redirect('got:ot-detail', pk=ot.pk)


class OtUpdate(UpdateView):
    model = Ot
    http_method_names = ['get', 'post']
    template_name = 'got/ots/ot_form.html'

    def get_form_class(self):
        if self.request.user.groups.filter(name='mto_members').exists():
            return OtForm
        else:
            return OtFormNoSup

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        ot_instance = self.get_object()
        kwargs['asset'] = ot_instance.system.asset
        return kwargs

    def form_valid(self, form):
        ot = form.save(commit=False)
        ot.modified_by = self.request.user
        ot.save()
        return super().form_valid(form)


class OtDelete(DeleteView):
    model = Ot
    success_url = reverse_lazy('got:ot-list')
    template_name = 'got/ots/ot_confirm_delete.html'


'TASKS VIEW'
class AssignedTaskByUserListView(LoginRequiredMixin, generic.ListView):
    model = Task
    template_name = 'got/ots/tasks_pendient.html'
    paginate_by = 25

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['assets'] = Asset.objects.all()
        context['all_users'] = operational_users(self.request.user)
        context['total_tasks'] = self.get_queryset().count()

        # Incluimos los parámetros de fecha para poder construir el enlace al PDF
        context['start_date'] = self.request.GET.get('start_date', '')
        context['end_date'] = self.request.GET.get('end_date', '')
        context['finalizados'] = self.request.GET.get('finalizados', '0')
        return context

    def get_queryset(self):
        return filter_tasks_queryset(self.request)
    

class TaskPDFView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        queryset = filter_tasks_queryset(request).order_by('ot__system__asset__name', 'start_date')
    
        # Obtener los parámetros de fecha
        start = request.GET.get('start_date', '')
        end = request.GET.get('end_date', '')
        # Formar el string con el rango de fechas (separado por coma)
        date_range = f"{start},{end}" if start != end else start

        context = {
            'tasks': queryset,
            'start': start,
            'end': end,
            'finalizados': request.GET.get('show_finalizadas', '0'),
            'date_range': date_range,
        }
        return pdf_render(request, 'got/ots/assigned_tasks_pdf.html', context, "REPORTE DE ACTIVIDADES")


@login_required
def assignedTasks_excel(request):
    # Usar el mismo queryset filtrado
    queryset = filter_tasks_queryset(request)
    start = request.GET.get('start_date', '')
    # Ordenar por asset, OT y start_date (la misma lógica que en la vista PDF)
    queryset = queryset.order_by('ot__system__asset__name', 'ot__num_ot', 'start_date')
    
    # Agrupar las tareas: grouped[asset][ot] = lista de tareas
    grouped = defaultdict(lambda: defaultdict(list))
    for task in queryset:
        asset = task.ot.system.asset
        ot = task.ot
        grouped[asset][ot].append(task)
    
    # Crear el workbook y la hoja de cálculo
    wb = Workbook()
    ws = wb.active
    ws.title = "Actividades"
    
    # Definir un borde delgado para las celdas
    thin_border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )
    
    # === Insertar el logo de la empresa ===
    url = "https://hivik.s3.us-east-2.amazonaws.com/static/Logo.png"
    img_response = requests.get(url)
    if img_response.status_code == 200:
        image_data = BytesIO(img_response.content)
        from openpyxl.drawing.image import Image as XLImage
        img = XLImage(image_data)
        img.width = 300   # Ajusta según tus necesidades
        img.height = 80   # Ajusta según tus necesidades
        ws.add_image(img, "A1")
    else:
        print("No se pudo descargar la imagen.")
    
    # --- Fusionar las primeras 4 filas para el título ---
    ws.merge_cells('A1:D4')
    title_cell = ws['A1']
    title_cell.value = "LISTADO DE ACTIVIDADES PROGRAMADAS" + (f" ({start})" if start else "")
    title_cell.font = Font(bold=True, size=20, color="FFFFFF")
    title_cell.fill = PatternFill(fill_type="solid", fgColor="4d93d9")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in range(1, 5):
        ws.row_dimensions[row].height = 30
    
    current_row = 5  # Comenzamos el contenido a partir de la fila 5
    
    # Recorrer cada asset (ordenado por nombre)
    for asset in sorted(grouped.keys(), key=lambda a: a.name):
        # Encabezado del asset: fusionar columnas 1 a 5
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        cell_asset = ws.cell(row=current_row, column=1, value=f"{asset.name} ({asset.abbreviation})")
        ws.row_dimensions[current_row].height = 20
        cell_asset.font = Font(bold=True, size=14)
        cell_asset.alignment = Alignment(horizontal="left")
        # Aplicar borde a todas las celdas del rango fusionado
        for row_cells in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=1, max_col=4):
            for cell in row_cells:
                cell.border = thin_border
        current_row += 1
        
        # Recorrer cada OT dentro del asset (ordenado por num_ot)
        for ot in sorted(grouped[asset].keys(), key=lambda ot: ot.num_ot):
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            cell_ot = ws.cell(row=current_row, column=1, value=f"OT-{ot.num_ot}: {ot.description}")
            cell_ot.font = Font(bold=True, size=12)
            cell_ot.alignment = Alignment(horizontal="left")
            for row_cells in ws.iter_rows(min_row=current_row, max_row=current_row, min_col=1, max_col=4):
                for cell in row_cells:
                    cell.border = thin_border
            current_row += 1
            
            # Encabezados para las actividades
            headers = ['Actividad', 'Responsable', 'Inicio', 'Fin']
            for col_num, header in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(fill_type="solid", fgColor="4d93d9")
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
            current_row += 1
            
            # Escribir cada actividad de esta OT
            for task in grouped[asset][ot]:
                ws.cell(row=current_row, column=1, value=task.description).border = thin_border
                responsable = f"{task.responsible.first_name} {task.responsible.last_name}" if task.responsible else ""
                ws.cell(row=current_row, column=2, value=responsable).border = thin_border
                ws.cell(row=current_row, column=3, value=task.start_date.strftime('%d/%m/%Y') if task.start_date else "").border = thin_border
                ws.cell(row=current_row, column=4, value=task.final_date.strftime('%d/%m/%Y') if task.final_date else "").border = thin_border
                current_row += 1

                # ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                # ws.cell(row=current_row, column=1, value=task.news).border = thin_border
                # cell.font = Font(name='Arial', italic=True)
                # cell.fill = PatternFill(fill_type="solid", fgColor="eeece1")
                # cell.border = thin_border
            # Espacio en blanco entre OT
            # current_row += 1
        # Espacio en blanco entre assets
        # current_row += 1

    # Ajustar el ancho de las columnas (evitando errores con celdas fusionadas)
    for i in range(1, ws.max_column + 1):
        col_letter = get_column_letter(i)
        max_length = 0
        for cell in ws[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2


    ws.column_dimensions['A'].width = 30

    for cell in ws['A']:
        cell.alignment = Alignment(wrap_text=True)        

    # Exportar el workbook a BytesIO y retornar la respuesta
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = "actividades.xlsx"
    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename={filename}'
    return response


class TaskCreate(CreateView):
    model = Task
    http_method_names = ['get', 'post']
    form_class = RutActForm
    template_name = 'got/rutinas/task_form_rut.html'

    def form_valid(self, form):
        pk = self.kwargs['pk']
        ruta = get_object_or_404(Ruta, pk=pk)
        form.instance.ruta = ruta
        form.instance.finished = False
        form.instance.modified_by = self.request.user

        return super().form_valid(form)

    def get_success_url(self):
        task = self.object
        return reverse('got:sys-detail', args=[task.ruta.system.id])
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        ruta = get_object_or_404(Ruta, pk=self.kwargs['pk'])
        asset = ruta.system.asset  # Obtenemos el asset desde la ruta
        kwargs['asset'] = asset
        return kwargs


class TaskUpdate(UpdateView):
    model = Task
    http_method_names = ['get', 'post']
    second_form_class = UploadImages

    def dispatch(self, request, *args, **kwargs):
        # Guarda la URL previa si está presente para redirigir luego
        self.next_url = request.GET.get('next') or request.POST.get('next') or ''
        return super().dispatch(request, *args, **kwargs)

    def get_template_names(self):
        """
        Devuelve la plantilla a utilizar según la asociación de la tarea:
         - Para tareas con OT se utiliza 'got/ots/task_form_ot.html'
         - Para tareas con Ruta se utiliza 'got/ots/task_form_rut.html'
        """
        task = self.get_object()
        if task.ot:
            return ['got/ots/task_form_ot.html']
        elif task.ruta:
            return ['got/rutinas/task_form_rut.html']
        else:
            raise ValueError("La tarea debe estar asociada a una OT o a una Ruta.")

    def get_form_class(self):
        task = self.get_object()
        if task.ot:
            return ActForm
        elif task.ruta:
            return RutActForm
        else:
            raise ValueError("La tarea debe estar asociada a una OT o a una Ruta.")
        
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        task = self.get_object()
        if task.ruta:
            kwargs['asset'] = task.ruta.system.asset
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        task = self.get_object()
        if task.ot:
            if 'image_form' not in context:
                context['image_form'] = self.second_form_class()
            context['images'] = Image.objects.filter(task=self.get_object())
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form()

        if self.object.ot:
            image_form = self.second_form_class(request.POST, request.FILES)
            if form.is_valid() and image_form.is_valid():
                response = self.form_valid(form)

                for img in request.FILES.getlist('file_field'):
                    Image.objects.create(task=self.object, image=img)
                return response
            else:
                return self.form_invalid(form)
        else:
            if form.is_valid():
                return self.form_valid(form)
            else:
                return self.form_invalid(form)
            
    def form_valid(self, form):
        form.instance.modified_by = self.request.user
        if form.instance.ot and 'responsible' in form.changed_data:
            responsible = form.cleaned_data.get('responsible')
            form.instance.user = responsible.get_full_name()
        return super().form_valid(form)

    def form_invalid(self, form):
        print("Errores en el formulario:", form.errors)
        return super().form_invalid(form)
    
    def get_success_url(self):
        referer = self.request.META.get('HTTP_REFERER')
        if self.next_url:
            return self.next_url
        else:    
            return referer


class TaskDelete(DeleteView):
    model = Task
    success_url = reverse_lazy('got:ot-list')
    template_name = 'got/ots/task_confirm_delete.html'


class TaskDeleterut(DeleteView):
    model = Task
    def get_success_url(self):
        sys_id = self.object.ruta.system.id
        return reverse_lazy('got:sys-detail', kwargs={'pk': sys_id})

    def get(self, request, *args, **kwargs):
        context = {'task': self.get_object()}
        return render(request, 'got/ots/task_confirm_delete.html', context)


class Finish_task(UpdateView):
    model = Task
    form_class = FinishTask
    template_name = 'got/ots/task_finish_form.html'
    second_form_class = UploadImages
    success_url = reverse_lazy('got:my-tasks')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if 'image_form' not in context:
            context['image_form'] = self.second_form_class()
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form()
        image_form = self.second_form_class(request.POST, request.FILES)
        if form.is_valid() and image_form.is_valid():
            response = super().form_valid(form)
            for img in request.FILES.getlist('file_field'):
                Image.objects.create(task=self.object, image=img)
            return response
        else:
            return self.form_invalid(form)

    def form_valid(self, form):
        form.instance.modified_by = self.request.user  # Asignar el usuario actual
        return super().form_valid(form)

    def form_invalid(self, form, **kwargs):
        return self.render_to_response(self.get_context_data(form=form, **kwargs))
    

class Finish_task_ot(UpdateView):
    model = Task
    form_class = FinishTask
    template_name = 'got/ots/task_finish_form.html'
    second_form_class = UploadImages

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if 'image_form' not in context:
            context['image_form'] = self.second_form_class()
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form()
        image_form = self.second_form_class(request.POST, request.FILES)
        if form.is_valid() and image_form.is_valid():
            return self.form_valid(form, image_form)
        else:
            return self.form_invalid(form)

    def form_valid(self, form, image_form):
        form.instance.modified_by = self.request.user 
        self.object = form.save()

        for img in self.request.FILES.getlist('file_field'):
            Image.objects.create(task=self.object, image=img)

        ot = self.object.ot
        success_url = reverse('got:ot-detail', kwargs={'pk': ot.pk})
        return HttpResponseRedirect(success_url)

    def form_invalid(self, form, **kwargs):
        return self.render_to_response(self.get_context_data(form=form, **kwargs))


class Reschedule_task(UpdateView):
    model = Task
    form_class = RescheduleTaskForm
    template_name = 'got/ots/task_reschedule.html'
    success_url = reverse_lazy('got:my-tasks')

    def form_valid(self, form):
        form.instance.modified_by = self.request.user
        return super().form_valid(form)
    

'RUTINAS VIEWS'
@login_required
def RutaListView(request):

    suministro_prefetch = Prefetch('suministros', queryset=Suministro.objects.all(), to_attr='all_suministros')
    equipo_prefetch = Prefetch('equipos', queryset=Equipo.objects.prefetch_related(suministro_prefetch).annotate(num_suministros=Count('suministros')).filter(num_suministros__gt=0), to_attr='all_equipos')
    system_prefetch = Prefetch('system_set', queryset=System.objects.prefetch_related(equipo_prefetch).annotate(num_equipos_with_suministros=Count('equipos__suministros')).filter(num_equipos_with_suministros__gt=0), to_attr='all_systems')
    assets = Asset.objects.filter(area='a').prefetch_related(system_prefetch).annotate(num_systems_with_equipos=Count('system__equipos__suministros')).filter(num_systems_with_equipos__gt=0)

    diques = Ruta.objects.filter(name__icontains='DIQUE')
    barcos = Asset.objects.filter(area='a')


    motores_data = []
    for barco in barcos:
        sistema = barco.system_set.filter(group=200).first()
        sistema2 = barco.system_set.filter(group=300).first()
        motores_info = {
            'name': barco.name,
            'estribor': {
                'marca': '---',
                'modelo': '---',
                'lubricante': '---',
                'capacidad': 0,
                'fecha': '---',
                },
            'babor': {
                'marca': '---',
                'modelo': '---'},
            'generador': {'marca': '---', 'modelo': '---', 'lubricante': '---', 'capacidad': 0}
        }

        if sistema:
            motor_estribor = sistema.equipos.filter(name__icontains='Motor propulsor estribor').first()
            motor_babor = sistema.equipos.filter(name__icontains='Motor propulsor babor').first()
            motor_generador1 = sistema2.equipos.filter(Q(name__icontains='Motor generador estribor') | Q(name__icontains='Motor generador 1')).first() if sistema2 else None
            motor_generador2 = sistema2.equipos.filter(Q(name__icontains='Motor generador babor') | Q(name__icontains='Motor generador 2')).first() if sistema2 else None
            
            if motor_estribor:
                motores_info['estribor'] = {
                    'marca': motor_estribor.marca,
                    'modelo': motor_estribor.model,
                    'lubricante': motor_estribor.lubricante,
                    'capacidad': motor_estribor.volumen,
                    'horometro': motor_estribor.horometro,
                    'fecha': motor_estribor.last_hour_report_date(),
                    }
            if motor_babor:
                motores_info['babor'] = {
                    'marca': motor_babor.marca,
                    'modelo': motor_babor.model,
                    'lubricante': motor_babor.lubricante,
                    'capacidad': motor_babor.volumen,
                    'horometro': motor_babor.horometro,
                    'fecha': motor_babor.last_hour_report_date()
                    }
            if motor_generador1:
                motores_info['generador1'] = {
                    'marca': motor_generador1.marca,
                    'modelo': motor_generador1.model,
                    'lubricante': motor_generador1.lubricante,
                    'capacidad': motor_generador1.volumen,
                    'horometro': motor_generador1.horometro,
                    'fecha': motor_generador1.last_hour_report_date()
                    }
            if motor_generador2:
                motores_info['generador2'] = {
                    'marca': motor_generador2.marca,
                    'modelo': motor_generador2.model,
                    'lubricante': motor_generador2.lubricante,
                    'capacidad': motor_generador2.volumen,
                    'horometro': motor_generador2.horometro,
                    'fecha': motor_generador2.last_hour_report_date()
                    }

        motores_data.append(motores_info)

    context = {
        'assets': assets,
        'dique_rutinas': diques,
        'barcos': barcos,
        'motores_data': motores_data,
    }
    return render(request, 'got/rutinas/ruta_list.html', context)


@login_required
def crear_ot_desde_ruta(request, ruta_id):
    ruta = get_object_or_404(Ruta, pk=ruta_id)

    if request.method == 'POST':
        asociar_otros = request.POST.get('asociar_otros', 'off') == 'on'
        rutinas_seleccionadas_ids = [str(ruta_id)]
        if asociar_otros:
            # Obtener rutinas seleccionadas
            rutinas_seleccionadas_ids += request.POST.getlist('rutinas_seleccionadas')

        rutinas_seleccionadas = Ruta.objects.filter(pk__in=rutinas_seleccionadas_ids)
        num_rutinas = rutinas_seleccionadas.count()

        primera_ruta = rutinas_seleccionadas.first()
        nueva_ot = Ot(
            description=f"Rutina de mantenimiento con código {primera_ruta.name}",
            state='x',
            supervisor=f"{request.user.first_name} {request.user.last_name}",
            tipo_mtto='p',
            system=primera_ruta.system,
            modified_by=request.user 
        )
        nueva_ot.save()

        def copiar_tasks_y_actualizar_ot(ruta, ot, modify_description=False):
            for task in ruta.task_set.all():
                if modify_description:
                    equipo_sistema = ruta.equipo.name if ruta.equipo else ruta.system.name
                    description = f"{task.description} ({equipo_sistema})"
                else:
                    description = task.description
                Task.objects.create(
                    ot=ot,
                    responsible=task.responsible,
                    description=description,
                    procedimiento=task.procedimiento,
                    hse=task.hse,
                    start_date=timezone.now().date(),
                    men_time=1,
                    finished=False,
                    modified_by=request.user 
                )

            ruta.ot = ot
            ruta.save()

            if ruta.dependencia:
                copiar_tasks_y_actualizar_ot(ruta.dependencia, ot)

        modify_description = num_rutinas > 1
        for selected_ruta in rutinas_seleccionadas:
            copiar_tasks_y_actualizar_ot(selected_ruta, nueva_ot, modify_description)

        # copiar_tasks_y_actualizar_ot(ruta, nueva_ot)
        return redirect('got:ot-detail', pk=nueva_ot.pk)
    else:
        # Manejar solicitudes GET para crear una OT para una sola rutina
        nueva_ot = Ot(
            description=f"Rutina de mantenimiento con código {ruta.name}",
            state='x',
            supervisor=f"{request.user.first_name} {request.user.last_name}",
            tipo_mtto='p',
            system=ruta.system,
            modified_by=request.user 
        )
        nueva_ot.save()

        def copiar_tasks_y_actualizar_ot(ruta, ot):
            for task in ruta.task_set.all():
                Task.objects.create(
                    ot=ot,
                    responsible=task.responsible,
                    description=task.description,
                    procedimiento=task.procedimiento,
                    hse=task.hse,
                    start_date=timezone.now().date(),
                    men_time=1,
                    finished=False,
                    modified_by=request.user 
                )

            ruta.ot = ot
            ruta.save()

            if ruta.dependencia:
                copiar_tasks_y_actualizar_ot(ruta.dependencia, ot)

        copiar_tasks_y_actualizar_ot(ruta, nueva_ot)
        return redirect('got:ot-detail', pk=nueva_ot.pk)


def rutina_form_view(request, ruta_id):
    ruta = get_object_or_404(Ruta, code=ruta_id)
    tasks_ruta_principal = Task.objects.filter(ruta=ruta)
    fecha_actual = timezone.now().date()
    fecha_seleccionada = request.POST.get('fecha', fecha_actual)

    if request.method == 'POST':
        formset_data = []
        procesar_tasks_y_dependencias(request, ruta, formset_data)

        # Crear OT finalizada
        # Guardar firma del supervisor
        signature_image = request.FILES.get('signature_image')
        signature_data = request.POST.get('signature')
        if signature_image:
            signature_file = signature_image
        if signature_data:
            format, imgstr = signature_data.split(';base64,')
            ext = format.split('/')[-1]
            filename = f'signature_{uuid.uuid4()}.{ext}'
            signature_file = ContentFile(base64.b64decode(imgstr), name=filename)

        # Datos de nueva OT a crear y cerrar
        new_ot = Ot.objects.create(
            system = ruta.system,
            description = f"Rutina de mantenimiento: {ruta.name}",
            supervisor = request.user.get_full_name(),
            state = 'f',
            tipo_mtto = 'p',
            sign_supervision = signature_file,
            modified_by = request.user
        )
            
        # Asignar tareas a la nueva OT
        for form_data in formset_data:
            new_task = Task.objects.create(
                ot = new_ot,
                description = form_data['task'].description,
                responsible = request.user,
                user = form_data['user'],
                news = form_data['observaciones'],
                finished = form_data['realizado'],
                start_date = fecha_seleccionada,
                modified_by = request.user
            )
            # Guardar evidencias si las hay
            for evidencia in form_data['evidencias']:
                Image.objects.create(task=new_task, image=evidencia)

        actualizar_rutas(ruta, fecha_seleccionada, new_ot)
        return redirect(reverse('got:sys-detail', args=[ruta.system.id]))

    else:
        formset_data = [{'task': task, 'form': ActivityForm(), 'upload_form': UploadImages()} for task in tasks_ruta_principal]
        dependencias = []
        # Agregar las tareas de las rutas dependientes
        if ruta.dependencia:
            dependencias = [ruta.dependencia]
            while dependencias[-1].dependencia:
                dependencias.append(dependencias[-1].dependencia)

            for dependencia in dependencias:
                tasks_dependencia = Task.objects.filter(ruta=dependencia)
                formset_data += [{'task': task, 'form': ActivityForm(), 'upload_form': UploadImages(), 'dependencia': dependencia.name} for task in tasks_dependencia]

    context = {
        'formset_data': formset_data,
        'ruta': ruta, 
        'dependencias': dependencias, 
        'fecha_seleccionada': fecha_seleccionada
    }
    return render(request, 'got/rutinas/ruta_ot_form.html', context)


'SOLICITUDES VIEWS'
class SolicitudesListView(LoginRequiredMixin, generic.ListView):
    model = Solicitud
    paginate_by = 20
    template_name = 'got/solicitud/solicitud_list.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['assets'] = Asset.objects.all()

        asset_filter = self.request.GET.get('asset')
        context['asset'] = Asset.objects.filter(abbreviation=asset_filter).first() if asset_filter else None

        user = self.request.user
        user_groups = set(user.groups.values_list('name', flat=True))
        context['user_groups'] = user_groups

        group_default_dpto = {
            'mto_members': 'm',  # Mantenimiento
            'operaciones': 'o'     # Operaciones
        }

        default_dpto = ''
        for group, dpto in group_default_dpto.items():
            if group in user_groups:
                default_dpto = dpto
                break 

        dpto = self.request.GET.get('dpto', default_dpto)
        current_dpto = dpto if dpto else ''

        context['default_dpto'] = default_dpto
        context['current_dpto'] = current_dpto
        return context

    def get_queryset(self):
        queryset = Solicitud.objects.all()
        user = self.request.user
        user_groups = set(user.groups.values_list('name', flat=True))

        group_default_dpto = {
            'mto_members': 'm',  # Mantenimiento
            'operaciones': 'o'     # Operaciones
        }

        default_dpto = ''
        for group, dpto in group_default_dpto.items():
            if group in user_groups:
                default_dpto = dpto
                break 
        
        dpto = self.request.GET.get('dpto', default_dpto)
        current_dpto = dpto if dpto else ''

        state = self.request.GET.get('state')
        asset_filter = self.request.GET.get('asset')
        keyword = self.request.GET.get('keyword')

        if asset_filter:
            queryset = queryset.filter(asset__abbreviation=asset_filter)
        if keyword:
            queryset = queryset.filter(suministros__icontains=keyword)

        if state == 'no_aprobada':
            queryset = queryset.filter(approved=False, cancel=False)
        elif state == 'aprobada':
            queryset = queryset.filter(approved=True, sc_change_date__isnull=True, cancel=False)
        elif state == 'tramitado':
            queryset = queryset.filter(approved=True, sc_change_date__isnull=False, cancel=False)
        elif state == 'parcialmente':
            queryset = queryset.filter(satisfaccion=False, recibido_por__isnull=False, cancel=False)
        elif state == 'recibido':
            queryset = queryset.filter(satisfaccion=True, cancel=False)
        elif state == 'cancel':
            queryset = queryset.filter(cancel=True)

        if current_dpto:
            queryset = queryset.filter(dpto=current_dpto)

        if user.has_perm('got.can_view_all_rqs'):
            return queryset
        else:
            filter_condition = Q()

            if 'maq_members' in user_groups:
                supervised_assets = Asset.objects.filter(Q(supervisor=user) | Q(capitan=user))
                filter_condition |= Q(asset__in=supervised_assets)
            if 'serport_members' in user_groups:
                filter_condition |= Q(solicitante=user)
            if 'buzos' in user_groups:
                buceo_assets = Asset.objects.filter(area='b')
                filter_condition |= Q(asset__in=buceo_assets)

            if not filter_condition:
                filter_condition = Q(solicitante=user)
            queryset = queryset.filter(filter_condition)

        return queryset


def detalle_pdf(request, pk):
    registro = Solicitud.objects.get(pk=pk)
    context = {'rq': registro}
    return render_to_pdf('got/solicitud/solicitud_detail.html', context)


class CreateSolicitudOt(LoginRequiredMixin, View):
    template_name = 'got/solicitud/create-solicitud-ot.html'

    def get(self, request, asset_id, ot_num=None):
        asset = get_object_or_404(Asset, abbreviation=asset_id)
        ot = None if not ot_num else get_object_or_404(Ot, num_ot=ot_num)
        items = Item.objects.all()

        return render(request, self.template_name, {
            'ot': ot,
            'asset': asset,
            'items': items
        })

    def post(self, request, asset_id, ot_num=None):
            asset = get_object_or_404(Asset, abbreviation=asset_id)
            ot = None if not ot_num else get_object_or_404(Ot, num_ot=ot_num)
            items_ids = request.POST.getlist('item_id[]') 
            cantidades = request.POST.getlist('cantidad[]')
            suministros = request.POST.get('suministros', '')
            dpto = request.POST.get('dpto')

            solicitud = Solicitud.objects.create(
                solicitante=request.user,
                requested_by=request.user.get_full_name(),
                ot=ot,
                asset=asset,
                suministros=suministros,
                dpto=dpto,
            )

            for item_id, cantidad in zip(items_ids, cantidades):
                if item_id and cantidad:
                    item = get_object_or_404(Item, id=item_id)
                    Suministro.objects.create(
                        item=item,
                        cantidad=int(cantidad),
                        Solicitud=solicitud
                    )
            return redirect('got:my-tasks')


class EditSolicitudView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        solicitud = get_object_or_404(Solicitud, pk=kwargs['pk'])
        suministros = request.POST.get('suministros')
        if suministros:
            solicitud.suministros = suministros
            solicitud.save()
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))
    

class ApproveSolicitudView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        solicitud = get_object_or_404(Solicitud, id=kwargs['pk'])
        solicitud.approved = not solicitud.approved
        if solicitud.approved:
            solicitud.approved_by = f"{request.user.first_name} {request.user.last_name}"
            solicitud.approval_date = timezone.now()
        else:
            solicitud.approved_by = ''
            solicitud.approval_date = None
        solicitud.save()
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))
    

class DeleteSolicitudView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        # Verificar si el usuario es superusuario (administrador)
        return self.request.user.is_superuser

    def handle_no_permission(self):
        return HttpResponseForbidden("No tiene permiso para eliminar esta solicitud.")

    def post(self, request, *args, **kwargs):
        solicitud = get_object_or_404(Solicitud, pk=kwargs['pk'])
        solicitud.delete()
        messages.success(request, "Solicitud eliminada exitosamente.")
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))

@login_required
def report_received(request, pk):
    if request.method == 'POST':
        solicitud = get_object_or_404(Solicitud, pk=pk)
        satisfaccion = request.POST.get('satisfaccion') == 'True'
        recibido_por = request.POST.get('recibido_por')
        solicitud.satisfaccion = satisfaccion
        solicitud.recibido_por = recibido_por
        solicitud.save()
        return redirect('got:rq-list')
    return redirect('got:rq-list')


@login_required
def update_sc(request, pk):
    if request.method == 'POST':
        solicitud = get_object_or_404(Solicitud, pk=pk)
        num_sc = request.POST.get('num_sc')
        solicitud.num_sc = num_sc
        solicitud.save()

        query_params = {
            'asset': request.POST.get('asset', ''),
            'state': request.POST.get('state', ''),
            'keyword': request.POST.get('keyword', ''),
        }
        
        query_string = '&'.join([f'{key}={value}' for key, value in query_params.items() if value])

        redirect_url = f'{reverse("got:rq-list")}'
        if query_string:
            redirect_url += f'?{query_string}'
        return redirect(redirect_url)
    return redirect('got:rq-list') 


@login_required
def cancel_sc(request, pk):
    if request.method == 'POST':
        solicitud = get_object_or_404(Solicitud, pk=pk)
        reason = request.POST.get('cancel_reason')
        solicitud.cancel_reason = reason
        solicitud.cancel = True
        solicitud.save()
        return redirect('got:rq-list')
    return redirect('got:rq-list')


def download_pdf(request):
    state = request.GET.get('state', '')
    asset_filter = request.GET.get('asset', '')
    keyword = request.GET.get('keyword', '')

    queryset = Solicitud.objects.all()

    if asset_filter:
        queryset = queryset.filter(asset__abbreviation=asset_filter)
    if state:
        if state == 'no_aprobada':
            queryset = queryset.filter(approved=False)
        elif state == 'aprobada':
            queryset = queryset.filter(approved=True, sc_change_date__isnull=True)
        elif state == 'tramitado':
            queryset = queryset.filter(approved=True, sc_change_date__isnull=False)
    if keyword:
        queryset = queryset.filter(suministros__icontains=keyword)

    context = {
        'rqs': queryset,
        'state': state, 
        'asset': Asset.objects.filter(pk=asset_filter).first(),
        'keyword': keyword,
        }
    return render_to_pdf('got/solicitud/rqs_report.html', context)


class TransferSolicitudView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'got.can_transfer_solicitud'

    def post(self, request, *args, **kwargs):
        solicitud = get_object_or_404(Solicitud, pk=kwargs['pk'])
        # Cambiar el departamento al otro valor
        if solicitud.dpto == 'o':
            solicitud.dpto = 'm'
        elif solicitud.dpto == 'm':
            solicitud.dpto = 'o'
        solicitud.save()
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))


'GENERAL VIEWS'
@login_required
def indicadores(request):
    m = 12
    y = 2024

    area_filter = request.GET.get('area', None)

    assets = Asset.objects.annotate(num_ots=Count('system__ot'))
    if area_filter:
        assets = assets.filter(area=area_filter)
    top_assets = assets.order_by('-num_ots')[:5]
    ots_per_asset = [a.num_ots for a in top_assets]
    asset_labels = [a.name for a in top_assets]

    labels = ['Preventivo', 'Correctivo', 'Modificativo']

    earliest_start_date = Task.objects.filter(ot=OuterRef('pk')).order_by('start_date').values('start_date')[:1]
    latest_final_date = Task.objects.filter(ot=OuterRef('pk')).annotate(
        final_date=ExpressionWrapper(
            F('start_date') + F('men_time'),
            output_field=DateField()
        )
    ).order_by('-final_date').values('final_date')[:1]

    bar = Ot.objects.annotate(
        start=Subquery(earliest_start_date),
        end=Subquery(latest_final_date)
    ).filter(state='x')

    if area_filter:
        bar = bar.filter(system__asset__area=area_filter)

    barcos = bar.filter(system__asset__area='a')

    if area_filter:
        ots = len(Ot.objects.filter(creation_date__month=m, creation_date__year=y, system__asset__area=area_filter))

        ot_finish = len(Ot.objects.filter(
            creation_date__month=m,
            creation_date__year=y, state='f',
            system__asset__area=area_filter))

        preventivo = len(Ot.objects.filter(
            creation_date__month=m,
            creation_date__year=y,
            tipo_mtto='p',
            system__asset__area=area_filter
            ))
        correctivo = len(Ot.objects.filter(
            creation_date__month=m,
            creation_date__year=y,
            tipo_mtto='c',
            system__asset__area=area_filter
            ))
        modificativo = len(Ot.objects.filter(
            creation_date__month=m,
            creation_date__year=y,
            tipo_mtto='m',
            system__asset__area=area_filter
            ))

    else:
        ots = len(Ot.objects.filter(
            creation_date__month=m, creation_date__year=y))
        ot_finish = len(Ot.objects.filter(
            creation_date__month=m, creation_date__year=y, state='f'))

        preventivo = len(Ot.objects.filter(
            creation_date__month=m, creation_date__year=y, tipo_mtto='p'))
        correctivo = len(Ot.objects.filter(
            creation_date__month=m, creation_date__year=y, tipo_mtto='c'))
        modificativo = len(Ot.objects.filter(
            creation_date__month=m, creation_date__year=y, tipo_mtto='m'))

    if ots == 0:
        ind_cumplimiento = 0
        data = 0
    else:
        ind_cumplimiento = round((ot_finish/ots)*100, 2)
        data = [
            round((preventivo/ots)*100, 2), round((correctivo/ots)*100, 2),
            round((modificativo/ots)*100, 2)
            ]
        
    assets_a = Asset.objects.filter(area='a')

    # Lista para almacenar los datos de combustible
    combustible_data = []

    # Iterar sobre cada asset y obtener la información de combustible
    for asset in assets_a:
        try:
            # Obtener el suministro de combustible para el asset actual
            suministro = Suministro.objects.get(asset=asset, item_id=132)
            total_quantity = suministro.cantidad

            last_transaction = Transaction.objects.filter(suministro=suministro, tipo='c').order_by('-fecha').first()
            if last_transaction:
                last_report_date = last_transaction.fecha
            else:
                last_report_date = None

            combustible_data.append({
                'asset_name': asset.name,
                'last_report_date': last_report_date,
                'total_quantity': total_quantity,
            })
        except Suministro.DoesNotExist:
            # Si el suministro de combustible no existe para este asset
            combustible_data.append({
                'asset_name': asset.name,
                'last_report_date': None,
                'total_quantity': None,
            })  

    context = {
        'ind_cumplimiento': ind_cumplimiento,
        'data': data,
        'labels': labels,
        'ots': ots,
        'ots_asset': ots_per_asset,
        'asset_labels': asset_labels,
        'ots_finished': ot_finish,
        'barcos': barcos,
        'combustible_data': combustible_data,
    }
    return render(request, 'got/assets/indicadores.html', context)
    

class MaintenanceDashboardView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'got/mantenimiento/maintenance_dashboard.html'

    def test_func(self):
        return self.request.user.groups.filter(name='mto_members').exists()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ships = Asset.objects.filter(area='a', show=True)

        # Obtener sistemas de todos los barcos y agrupar por 'group' para obtener sistemas únicos
        systems_queryset = System.objects.filter(asset__in=ships).select_related('asset')
        unique_systems = {}
        for system in systems_queryset:
            key = system.name
            if key not in unique_systems:
                unique_systems[key] = system

        # Ordenar los sistemas únicos por 'group' o como prefieras
        systems = sorted(unique_systems.values(), key=lambda s: s.name)

        # Preparar la estructura de datos
        data = []
        for ship in ships:
            ship_row = {'ship': ship.name}
            ship_systems = System.objects.filter(asset=ship).prefetch_related(
                'rutas',
                'equipos__failurereport_set',
                'ot_set',
            )

            # Crear un diccionario de sistemas del barco actual por 'group'
            ship_systems_dict = {system.name: system for system in ship_systems}

            for system in systems:
                system_group = system.name
                ship_system = ship_systems_dict.get(system_group)
                if ship_system:
                    states, state_data = self.get_system_states(ship_system)
                    ship_row[system_group] = {
                        'states': states,
                        'state_data': state_data,
                        'system_id': ship_system.id,
                    }
                else:
                    ship_row[system_group] = {'states': [], 'state_data': {}, 'system_id': None}
            data.append(ship_row)

        context['systems'] = systems
        context['data'] = data
        return context

    def get_system_states(self, system):
        states = []
        state_data = {}
        today = date.today()
        
        rutas = list(system.rutas.all())
        equipos = list(system.equipos.all())
        failure_reports = []
        
        failure_reports = FailureReport.objects.filter(
            equipo__in=equipos,
            closed=False
        ).select_related('equipo', 'related_ot').prefetch_related(
            Prefetch(
                'related_ot__task_set',
                queryset=Task.objects.filter(finished=False),
                to_attr='tasks_in_execution'
            )
        )

        ots = system.ot_set.filter(state='x').prefetch_related(
            Prefetch(
                'task_set',
                queryset=Task.objects.filter(finished=False),
                to_attr='tasks_in_execution'
            )
        )

        ots_with_tasks_in_execution = []
        ots_without_tasks_in_execution = []

        for ot in ots:
            if ot.tasks_in_execution:
                ots_with_tasks_in_execution.append(ot)
            else:
                ots_without_tasks_in_execution.append(ot)
        
        requires_maintenance = False
        all_up_to_date = True
        has_planeacion = False
        overdue_rutinas = []
        planeacion_rutinas = []

        for ruta in rutas:
            next_date = ruta.next_date
            if next_date and next_date < today:
                requires_maintenance = True
                all_up_to_date = False
                overdue_rutinas.append(ruta)
            elif not next_date:
                requires_maintenance = True
                all_up_to_date = False
                overdue_rutinas.append(ruta)
            percentage = ruta.percentage_remaining
            if percentage and 0 < percentage < 15:
                has_planeacion = True
                planeacion_rutinas.append(ruta)

        if requires_maintenance:
            states.append(('Requiere', '#ff00ff'))  # Rojo
            state_data['Requiere'] = overdue_rutinas

        if has_planeacion:
            states.append(('Planeación', '#ffff00'))  # Amarillo
            state_data['Planeación'] = planeacion_rutinas

        # Estado "Alerta"
        alerta_reports = failure_reports.filter(critico=True)
        if alerta_reports.exists():
            states.append(('Alerta', '#cc0000'))  # Rojo intenso
            state_data['Alerta'] = alerta_reports

        # Estado "Novedades"
        novedades_reports = failure_reports.filter(critico=False)
        if novedades_reports.exists():
            states.append(('Novedades', '#ffa500'))  # Naranja
            state_data['Novedades'] = novedades_reports

        # Estado "Trabajando"
        if ots_with_tasks_in_execution:
            states.append(('Trabajando', '#800080'))  # Morado
            state_data['Trabajando'] = ots_with_tasks_in_execution

        if ots_without_tasks_in_execution:
            states.append(('Pendientes', '#025669'))  # Color personalizado
            state_data['Pendientes'] = ots_without_tasks_in_execution

        if not alerta_reports.exists() and not novedades_reports.exists():
            if not requires_maintenance and all_up_to_date:
                states.append(('Ok', '#86e49d'))  # Verde

        return states, state_data

    
class BudgetView(TemplateView):
    template_name = 'got/mantenimiento/budget_view.html'

    def get_assets_list(self):
        return Asset.objects.filter(area='a', show=True).order_by('name')

    def get_systems_list(self, asset_abbr):
        return System.objects.filter(asset__abbreviation=asset_abbr).order_by('name')

    def get_equipos_list(self, system_id):
        return Equipo.objects.filter(system_id=system_id).order_by('name')

    def get(self, request, *args, **kwargs):
        asset_abbr = request.GET.get('asset_abbr', '')
        system_id = request.GET.get('system_id', '')
        equipo_code = request.GET.get('equipo_code', '')

        period_start = date(2025, 1, 1)
        period_end = date(2025, 12, 31)

        # Obtener rutas según filtros
        rutas = Ruta.objects.all()
        if asset_abbr:
            rutas = rutas.filter(system__asset__abbreviation=asset_abbr)
        if system_id:
            rutas = rutas.filter(system_id=system_id)
        if equipo_code:
            rutas = rutas.filter(equipo__code=equipo_code)

        item_totals = {}
        service_totals = {}

        systems_with_reqs = set()

        # Cálculo de num_executions y totales
        for ruta in rutas:
            num_executions = calculate_executions(ruta, period_start, period_end)
            if num_executions == 0:
                continue

            requisitos = ruta.requisitos.all()
            if requisitos.exists():
                # Si esta ruta tiene requisitos, entonces este system_id tiene requerimientos
                systems_with_reqs.add(ruta.system_id)

            for req in requisitos:
                total_quantity = req.cantidad * num_executions
                if req.tipo in ['m', 'h']:
                    if req.item:
                        key = req.item.id
                        name = req.item
                        presentacion = req.item.presentacion
                        unit_price = req.item.unit_price if getattr(req.item, 'unit_price', None) else Decimal('0.00')
                        reference = req.item.reference if req.item.reference else ''
                    else:
                        key = ('desc', req.descripcion)
                        name = req.descripcion
                        presentacion = 'Sin presentación'
                        unit_price = Decimal('0.00')
                        reference = ''

                    if key not in item_totals:
                        item_totals[key] = {
                            'name': name,
                            'presentacion': presentacion,
                            'total_quantity': total_quantity,
                            'num_executions': num_executions,
                            'unit_price': unit_price,
                            'reference': reference,
                            'type': 'item',
                            'id': key if isinstance(key, int) else None
                        }
                    else:
                        item_totals[key]['total_quantity'] += total_quantity
                        item_totals[key]['num_executions'] += num_executions

                elif req.tipo == 's':
                    if req.service:
                        key = req.service.id
                        name = req.service.description
                        presentacion = 'Serv'
                        unit_price = req.service.unit_price if req.service.unit_price else Decimal('0.00')
                        reference = ''
                    else:
                        key = ('desc_s', req.descripcion)
                        name = req.descripcion
                        presentacion = 'Serv'
                        unit_price = Decimal('0.00')
                        reference = ''

                    if key not in service_totals:
                        service_totals[key] = {
                            'name': name,
                            'presentacion': presentacion,
                            'total_quantity': total_quantity,
                            'num_executions': num_executions,
                            'unit_price': unit_price,
                            'type': 'service',
                            'id': key if isinstance(key, int) else None
                        }
                    else:
                        service_totals[key]['total_quantity'] += total_quantity
                        service_totals[key]['num_executions'] += num_executions

        # Calcular costos
        for item in item_totals.values():
            item['total_cost'] = item['total_quantity'] * item['unit_price']

        for svc in service_totals.values():
            svc['total_cost'] = svc['total_quantity'] * svc['unit_price']

        # Ordenar
        item_list = list(item_totals.values())
        item_list.sort(key=lambda x: (
            str(x['name'].name).lower() if hasattr(x['name'], 'name') and x['name'].name else str(x['name']).lower(),
            x['reference'].lower() if x['reference'] else ''
        ))
        service_list = list(service_totals.values())
        service_list.sort(key=lambda x: str(x['name'] or '').lower())

        total_articulos = sum(i['total_cost'] for i in item_list)
        total_servicios = sum(s['total_cost'] for s in service_list)

        # Calcular micelanios (10%)
        micelanios_articulos = total_articulos * Decimal('0.10')
        micelanios_servicios = total_servicios * Decimal('0.10')

        # Nuevos totales incluyendo micelanios
        total_articulos_with_micelanios = total_articulos + micelanios_articulos
        total_servicios_with_micelanios = total_servicios + micelanios_servicios

        # Calcular total combinado
        combined_total = total_articulos_with_micelanios + total_servicios_with_micelanios

        # Obtener listas para filtros
        assets = self.get_assets_list()
        systems = self.get_systems_list(asset_abbr) if asset_abbr else []
        equipos = self.get_equipos_list(system_id) if system_id else []

        enabled_systems = set()
        if asset_abbr:
            for sys in systems:
                if sys.id in systems_with_reqs:
                    enabled_systems.add(sys.id)

        context = {
            'item_totals': item_list,
            'service_totals': service_list,
            'period_start': period_start,
            'period_end': period_end,
            'total_articulos': total_articulos,
            'micelanios_articulos': micelanios_articulos,
            'total_articulos_with_micelanios': total_articulos_with_micelanios,
            'total_servicios': total_servicios,
            'micelanios_servicios': micelanios_servicios,
            'total_servicios_with_micelanios': total_servicios_with_micelanios,
            'combined_total': combined_total,
            'assets_list': assets,
            'selected_asset': asset_abbr,
            'systems_list': systems,
            'selected_system': system_id,
            'equipos_list': equipos,
            'selected_equipo': equipo_code,
            'enabled_systems': enabled_systems
        }

        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        action = request.POST.get('action')
        if action == 'update_unit_price':
            obj_type = request.POST.get('obj_type')
            obj_id = request.POST.get('obj_id')
            new_price = request.POST.get('new_price')
            try:
                new_price = Decimal(new_price)
            except:
                new_price = Decimal('0.00')

            if obj_type == 'item':
                item = get_object_or_404(Item, pk=obj_id)
                item.unit_price = new_price
                item.save()
                messages.success(request, 'Precio del artículo actualizado.')
            elif obj_type == 'service':
                svc = get_object_or_404(Service, pk=obj_id)
                svc.unit_price = new_price
                svc.save()
                messages.success(request, 'Precio del servicio actualizado.')
            else:
                messages.error(request, 'Tipo no reconocido.')

            # Mantener filtros
            asset_abbr = request.GET.get('asset_abbr', '')
            system_id = request.GET.get('system_id', '')
            equipo_code = request.GET.get('equipo_code', '')

            redirect_url = reverse('got:budget_view')
            query_params = []
            if asset_abbr:
                query_params.append(f"asset_abbr={asset_abbr}")
            if system_id:
                query_params.append(f"system_id={system_id}")
            if equipo_code:
                query_params.append(f"equipo_code={equipo_code}")
            if query_params:
                redirect_url += "?" + "&".join(query_params)

            return redirect(redirect_url)
        else:
            messages.error(request, 'Acción no reconocida.')
            return redirect('got:budget_view')


class BudgetSummaryByAssetView(TemplateView):
    template_name = 'got/mantenimiento/budget_summary_view.html'

    def get(self, request, *args, **kwargs):
        period_start = date(2025, 1, 1)
        period_end   = date(2025, 12, 31)

        # 1. Definir las condiciones usando Q
        condition_barcos = Q(area='a', show=True)
        condition_otros   = Q(show=True) & ~Q(area='a') & Q(system__rutas__requisitos__isnull=False)

        # 2. Filtrar los assets que cumplen alguna de las condiciones
        all_assets = Asset.objects.filter(
            condition_barcos | condition_otros
        ).distinct().order_by('name')

        assets_data = []
        total_global = Decimal('0.00')
        micelanios_global = Decimal('0.00')

        for barco in all_assets:
            rutas = Ruta.objects.filter(system__asset=barco)
            costo_total_barco = Decimal('0.00')
            equipos_dict = {}
            categories_dict = {
                'Aceite y Filtros': Decimal('0.00'),
                'Servicios':        Decimal('0.00'),
                'Repuestos':        Decimal('0.00')
            }

            for ruta in rutas:
                num_exec = calculate_executions(ruta, period_start, period_end)
                if num_exec == 0:
                    continue

                equipo_id = ruta.equipo_id or 'otros'

                for req in ruta.requisitos.all():
                    unit_price = Decimal('0.00')
                    if req.tipo in ['m', 'h'] and req.item:
                        unit_price = req.item.unit_price or Decimal('0.00')
                    elif req.tipo == 's' and req.service:
                        unit_price = req.service.unit_price or Decimal('0.00')

                    subtotal = req.cantidad * unit_price * num_exec
                    equipos_dict.setdefault(equipo_id, Decimal('0.00'))
                    equipos_dict[equipo_id] += subtotal

                    # Clasificar en categories_dict
                    if req.tipo in ['m', 'h'] and req.item:
                        if req.item.name in ["Aceite", "Filtros"]:
                            categories_dict['Aceite y Filtros'] += subtotal
                        else:
                            categories_dict['Repuestos'] += subtotal
                    elif req.tipo == 's' and req.service:
                        categories_dict['Servicios'] += subtotal

            # 3. Calcular costos + micelanios
            for value in equipos_dict.values():
                costo_total_barco += value

            micelanios_barco = costo_total_barco * Decimal('0.10')
            costo_total_barco_with_micelanios = costo_total_barco + micelanios_barco

            total_global += costo_total_barco_with_micelanios
            micelanios_global += micelanios_barco

            # 4. Convertir dict equipos a lista
            equipos_list = []
            for eq_id, cost_eq in equipos_dict.items():
                if eq_id == 'otros':
                    eq_name = 'Otros'
                else:
                    try:
                        eq_obj = Equipo.objects.get(pk=eq_id)
                        eq_name = eq_obj.name
                    except Equipo.DoesNotExist:
                        eq_name = 'Desconocido'
                equipos_list.append({
                    'equipo_name': eq_name,
                    'cost': float(cost_eq)
                })
            equipos_list.sort(key=lambda e: e['equipo_name'].lower())

            # Datos para gráficas
            labels_equipos = [eq['equipo_name'] for eq in equipos_list]
            data_equipos = [eq['cost'] for eq in equipos_list]

            labels_categories = ['Aceite y Filtros', 'Servicios', 'Repuestos']
            data_categories = [
                float(categories_dict['Aceite y Filtros']),
                float(categories_dict['Servicios']),
                float(categories_dict['Repuestos'])
            ]

            assets_data.append({
                'asset_id':        barco.abbreviation,
                'asset_name':      barco.name,
                'asset_cost':      float(costo_total_barco_with_micelanios),
                'micelanios':      float(micelanios_barco),
                'equipos':         equipos_list,
                'labels_equipos':  labels_equipos,
                'data_equipos':    data_equipos,
                'types_labels':    labels_categories,
                'types_data':      data_categories,
            })

        # 5. Preparar datos para la gráfica principal
        pie_labels = []
        pie_data   = []
        for asset_info in assets_data:
            pie_labels.append(asset_info['asset_name'])
            pie_data.append(asset_info['asset_cost'])

        # Opcional: agregar "Micelanios" global
        pie_labels.append("Micelanios")
        pie_data.append(float(micelanios_global))

        # Convertir a JSON
        assets_data_json = json.dumps(assets_data)
        pie_labels_json  = json.dumps(pie_labels)
        pie_data_json    = json.dumps(pie_data)

        context = {
            'period_start':       period_start,
            'period_end':         period_end,
            'total_global':       float(total_global),
            'micelanios_global':  float(micelanios_global),
            'assets_data':        assets_data,
            'assets_data_json':   assets_data_json,
            'pie_labels':         pie_labels_json,
            'pie_data':           pie_data_json,
        }
        return self.render_to_response(context)


def managerial_asset_report_pdf(request, abbreviation):
    asset = get_object_or_404(Asset, abbreviation=abbreviation)
    systems = asset.system_set.all()

    systems_data = []
    for sys in systems:
        # Equipos y sus imágenes
        equipos_qs = sys.equipos.all()
        equipos_data = []
        for eq in equipos_qs:
            # Filtramos imágenes asociadas al equipo
            eq_images = eq.images.all().order_by('id')
            # Omitir los campos None => lo haremos en la plantilla

            equipos_data.append({
                'name': eq.name,
                'ubicacion': eq.ubicacion,
                'code': eq.code,
                'model': eq.model,
                'marca': eq.marca,
                'serial': eq.serial,
                'fabricante': eq.fabricante,
                'tipo_display': eq.get_tipo_display(),
                'feature': eq.feature,
                'all_images': eq_images,
            })

        # OTs en ejecución
        ots_en_x = sys.ot_set.filter(state='x')
        ots_en_ejecucion_data = []
        for ot in ots_en_x:
            open_tasks = ot.task_set.filter(finished=False)
            ots_en_ejecucion_data.append({
                'num_ot': ot.num_ot,
                'description': ot.description,
                'open_tasks': open_tasks,
            })

        # Fallas abiertas sin OT
        # => Recolectamos todos los equipos IDs
        eq_ids = equipos_qs.values_list('pk', flat=True)
        failures_abiertas = FailureReport.objects.filter(
            equipo_id__in=eq_ids,
            closed=False,
            related_ot__isnull=True
        )

        systems_data.append({
            'system': sys,
            'equipments': equipos_data,
            'ots_en_ejecucion': ots_en_ejecucion_data,
            'failures_abiertas': failures_abiertas,
        })

    context = {
        'asset': asset,
        'systems_data': systems_data,
        'today': timezone.now().date(),
    }

    return render_to_pdf('got/systems/managerial_asset_report.html', context)

class CustomPasswordResetView(PasswordResetView):
    email_template_name = 'registration/password_reset_email.txt'  # Plantilla de texto plano
    html_email_template_name = 'registration/password_reset_email.html'  # Plantilla HTML
    subject_template_name = 'registration/password_reset_subject.txt'  # Plantilla para el asunto
    domain_override = settings.MY_SITE_DOMAIN  # Establece el dominio

    def get_extra_email_context(self):
        # Determinar el protocolo
        protocol = 'https' if self.request.is_secure() else 'http'

        # Obtener el nombre del sitio
        site_name = settings.MY_SITE_NAME  # Define esto en settings.py

        return {
            'protocol': protocol,
            'domain': self.domain_override,
            'site_name': site_name,
        }
    
