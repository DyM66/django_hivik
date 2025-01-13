# inventory_management/views.py
import base64
import qrcode
import qrcode.image.svg  # si prefieres SVG
from io import BytesIO
from django.core.files.base import ContentFile
from django.utils import timezone
from django.shortcuts import render, get_object_or_404
from django.urls import reverse
from django.http import HttpResponseRedirect
from got.models import Asset, System, Equipo, Suministro, Item, DarBaja  # ← Importamos de la app "got"
from got.forms import DarBajaForm
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views import View
from django.views.generic import ListView, DetailView
from django.views.generic.edit import CreateView


class AssetListView(ListView):
    model = Asset
    template_name = 'inventory_management/asset_list.html'
    context_object_name = 'assets'

    def get_queryset(self):
        """
        Excluimos las áreas 'b' (Buceo), 'o' (Oceanografía) y 'x' (Apoyo).
        Filtramos show=True.
        """
        qs = super().get_queryset()
        qs = qs.filter(show=True).exclude(area__in=['b','o','x'])
        return qs.select_related('supervisor','capitan')


class ActivoEquipmentListView(View):
    template_name = 'inventory_management/asset_equipment_list.html'

    def get(self, request, abbreviation):
        """
        Muestra un scroll horizontal de todos los Activos (excluyendo b, o, x),
        y a la derecha o abajo la lista de Equipos para el Activo seleccionado,
        junto con la tabla de Suministros, etc.
        """
        # 1) Lista de Activos (excluyendo area b, o, x), show=True
        #    Ordenados por nombre. Utilizamos select_related para traer supervisor/capitan.
        all_activos = (
            Asset.objects.filter(show=True)
            .exclude(area__in=['b', 'o', 'x'])
            .select_related('supervisor', 'capitan')
            .order_by('name')
        )

        # 2) Obtener el Activo seleccionado por abbreviation
        activo = get_object_or_404(Asset, abbreviation=abbreviation)

        # 3) Tomar todos los sistemas del Activo
        sistemas = activo.system_set.all()

        # 4) Listar equipos de dichos sistemas
        equipos = Equipo.objects.filter(system__in=sistemas).order_by('name')

        # 5) Suministros ligados a este Activo
        suministros = Suministro.objects.filter(asset=activo).select_related('item')


        # 6) Generar un QR para cada equipo en base64
        domain = "https://got.serport.co"
        # equipos_con_qr = []
        for eq in equipos:
            public_url = f"{domain}/inv/public/equipo/{eq.code}/"

            # detail_url = f"/got/systems/{eq.system.pk}/{eq.code}/"

            # Crear el QRCode
            # qr = qrcode.QRCode(version=1, box_size=4, border=2)
            # qr.add_data(detail_url)
            # qr.make(fit=True)

            qr = qrcode.QRCode(version=1, box_size=4, border=2)
            qr.add_data(public_url)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")

            # Convertir la imagen a base64
            qr_io = BytesIO()
            img.save(qr_io, format='PNG')
            eq.qr_code_b64 = base64.b64encode(qr_io.getvalue()).decode('utf-8')

        all_items = Item.objects.all().order_by('name')
        context = {
            'activo': activo,
            'all_activos': all_activos,   # Para el scroll horizontal
            'equipos': equipos,           # Para la tabla de equipos
            # 'equipos_qr': equipos_con_qr,
            'suministros': suministros,   # Para la tabla de suministros
            'all_items': all_items,
        }

        context['fecha_actual'] = timezone.now().date()

        return render(request, self.template_name, context)
    
    

from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def public_equipo_detail(request, eq_code):
    """
    Vista pública (sin login) con la info completa de un equipo.
    Sin barra de navegación ni estilos de la app "base".
    """
    equipo = get_object_or_404(Equipo, code=eq_code)

    # A qué sistema y activo pertenece:
    system = equipo.system  # model: System
    asset = system.asset    # model: Asset

    images = equipo.images.all().order_by('id')
    # Podrías traer otras relaciones si deseas:
    # daily_fuel = equipo.fuel_consumptions.all() ...
    # rutas = Ruta.objects.filter(equipo=equipo) ...
    # etc.

    context = {
        'equipo': equipo,
        'system': system,
        'asset': asset,
        'images': images,
        # 'rutas': rutas, 'daily_fuel': daily_fuel, etc.
    }
    return render(request, 'inventory_management/public_equipo_detail.html', context)


import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from got.utils import pro_export_to_excel 

def export_equipment_supplies(request, abbreviation):
    """
    Genera un archivo Excel con dos secciones:
    1) Lista de Equipos del Activo
    2) Lista de Suministros del Activo
    usando la función pro_export_to_excel (adaptándola para 2 sheets).
    """
    # 1. Obtener el Activo
    asset = get_object_or_404(Asset, abbreviation=abbreviation)

    # 2. Recolectar datos de Equipos
    #    Filtra todos los equipos de los sistemas de este activo
    systems = asset.system_set.all()
    equipos_qs = Equipo.objects.filter(system__in=systems).order_by('name')

    # Cabeceras (columnas) para "Equipos"
    equipment_headers = [
        'Código', 'Nombre', 'Tipo', 'Modelo', 'Marca', 'Serial',
        'Fabricante', 'Ubicación', 'Sistema', 'Activo',
        'Horómetro/Km', 'Promedio horas/día', 'Volumen', 'Recomendaciones'
        # Agrega las columnas que desees
    ]

    # Construimos las filas
    equipment_data = []
    for eq in equipos_qs:
        system_name = eq.system.name
        asset_name = eq.system.asset.name
        # Para Horómetro/kilometraje
        if eq.system.asset.area == 'v':
            # Vehículo => interpretamos eq.horometro como km
            horo_value = f"{eq.horometro} km"
        else:
            horo_value = f"{eq.horometro} h"

        row = [
            eq.code,
            eq.name,
            eq.get_tipo_display(),
            eq.model or "",
            eq.marca or "",
            eq.serial or "",
            eq.fabricante or "",
            eq.ubicacion or system_name,  # default eq.system.location, ajusta según gustes
            system_name,
            asset_name,
            horo_value,
            str(eq.prom_hours or 0),
            str(eq.volumen or ""),
            (eq.recomendaciones or "").replace('\n', ' '),
        ]
        equipment_data.append(row)

    # 3. Recolectar datos de Suministros
    suministros_qs = Suministro.objects.filter(asset=asset).select_related('item')
    # Cabeceras para "Suministros"
    supply_headers = [
        'Artículo', 'Referencia', 'Presentación', 'Cantidad'
        # añade más si tu Suministro/item tiene más info
    ]

    supply_data = []
    for s in suministros_qs:
        if s.item:
            supply_data.append([
                s.item.name,
                s.item.reference or "",
                s.item.presentacion or "",
                str(s.cantidad)
            ])
        else:
            # Suministro sin item
            supply_data.append([
                "(Sin artículo)",
                "",
                "",
                str(s.cantidad)
            ])

    # 4. Crear un Workbook con 2 hojas:
    #    Notaremos que tu pro_export_to_excel, tal como está, genera un HttpResponse con 1 sheet.
    #    Para hacer 2 sheets, hay 2 vías:
    #      A) Llamar 2 veces => no se puede, pues retorna la response
    #      B) Reescribir algo interno
    #      C) “Hack” => generar 2 Excel independientes no es lo que quieres...
    #    Lo más adecuado: te muestro un EJEMPLO de la lógica interna “similar” a pro_export_to_excel
    #    pero adaptada a 2 sheets. Te muestro cómo:
    
    import io
    from django.http import HttpResponse
    from openpyxl import Workbook

    wb = Workbook()
    ws_equips = wb.active
    ws_equips.title = "Equipos"

    # Título en la hoja
    current_row = 1
    # Encabezados
    for col_num, header in enumerate(equipment_headers, 1):
        cell = ws_equips.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Datos
    for row_data in equipment_data:
        current_row += 1
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws_equips.cell(row=current_row, column=col_num)
            cell.value = cell_value

    # Ajuste de anchos
    for idx, column_cells in enumerate(ws_equips.columns, 1):
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        col_letter = get_column_letter(idx)
        ws_equips.column_dimensions[col_letter].width = length + 2

    # 2DA HOJA => Suministros
    ws_supp = wb.create_sheet(title="Suministros")

    # Encabezados
    current_row = 1
    for col_num, header in enumerate(supply_headers, 1):
        cell = ws_supp.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Datos
    for row_data in supply_data:
        current_row += 1
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws_supp.cell(row=current_row, column=col_num)
            cell.value = cell_value

    for idx, column_cells in enumerate(ws_supp.columns, 1):
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        col_letter = get_column_letter(idx)
        ws_supp.column_dimensions[col_letter].width = length + 2

    # 5. Exportar a HttpResponse
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{asset.abbreviation}_Equipos_y_Suministros.xlsx"
    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename={filename}'
    return response

from django.shortcuts import redirect
from django.contrib import messages

def create_supply_view(request, abbreviation):
    """
    Crea un nuevo Suministro asociado a un Activo, según
    el item seleccionado en el modal y la cantidad ingresada.
    """
    asset = get_object_or_404(Asset, abbreviation=abbreviation)

    if request.method == 'POST':
        item_id = request.POST.get('item_id')
        cantidad_str = request.POST.get('cantidad')

        if not item_id or not cantidad_str:
            messages.error(request, "Debe seleccionar un artículo y especificar la cantidad.")
            return redirect('inventory_management:asset_equipment_list', abbreviation=abbreviation)

        try:
            cantidad = float(cantidad_str)
        except ValueError:
            messages.error(request, "Cantidad inválida.")
            return redirect('inventory_management:asset_equipment_list', abbreviation=abbreviation)

        # Verificar que el item exista
        try:
            item = Item.objects.get(id=item_id)
        except Item.DoesNotExist:
            messages.error(request, "No se encontró el artículo seleccionado.")
            return redirect('inventory_management:asset_equipment_list', abbreviation=abbreviation)

        # Crear el nuevo Suministro
        Suministro.objects.create(
            item=item,
            cantidad=cantidad,
            asset=asset  # Asignado al Activo
        )
        messages.success(request, f"Se ha creado un nuevo suministro para {asset.name}.")
        return redirect('inventory_management:asset_equipment_list', abbreviation=abbreviation)

    # Si no es POST => redirigir sin hacer nada
    return redirect('inventory_management:asset_equipment_list', abbreviation=abbreviation)


class AllAssetsEquipmentListView(View):
    """
    Muestra una vista muy similar a ActivoEquipmentListView,
    pero incluyendo TODOS los activos (show=True),
    y en la tabla de equipos y suministros se incluye una columna adicional con el Activo.
    """
    template_name = 'inventory_management/all_equipment_list.html'

    def get(self, request):
        # 1) Lista de TODOS los Activos (show=True), sin excluir nada:
        all_activos = (
            Asset.objects.filter(show=True)
            .select_related('supervisor', 'capitan')
            .order_by('name')
        )

        # 2) Obtener todos los sistemas de esos activos
        all_systems = System.objects.filter(asset__in=all_activos)

        # 3) Todos los equipos de esos sistemas
        equipos = Equipo.objects.filter(system__in=all_systems).order_by('name')

        # 4) Todos los suministros asociados a esos activos
        suministros = Suministro.objects.filter(asset__in=all_activos).select_related('item')

        # 5) Generar un QR para cada equipo (opcional, si lo deseas)
        domain = "https://got.serport.co"
        for eq in equipos:
            # Por ejemplo, generamos un link público:
            public_url = f"{domain}/inv/public/equipo/{eq.code}/"
            qr = qrcode.QRCode(version=1, box_size=4, border=2)
            qr.add_data(public_url)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            qr_io = BytesIO()
            img.save(qr_io, format='PNG')
            eq.qr_code_b64 = base64.b64encode(qr_io.getvalue()).decode('utf-8')

        # 6) Si deseas, obtén también la lista de Item para "crear nuevo suministro"
        all_items = Item.objects.all().order_by('name')

        context = {
            'all_activos': all_activos,  # Para scroll horizontal, si quieres
            'equipos': equipos,
            'suministros': suministros,
            'all_items': all_items,
            # Fecha actual, etc.
            'fecha_actual': timezone.now().date(),
        }
        return render(request, self.template_name, context)


class DarBajaCreateView(LoginRequiredMixin, CreateView):
    model = DarBaja
    form_class = DarBajaForm
    template_name = 'inventory_management/dar_baja_form.html'

    def get_equipo(self):
        equipo_code = self.kwargs.get('equipo_code')
        equipo = get_object_or_404(Equipo, code=equipo_code)
        return equipo

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        equipo = self.get_equipo()
        context['equipo'] = equipo
        return context

    def form_valid(self, form):
        equipo = self.get_equipo()
        user = self.request.user
        form.instance.equipo = equipo
        form.instance.reporter = user.get_full_name() if user.get_full_name() else user.username
        form.instance.activo = f"{equipo.system.asset.name} - {equipo.system.name}"
        response = super().form_valid(form)

        # Manejar las firmas
        firma_responsable_data = self.request.POST.get('firma_responsable_data')
        firma_autorizado_data = self.request.POST.get('firma_autorizado_data')

        if firma_responsable_data:
            format, imgstr = firma_responsable_data.split(';base64,')
            ext = format.split('/')[-1]
            data = ContentFile(base64.b64decode(imgstr))
            self.object.firma_responsable.save(f'firma_responsable_{self.object.pk}.{ext}', data, save=True)

        if firma_autorizado_data:
            format, imgstr = firma_autorizado_data.split(';base64,')
            ext = format.split('/')[-1]
            data = ContentFile(base64.b64decode(imgstr))
            self.object.firma_autorizado.save(f'firma_autorizado_{self.object.pk}.{ext}', data, save=True)

        # Eliminar rutinas asociadas al equipo
        routines = equipo.equipos.all()
        if routines.exists():
            routines.delete()
            messages.warning(self.request, 'Las rutinas asociadas al equipo han sido eliminadas.')

        # Eliminar registros de horas (HistoryHour)
        history_hours = equipo.hours.all()
        if history_hours.exists():
            history_hours.delete()
            messages.warning(self.request, 'Los registros de horas asociados al equipo han sido eliminados.')

        # Eliminar registros de consumo diario de combustible (DailyFuelConsumption)
        daily_fuel_consumption = equipo.fuel_consumptions.all()
        if daily_fuel_consumption.exists():
            daily_fuel_consumption.delete()
            messages.warning(self.request, 'Los registros de consumo diario de combustible han sido eliminados.')

        # Eliminar registros de Megger
        megger_records = equipo.megger_set.all()
        if megger_records.exists():
            megger_records.delete()
            messages.warning(self.request, 'Los registros de Megger asociados al equipo han sido eliminados.')

        # Eliminar registros preoperacionales
        preoperational_records = equipo.preoperacional_set.all()
        if preoperational_records.exists():
            preoperational_records.delete()
            messages.warning(self.request, 'Los registros preoperacionales asociados al equipo han sido eliminados.')

        # Eliminar suministros (Suministro)
        supplies = equipo.suministros.all()
        if supplies.exists():
            supplies.delete()
            messages.warning(self.request, 'Los suministros asociados al equipo han sido eliminados.')

        # Mover el equipo al sistema con id 445
        new_system = System.objects.get(id=445)
        old_system = equipo.system
        equipo.system = new_system
        equipo.save()
        messages.success(self.request, f'El equipo ha sido trasladado al sistema {new_system.name}.')

        # Redirigir al usuario a la vista del sistema original
        return response
    
    def get_success_url(self):
        old_system = self.get_equipo().system
        return reverse('got:sys-detail', kwargs={'pk': old_system.id})
